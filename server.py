import os, sys, io, random, smtplib, threading, time, ctypes, ctypes.wintypes, tempfile

if hasattr(sys.stdout, 'encoding') and sys.stdout.encoding.lower() != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

from datetime import timedelta, datetime, date
from email.mime.text import MIMEText
from flask import Flask, request, jsonify, render_template
from flask_sqlalchemy import SQLAlchemy
from flask_cors import CORS
from flask_jwt_extended import JWTManager, create_access_token, jwt_required, get_jwt_identity
from werkzeug.security import generate_password_hash, check_password_hash
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from sqlalchemy import text, inspect

# ══════════════════════════════════════════════════════════════════════
#  SAFE DATA FOLDER
# ══════════════════════════════════════════════════════════════════════
DATA_FOLDER  = os.path.join(os.path.expanduser("~"), "SanjanaData")
os.makedirs(DATA_FOLDER, exist_ok=True)
DB_PATH      = os.path.join(DATA_FOLDER, "sanjana.db")
BACKUP_PATH  = os.path.join(DATA_FOLDER, "backup.xlsx")
LOG_PATH     = os.path.join(DATA_FOLDER, "backup_log.txt")

if getattr(sys, 'frozen', False):
    _BUNDLE = sys._MEIPASS
    _APP    = os.path.dirname(sys.executable)
else:
    _BUNDLE = os.path.abspath(os.path.dirname(__file__))
    _APP    = _BUNDLE

TEMPLATE_FOLDER = os.environ.get('FLASK_TEMPLATE_FOLDER', os.path.join(_BUNDLE, 'templates'))
STATIC_FOLDER   = os.environ.get('FLASK_STATIC_FOLDER',   os.path.join(_BUNDLE, 'static'))

print(f"[SANJANA] Data      : {DATA_FOLDER}")
print(f"[SANJANA] Database  : {DB_PATH}")

# ══════════════════════════════════════════════════════════════════════
#  FLASK APP
# ══════════════════════════════════════════════════════════════════════
app = Flask(__name__, template_folder=TEMPLATE_FOLDER, static_folder=STATIC_FOLDER)
CORS(app)
app.config['SQLALCHEMY_DATABASE_URI']        = 'sqlite:///' + DB_PATH
app.config['JWT_SECRET_KEY']                 = 'sanjana-pro-max-secret-2026-stable'
app.config['JWT_ACCESS_TOKEN_EXPIRES']       = timedelta(hours=72)
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Prevent browser caching so updates to HTML/JS take effect immediately
@app.after_request
def add_header(response):
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

# Ensure static files are updated
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0

limiter = Limiter(get_remote_address, app=app,
                  default_limits=["500 per day", "100 per hour"],
                  storage_uri="memory://")
db  = SQLAlchemy(app)
jwt = JWTManager(app)

ADMIN_PASSWORD    = "dhilip@25"
CLOUD_ADMIN_URL   = os.environ.get('CLOUD_ADMIN_URL', 'https://sanjana-cloud-admin.onrender.com')
temp_users        = {}
temp_password_otps = {}
temp_license_renewal_otps = {}
temp_email_change_otps = {}
active_sessions   = {}
_last_backup_date = None

def _time_ago(dt):
    """Return a human-readable 'X mins ago' string from a datetime, or 'Never'."""
    if not dt:
        return 'Never'
    diff = datetime.now() - dt
    seconds = int(diff.total_seconds())
    if seconds < 60:
        return 'Just now'
    elif seconds < 3600:
        m = seconds // 60
        return f'{m} min{"s" if m != 1 else ""} ago'
    elif seconds < 86400:
        h = seconds // 3600
        return f'{h} hr{"s" if h != 1 else ""} ago'
    else:
        d = seconds // 86400
        return f'{d} day{"s" if d != 1 else ""} ago'


# ══════════════════════════════════════════════════════════════════════
#  MODELS
# ══════════════════════════════════════════════════════════════════════
class Shop(db.Model):
    __tablename__  = 'shop'
    id            = db.Column(db.Integer,     primary_key=True)
    shop_name     = db.Column(db.String(100), nullable=False)
    owner_name    = db.Column(db.String(100), nullable=False)
    email         = db.Column(db.String(120), unique=True, nullable=False)
    phone         = db.Column(db.String(20),  nullable=False)
    address       = db.Column(db.String(255), default='')
    shop_number   = db.Column(db.String(50),  default='')
    bill_number   = db.Column(db.String(50),  default='1001')
    return_number = db.Column(db.String(50),  default='1001')
    username      = db.Column(db.String(50),  unique=True, nullable=True)
    password_hash = db.Column(db.String(256), nullable=False)
    created_at    = db.Column(db.DateTime,    default=datetime.now)
    license_start = db.Column(db.DateTime,    default=datetime.now)
    license_end   = db.Column(db.DateTime,    default=lambda: datetime.now().replace(year=datetime.now().year+1))
    approved      = db.Column(db.Boolean,     default=False)
    is_stopped    = db.Column(db.Boolean,     default=False)
    gst_number    = db.Column(db.String(50),  default='')
    last_online   = db.Column(db.DateTime,    nullable=True)

class Doctor(db.Model):
    __tablename__ = 'doctor'
    id      = db.Column(db.Integer,     primary_key=True)
    shop_id = db.Column(db.Integer,     db.ForeignKey('shop.id'), nullable=False)
    name    = db.Column(db.String(100), nullable=False)
    phone   = db.Column(db.String(20),  default='')
    spec    = db.Column(db.String(100), default='')

# ── Patient — saved like Doctor list ──────────────────────────
class Patient(db.Model):
    __tablename__ = 'patient'
    id      = db.Column(db.Integer,     primary_key=True)
    shop_id = db.Column(db.Integer,     db.ForeignKey('shop.id'), nullable=False)
    name    = db.Column(db.String(100), nullable=False)
    phone   = db.Column(db.String(20),  default='')
    address = db.Column(db.String(255), default='')

# ── Supplier — saved like Doctor list ─────────────────────────
class Supplier(db.Model):
    __tablename__ = 'supplier'
    id          = db.Column(db.Integer,     primary_key=True)
    shop_id     = db.Column(db.Integer,     db.ForeignKey('shop.id'), nullable=False)
    name        = db.Column(db.String(100), nullable=False)
    phone       = db.Column(db.String(20),  default='')
    address     = db.Column(db.String(255), default='')
    gst_number  = db.Column(db.String(50),  default='')
    company     = db.Column(db.String(100), default='')

# ── Purchase Entry — like the screen in the photo ─────────────
class PurchaseEntry(db.Model):
    __tablename__   = 'purchase_entry'
    id              = db.Column(db.Integer,     primary_key=True)
    shop_id         = db.Column(db.Integer,     db.ForeignKey('shop.id'), nullable=False)
    entry_number    = db.Column(db.String(20),  default='')
    supplier_name   = db.Column(db.String(100), default='')
    party_number    = db.Column(db.String(50),  default='')
    entry_date      = db.Column(db.String(20),  default='')
    entry_type      = db.Column(db.String(30),  default='Purchase')
    value_of_goods  = db.Column(db.Float,       default=0)
    discount        = db.Column(db.Float,       default=0)
    gst             = db.Column(db.Float,       default=0)
    net_amount      = db.Column(db.Float,       default=0)
    items_json      = db.Column(db.Text,        default='[]')
    created_at      = db.Column(db.DateTime,    default=datetime.now)
    supplier_gstin  = db.Column(db.String(50),  default='')
    place_of_supply = db.Column(db.String(100), default='')
    hsn_code        = db.Column(db.String(100), default='')
    gst_rate        = db.Column(db.Float,       default=0.0)
    taxable_amount  = db.Column(db.Float,       default=0.0)
    cgst_rate       = db.Column(db.Float,       default=0.0)
    cgst_amount     = db.Column(db.Float,       default=0.0)
    sgst_rate       = db.Column(db.Float,       default=0.0)
    sgst_amount     = db.Column(db.Float,       default=0.0)
    igst_rate       = db.Column(db.Float,       default=0.0)
    igst_amount     = db.Column(db.Float,       default=0.0)
    total_gst       = db.Column(db.Float,       default=0.0)
    grand_total     = db.Column(db.Float,       default=0.0)
    financial_year_id = db.Column(db.Integer,   db.ForeignKey('financial_year.id'), nullable=True)

class Medicine(db.Model):
    __tablename__  = 'medicine'
    id            = db.Column(db.Integer,     primary_key=True)
    shop_id       = db.Column(db.Integer,     db.ForeignKey('shop.id'), nullable=False)
    name          = db.Column(db.String(100), nullable=False)
    category      = db.Column(db.String(50),  default='General')
    batch         = db.Column(db.String(50),  default='')
    quantity      = db.Column(db.Float,       default=0.0)
    price         = db.Column(db.Float,       nullable=False)
    mrp           = db.Column(db.Float,       default=0.0)
    gst           = db.Column(db.Float,       default=0.0)
    expiry_date   = db.Column(db.String(20),  default='')
    supplier_name = db.Column(db.String(100), default='')
    company_name  = db.Column(db.String(100), default='')
    pack_size     = db.Column(db.String(50),  default='10')
    # Discount from purchase — auto-applied in billing
    sale_discount = db.Column(db.Float,       default=0.0)

class Bill(db.Model):
    __tablename__  = 'bill'
    id            = db.Column(db.Integer,     primary_key=True)
    shop_id       = db.Column(db.Integer,     db.ForeignKey('shop.id'), nullable=False)
    bill_number   = db.Column(db.String(20),  default='')
    customer_name = db.Column(db.String(100), default='Walk-in')
    customer_phone= db.Column(db.String(20),  default='')
    doctor_name   = db.Column(db.String(100), default='')
    subtotal      = db.Column(db.Float,       default=0)
    cgst          = db.Column(db.Float,       default=0)
    sgst          = db.Column(db.Float,       default=0)
    discount      = db.Column(db.Float,       default=0)
    total_amount  = db.Column(db.Float,       default=0)
    bill_date     = db.Column(db.DateTime,    default=datetime.now)
    # ── FEATURE 2: Custom date ─────────────────────────────────
    custom_date   = db.Column(db.String(20),  default='')
    items_json    = db.Column(db.Text,        default='[]')
    status        = db.Column(db.String(20),  default='active')
    returned_amount=db.Column(db.Float,       default=0.0)
    customer_gstin  = db.Column(db.String(50),  default='')
    place_of_supply = db.Column(db.String(100), default='')
    hsn_code        = db.Column(db.String(100), default='')
    gst_rate        = db.Column(db.Float,       default=0.0)
    taxable_amount  = db.Column(db.Float,       default=0.0)
    cgst_rate       = db.Column(db.Float,       default=0.0)
    cgst_amount     = db.Column(db.Float,       default=0.0)
    sgst_rate       = db.Column(db.Float,       default=0.0)
    sgst_amount     = db.Column(db.Float,       default=0.0)
    igst_rate       = db.Column(db.Float,       default=0.0)
    igst_amount     = db.Column(db.Float,       default=0.0)
    total_gst       = db.Column(db.Float,       default=0.0)
    grand_total     = db.Column(db.Float,       default=0.0)
    financial_year_id = db.Column(db.Integer,   db.ForeignKey('financial_year.id'), nullable=True)

class BillReturn(db.Model):
    __tablename__  = 'bill_return'
    id            = db.Column(db.Integer,     primary_key=True)
    shop_id       = db.Column(db.Integer,     db.ForeignKey('shop.id'), nullable=False)
    return_number = db.Column(db.String(20),  default='')
    bill_id       = db.Column(db.Integer,     db.ForeignKey('bill.id'), nullable=False)
    bill_number   = db.Column(db.String(20),  default='')
    customer_name = db.Column(db.String(100), default='Walk-in')
    customer_phone= db.Column(db.String(20),  default='')
    return_date   = db.Column(db.DateTime,    default=datetime.now)
    items_json    = db.Column(db.Text,        default='[]')
    subtotal      = db.Column(db.Float,       default=0.0)
    tax_amount    = db.Column(db.Float,       default=0.0)
    refund_amount = db.Column(db.Float,       default=0.0)
    reason        = db.Column(db.String(255), default='')
    notes         = db.Column(db.Text,        default='')

class LicenseRenewalRequest(db.Model):
    __tablename__ = 'license_renewal_request'
    id           = db.Column(db.Integer,     primary_key=True)
    shop_id      = db.Column(db.Integer,     db.ForeignKey('shop.id'), nullable=False)
    requested_at = db.Column(db.DateTime,    default=datetime.now)
    status       = db.Column(db.String(20),  default='pending')  # 'pending', 'approved', 'rejected'
    processed_at = db.Column(db.DateTime,    nullable=True)

class FinancialYear(db.Model):
    __tablename__ = 'financial_year'
    id            = db.Column(db.Integer,     primary_key=True)
    fy_name       = db.Column(db.String(50),  nullable=False)
    start_date    = db.Column(db.String(20),  nullable=False)
    end_date      = db.Column(db.String(20),  nullable=False)
    is_active     = db.Column(db.Boolean,     default=False)

class GSTReturnStatus(db.Model):
    __tablename__ = 'gst_return_status'
    id            = db.Column(db.Integer,     primary_key=True)
    shop_id       = db.Column(db.Integer,     db.ForeignKey('shop.id'), nullable=False)
    financial_year_id = db.Column(db.Integer, db.ForeignKey('financial_year.id'), nullable=False)
    month_val     = db.Column(db.String(7),   nullable=False) # e.g. "2026-04"
    status        = db.Column(db.String(20),  default='Pending') # 'Filed', 'Pending'
    filed_date    = db.Column(db.String(20),  default='')

def send_email_notification(to_email, subject, text_body, html_body):
    SENDER_EMAIL    = 'sanjanasoftware03@gmail.com'
    SENDER_PASSWORD = 'xqgylrjhthhnxfmd'
    try:
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText as MT
        
        msg_root = MIMEMultipart('alternative')
        msg_root['Subject'] = subject
        msg_root['From']    = f'Sanjana Software <{SENDER_EMAIL}>'
        msg_root['To']      = to_email
        msg_root.attach(MT(text_body, 'plain'))
        msg_root.attach(MT(html_body, 'html'))
        
        try:
            s = smtplib.SMTP_SSL('smtp.gmail.com', 465, timeout=15)
            s.login(SENDER_EMAIL, SENDER_PASSWORD)
            s.sendmail(SENDER_EMAIL, to_email, msg_root.as_string())
            s.quit()
            return True
        except Exception as ssl_err:
            s2 = smtplib.SMTP('smtp.gmail.com', 587, timeout=15)
            s2.ehlo(); s2.starttls(); s2.ehlo()
            s2.login(SENDER_EMAIL, SENDER_PASSWORD)
            s2.sendmail(SENDER_EMAIL, to_email, msg_root.as_string())
            s2.quit()
            return True
    except Exception as e:
        print(f"[EMAIL ERROR] Failed to send email to {to_email}: {e}")
        return False


# ══════════════════════════════════════════════════════════════════════
#  DATABASE INIT
# ══════════════════════════════════════════════════════════════════════
def setup_database():
    with app.app_context():
        db.create_all()
        ins = inspect(db.engine)
        tbl = ins.get_table_names()
        migrations = [
            ("shop",     "license_start",  "ALTER TABLE shop     ADD COLUMN license_start DATETIME"),
            ("shop",     "license_end",    "ALTER TABLE shop     ADD COLUMN license_end   DATETIME"),
            ("medicine", "batch",          "ALTER TABLE medicine ADD COLUMN batch         VARCHAR(50)  DEFAULT ''"),
            ("medicine", "mrp",            "ALTER TABLE medicine ADD COLUMN mrp           FLOAT        DEFAULT 0.0"),
            ("medicine", "expiry_date",    "ALTER TABLE medicine ADD COLUMN expiry_date   VARCHAR(20)  DEFAULT ''"),
            ("medicine", "supplier_name",  "ALTER TABLE medicine ADD COLUMN supplier_name VARCHAR(100) DEFAULT ''"),
            ("medicine", "company_name",   "ALTER TABLE medicine ADD COLUMN company_name  VARCHAR(100) DEFAULT ''"),
            ("medicine", "gst",            "ALTER TABLE medicine ADD COLUMN gst           FLOAT        DEFAULT 0.0"),
            ("medicine", "pack_size",      "ALTER TABLE medicine ADD COLUMN pack_size     VARCHAR(50)  DEFAULT '10'"),
            ("medicine", "sale_discount",  "ALTER TABLE medicine ADD COLUMN sale_discount FLOAT        DEFAULT 0.0"),
            ("bill",     "bill_number",    "ALTER TABLE bill     ADD COLUMN bill_number   VARCHAR(20)  DEFAULT ''"),
            ("bill",     "customer_name",  "ALTER TABLE bill     ADD COLUMN customer_name VARCHAR(100) DEFAULT 'Walk-in'"),
            ("bill",     "customer_phone", "ALTER TABLE bill     ADD COLUMN customer_phone VARCHAR(20) DEFAULT ''"),
            ("bill",     "doctor_name",    "ALTER TABLE bill     ADD COLUMN doctor_name   VARCHAR(100) DEFAULT ''"),
            ("bill",     "subtotal",       "ALTER TABLE bill     ADD COLUMN subtotal      FLOAT        DEFAULT 0"),
            ("bill",     "cgst",           "ALTER TABLE bill     ADD COLUMN cgst          FLOAT        DEFAULT 0"),
            ("bill",     "sgst",           "ALTER TABLE bill     ADD COLUMN sgst          FLOAT        DEFAULT 0"),
            ("bill",     "discount",       "ALTER TABLE bill     ADD COLUMN discount      FLOAT        DEFAULT 0"),
            ("bill",     "items_json",     "ALTER TABLE bill     ADD COLUMN items_json    TEXT         DEFAULT '[]'"),
            ("bill",     "custom_date",    "ALTER TABLE bill     ADD COLUMN custom_date   VARCHAR(20)  DEFAULT ''"),
            ("shop",     "approved",       "ALTER TABLE shop     ADD COLUMN approved      BOOLEAN      DEFAULT 1"),
            ("shop",     "username",       "ALTER TABLE shop     ADD COLUMN username      VARCHAR(50)"),
            ("shop",     "is_stopped",     "ALTER TABLE shop     ADD COLUMN is_stopped    BOOLEAN      DEFAULT 0"),
            ("shop",     "return_number",  "ALTER TABLE shop     ADD COLUMN return_number VARCHAR(50) DEFAULT '1001'"),
            ("bill",     "status",         "ALTER TABLE bill     ADD COLUMN status        VARCHAR(20) DEFAULT 'active'"),
            ("bill",     "returned_amount", "ALTER TABLE bill    ADD COLUMN returned_amount FLOAT      DEFAULT 0.0"),
            ("shop",     "gst_number",      "ALTER TABLE shop     ADD COLUMN gst_number VARCHAR(50) DEFAULT ''"),
            ("shop",     "last_online",      "ALTER TABLE shop     ADD COLUMN last_online DATETIME"),
            ("bill",     "customer_gstin",  "ALTER TABLE bill     ADD COLUMN customer_gstin VARCHAR(50) DEFAULT ''"),
            ("bill",     "place_of_supply", "ALTER TABLE bill     ADD COLUMN place_of_supply VARCHAR(100) DEFAULT ''"),
            ("bill",     "hsn_code",        "ALTER TABLE bill     ADD COLUMN hsn_code VARCHAR(100) DEFAULT ''"),
            ("bill",     "gst_rate",        "ALTER TABLE bill     ADD COLUMN gst_rate FLOAT DEFAULT 0.0"),
            ("bill",     "taxable_amount",  "ALTER TABLE bill     ADD COLUMN taxable_amount FLOAT DEFAULT 0.0"),
            ("bill",     "cgst_rate",       "ALTER TABLE bill     ADD COLUMN cgst_rate FLOAT DEFAULT 0.0"),
            ("bill",     "cgst_amount",     "ALTER TABLE bill     ADD COLUMN cgst_amount FLOAT DEFAULT 0.0"),
            ("bill",     "sgst_rate",       "ALTER TABLE bill     ADD COLUMN sgst_rate FLOAT DEFAULT 0.0"),
            ("bill",     "sgst_amount",     "ALTER TABLE bill     ADD COLUMN sgst_amount FLOAT DEFAULT 0.0"),
            ("bill",     "igst_rate",       "ALTER TABLE bill     ADD COLUMN igst_rate FLOAT DEFAULT 0.0"),
            ("bill",     "igst_amount",     "ALTER TABLE bill     ADD COLUMN igst_amount FLOAT DEFAULT 0.0"),
            ("bill",     "total_gst",       "ALTER TABLE bill     ADD COLUMN total_gst FLOAT DEFAULT 0.0"),
            ("bill",     "grand_total",     "ALTER TABLE bill     ADD COLUMN grand_total FLOAT DEFAULT 0.0"),
            ("bill",     "financial_year_id", "ALTER TABLE bill   ADD COLUMN financial_year_id INTEGER DEFAULT NULL"),
            ("purchase_entry", "supplier_gstin",  "ALTER TABLE purchase_entry ADD COLUMN supplier_gstin VARCHAR(50) DEFAULT ''"),
            ("purchase_entry", "place_of_supply", "ALTER TABLE purchase_entry ADD COLUMN place_of_supply VARCHAR(100) DEFAULT ''"),
            ("purchase_entry", "hsn_code",        "ALTER TABLE purchase_entry ADD COLUMN hsn_code VARCHAR(100) DEFAULT ''"),
            ("purchase_entry", "gst_rate",        "ALTER TABLE purchase_entry ADD COLUMN gst_rate FLOAT DEFAULT 0.0"),
            ("purchase_entry", "taxable_amount",  "ALTER TABLE purchase_entry ADD COLUMN taxable_amount FLOAT DEFAULT 0.0"),
            ("purchase_entry", "cgst_rate",       "ALTER TABLE purchase_entry ADD COLUMN cgst_rate FLOAT DEFAULT 0.0"),
            ("purchase_entry", "cgst_amount",     "ALTER TABLE purchase_entry ADD COLUMN cgst_amount FLOAT DEFAULT 0.0"),
            ("purchase_entry", "sgst_rate",       "ALTER TABLE purchase_entry ADD COLUMN sgst_rate FLOAT DEFAULT 0.0"),
            ("purchase_entry", "sgst_amount",     "ALTER TABLE purchase_entry ADD COLUMN sgst_amount FLOAT DEFAULT 0.0"),
            ("purchase_entry", "igst_rate",       "ALTER TABLE purchase_entry ADD COLUMN igst_rate FLOAT DEFAULT 0.0"),
            ("purchase_entry", "igst_amount",     "ALTER TABLE purchase_entry ADD COLUMN igst_amount FLOAT DEFAULT 0.0"),
            ("purchase_entry", "total_gst",       "ALTER TABLE purchase_entry ADD COLUMN total_gst FLOAT DEFAULT 0.0"),
            ("purchase_entry", "grand_total",     "ALTER TABLE purchase_entry ADD COLUMN grand_total FLOAT DEFAULT 0.0"),
            ("purchase_entry", "financial_year_id", "ALTER TABLE purchase_entry ADD COLUMN financial_year_id INTEGER DEFAULT NULL"),
        ]
        for table, col, sql in migrations:
            try:
                existing = [c['name'] for c in ins.get_columns(table)] if table in tbl else []
                if col not in existing:
                    db.session.execute(text(sql))
                    db.session.commit()
            except:
                db.session.rollback()

        try:
            missing_custom = Bill.query.filter(db.or_(Bill.custom_date == None, Bill.custom_date == '')).all()
            for b in missing_custom:
                if b.bill_date:
                    b.custom_date = b.bill_date.strftime('%Y-%m-%d')
            if missing_custom:
                db.session.commit()
        except:
            db.session.rollback()

        # Ensure we have default Financial Years and link existing transactions
        try:
            tbl_check = inspect(db.engine).get_table_names()
            if 'financial_year' in tbl_check:
                count = db.session.query(db.func.count(FinancialYear.id)).scalar()
                if count == 0:
                    fy25 = FinancialYear(fy_name="FY 2025-26", start_date="2025-04-01", end_date="2026-03-31", is_active=False)
                    fy26 = FinancialYear(fy_name="FY 2026-27", start_date="2026-04-01", end_date="2027-03-31", is_active=True)
                    db.session.add(fy25)
                    db.session.add(fy26)
                    db.session.commit()

                fys = FinancialYear.query.all()
                def get_fy_id(date_str_or_obj):
                    if not date_str_or_obj:
                        active = next((f for f in fys if f.is_active), None)
                        return active.id if active else None
                    if isinstance(date_str_or_obj, datetime):
                        dt = date_str_or_obj.date()
                    else:
                        try:
                            dt = datetime.strptime(str(date_str_or_obj)[:10], "%Y-%m-%d").date()
                        except:
                            active = next((f for f in fys if f.is_active), None)
                            return active.id if active else None
                    for f in fys:
                        f_start = datetime.strptime(f.start_date, "%Y-%m-%d").date()
                        f_end = datetime.strptime(f.end_date, "%Y-%m-%d").date()
                        if f_start <= dt <= f_end:
                            return f.id
                    active = next((f for f in fys if f.is_active), None)
                    return active.id if active else None

                bills_to_update = Bill.query.filter(Bill.financial_year_id == None).all()
                for b in bills_to_update:
                    date_val = b.custom_date or b.bill_date
                    b.financial_year_id = get_fy_id(date_val)

                purchases_to_update = PurchaseEntry.query.filter(PurchaseEntry.financial_year_id == None).all()
                for p in purchases_to_update:
                    p.financial_year_id = get_fy_id(p.entry_date)

                if bills_to_update or purchases_to_update:
                    db.session.commit()
        except Exception as e:
            print(f"[SANJANA FY MIGRATION ERROR] {e}")
            db.session.rollback()

        print("[SANJANA] ✅ Database ready!")

setup_database()

@app.before_request
def guard():
    if 'shop' not in inspect(db.engine).get_table_names():
        setup_database()
    daily_backup_check()

# ══════════════════════════════════════════════════════════════════════
#  BACKUP
# ══════════════════════════════════════════════════════════════════════
def write_log(msg):
    try:
        with open(LOG_PATH, 'a', encoding='utf-8') as f:
            f.write(f"[{datetime.now():%Y-%m-%d %H:%M:%S}] {msg}\n")
    except: pass

def create_excel_backup():
    try:
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        with app.app_context():
            meds = db.session.execute(text(
                "SELECT s.shop_name,m.name,m.category,m.batch,m.company_name,"
                "m.supplier_name,m.price,m.mrp,m.gst,m.quantity,m.expiry_date "
                "FROM medicine m JOIN shop s ON s.id=m.shop_id ORDER BY s.shop_name,m.name"
            )).fetchall()
            df_meds = pd.DataFrame(meds, columns=["Shop","Name","Category","Batch",
                "Company","Supplier","Price","MRP","GST%","Qty","Expiry"])
            bills = db.session.execute(text(
                "SELECT s.shop_name,b.bill_number,b.customer_name,b.doctor_name,"
                "b.subtotal,b.cgst,b.sgst,b.discount,b.total_amount,"
                "strftime('%Y-%m-%d %H:%M',b.bill_date) "
                "FROM bill b JOIN shop s ON s.id=b.shop_id ORDER BY b.bill_date DESC"
            )).fetchall()
            df_bills = pd.DataFrame(bills, columns=["Shop","Bill No","Customer","Doctor",
                "Subtotal","CGST","SGST","Discount","Total","Date"])
            shops = db.session.execute(text(
                "SELECT shop_name,owner_name,email,phone,address,bill_number,"
                "strftime('%Y-%m-%d',created_at) FROM shop"
            )).fetchall()
            df_shops = pd.DataFrame(shops, columns=["Shop","Owner","Email","Phone",
                "Address","Next Bill No","Joined"])
        with pd.ExcelWriter(BACKUP_PATH, engine='openpyxl') as w:
            df_meds.to_excel(w,  sheet_name='Medicines', index=False)
            df_bills.to_excel(w, sheet_name='Bills',     index=False)
            df_shops.to_excel(w, sheet_name='Shops',     index=False)
        wb  = load_workbook(BACKUP_PATH)
        clr = {'Medicines':'1A237E','Bills':'1B5E20','Shops':'880E4F'}
        t   = Border(left=Side(style='thin',color='CCCCCC'),right=Side(style='thin',color='CCCCCC'),
                     top=Side(style='thin',color='CCCCCC'),bottom=Side(style='thin',color='CCCCCC'))
        for sn, c in clr.items():
            if sn not in wb.sheetnames: continue
            ws = wb[sn]
            for cell in ws[1]:
                cell.font=Font(bold=True,color='FFFFFF',size=11,name='Arial')
                cell.fill=PatternFill('solid',start_color=c)
                cell.alignment=Alignment(horizontal='center',vertical='center')
                cell.border=t
            ws.row_dimensions[1].height=26
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.font=Font(name='Arial',size=10); cell.border=t
            for col in ws.columns:
                w2=max((len(str(c2.value or '')) for c2 in col),default=10)
                ws.column_dimensions[get_column_letter(col[0].column)].width=min(w2+4,35)
            ws.freeze_panes='A2'
        wb.save(BACKUP_PATH)
        msg = f"Backup OK → {BACKUP_PATH}"
        write_log(msg); print(f"[BACKUP] ✅ {msg}")
        return True, BACKUP_PATH
    except Exception as e:
        write_log(f"Backup FAILED: {e}"); return False, str(e)

def backup_bg():
    threading.Thread(target=create_excel_backup, daemon=True).start()

def daily_backup_check():
    global _last_backup_date
    today = date.today()
    if _last_backup_date != today:
        _last_backup_date = today
        backup_bg()

# ══════════════════════════════════════════════════════════════════════
#  PAGE ROUTES
# ══════════════════════════════════════════════════════════════════════
@app.route('/')
def index():      return render_template('login.html')
@app.route('/dashboard')
def dashboard():  return render_template('dashboard.html')
@app.route('/billing')
def billing():    return render_template('billing.html')
@app.route('/purchase')
def purchase():   return render_template('purchase.html')
@app.route('/owner-admin')
def admin_page(): return render_template('admin.html')

# ══════════════════════════════════════════════════════════════════════
#  AUTH
# ══════════════════════════════════════════════════════════════════════
@app.route('/api/send-otp', methods=['POST'])
@limiter.limit("10 per minute")
def send_otp():
    try:
        data  = request.json
        email = data.get('email','').strip()

        if not email:
            return jsonify({'error': 'Email is required'}), 400
        if Shop.query.filter_by(email=email).first():
            return jsonify({'error': 'Email already registered'}), 409

        otp = str(random.randint(100000, 999999))

        # ── Always store OTP first ─────────────────────────────────
        temp_users[email] = {'otp': otp, 'data': data}
        print(f"\n{'='*40}")
        print(f"  OTP FOR: {email}")
        print(f"  OTP CODE: {otp}")
        print(f"{'='*40}\n")

        # ── Try sending email ──────────────────────────────────────
        SENDER_EMAIL    = 'sanjanasoftware03@gmail.com'
        SENDER_PASSWORD = 'xqgylrjhthhnxfmd'   # Gmail app password

        email_sent = False
        err_msg    = ''
        try:
            from email.mime.multipart import MIMEMultipart
            # Build a proper HTML email
            msg_root = MIMEMultipart('alternative')
            msg_root['Subject'] = f'Your Sanjana Software OTP: {otp}'
            msg_root['From']    = f'Sanjana Software <{SENDER_EMAIL}>'
            msg_root['To']      = email

            text_body = f"Hello {data.get('owner_name','User')},\n\nYour OTP is: {otp}\n\nValid for 10 minutes.\n\nSanjana Software"
            html_body = f"""<div style="font-family:Arial,sans-serif;max-width:400px;margin:0 auto;padding:20px">
  <h2 style="color:#1a237e">Sanjana Software</h2>
  <p>Hello <b>{data.get('owner_name','User')}</b>,</p>
  <p>Your registration OTP is:</p>
  <div style="background:#f0f4ff;border:2px solid #1a237e;border-radius:10px;
    padding:20px;text-align:center;margin:16px 0">
    <span style="font-size:32px;font-weight:900;letter-spacing:8px;color:#1a237e">{otp}</span>
  </div>
  <p style="color:#888;font-size:12px">Valid for 10 minutes. Do not share this OTP.</p>
  <p style="color:#888;font-size:12px">Sanjana Software Team</p>
</div>"""
            from email.mime.text import MIMEText as MT
            msg_root.attach(MT(text_body, 'plain'))
            msg_root.attach(MT(html_body, 'html'))

            # Try SMTP_SSL first (port 465)
            try:
                s = smtplib.SMTP_SSL('smtp.gmail.com', 465, timeout=15)
                s.login(SENDER_EMAIL, SENDER_PASSWORD)
                s.sendmail(SENDER_EMAIL, email, msg_root.as_string())
                s.quit()
                email_sent = True
                print(f"[OTP] ✅ Email sent via SSL to {email}")
            except Exception as ssl_err:
                print(f"[OTP] SSL failed ({ssl_err}), trying TLS...")
                # Fallback: try STARTTLS (port 587)
                s2 = smtplib.SMTP('smtp.gmail.com', 587, timeout=15)
                s2.ehlo()
                s2.starttls()
                s2.ehlo()
                s2.login(SENDER_EMAIL, SENDER_PASSWORD)
                s2.sendmail(SENDER_EMAIL, email, msg_root.as_string())
                s2.quit()
                email_sent = True
                print(f"[OTP] ✅ Email sent via TLS to {email}")

        except Exception as mail_err:
            err_msg = str(mail_err)
            print(f"[OTP] ❌ Email failed: {err_msg}")

        if email_sent:
            return jsonify({'message': f'OTP sent to {email}! Check your inbox.'}), 200
        else:
            # Return OTP in response so UI can show it — fallback for offline use
            return jsonify({
                'message': 'OTP generated! Email could not be sent.',
                'otp_console': True,
                'otp_hint': otp,          # show directly on screen
                'error_detail': err_msg[:120]
            }), 200

    except Exception as e:
        print(f"[OTP] Fatal error: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/verify-otp', methods=['POST'])
def verify_otp():
    try:
        data  = request.json
        email = data.get('email','').strip()
        otp   = str(data.get('otp','')).strip()

        if not email or not otp:
            return jsonify({'error': 'Email and OTP required'}), 400

        if email not in temp_users:
            return jsonify({'error': 'OTP expired. Please request a new one.'}), 400

        if temp_users[email]['otp'] != otp:
            return jsonify({'error': 'Wrong OTP. Please check and try again.'}), 400

        u = temp_users[email]['data']
        now = datetime.now()
        new_shop = Shop(
            shop_name     = u.get('shop_name',''),
            owner_name    = u.get('owner_name',''),
            email         = email,
            phone         = u.get('phone',''),
            address       = u.get('address',''),
            shop_number   = u.get('shop_number',''),
            bill_number   = str(u.get('bill_number','1001')),
            password_hash = generate_password_hash(u.get('password','')),
            created_at    = now,
            license_start = now,
            license_end   = now.replace(year=now.year+1)  # exactly 1 year
        )
        db.session.add(new_shop)
        db.session.commit()
        del temp_users[email]
        backup_bg()
        print(f"[SANJANA] ✅ Registered (pending approval): {u.get('shop_name')} ({email})")

        # ── Send registration details email to admin ──────────────
        try:
            ADMIN_EMAIL     = 'sanjanasoftware03@gmail.com'
            NOTIFY_SENDER   = 'sanjanasoftware03@gmail.com'
            NOTIFY_PASSWORD = 'xqgylrjhthhnxfmd'
            join_time = now.strftime('%d-%b-%Y %I:%M %p')

            from email.mime.multipart import MIMEMultipart
            from email.mime.text import MIMEText as MT
            notify_msg = MIMEMultipart('alternative')
            notify_msg['Subject'] = f'🆕 New Registration: {u.get("shop_name","")} — Approval Required'
            notify_msg['From']    = f'Sanjana Software <{NOTIFY_SENDER}>'
            notify_msg['To']      = ADMIN_EMAIL

            notify_text = (f"New Registration Request\n\n"
                f"Shop Name: {u.get('shop_name','')}\n"
                f"Owner Name: {u.get('owner_name','')}\n"
                f"Phone: {u.get('phone','')}\n"
                f"Email: {email}\n"
                f"Address: {u.get('address','N/A')}\n"
                f"Shop License: {u.get('shop_number','N/A')}\n"
                f"Joined: {join_time}\n\n"
                f"Please login to Owner Admin Panel to approve or reject.")

            notify_html = f"""<div style="font-family:Arial,sans-serif;max-width:500px;margin:0 auto;padding:20px">
  <h2 style="color:#1a237e;border-bottom:2px solid #1a237e;padding-bottom:10px">🆕 New Registration Request</h2>
  <p style="color:#555">A new shop has registered and is waiting for your approval:</p>
  <table style="width:100%;border-collapse:collapse;margin:16px 0">
    <tr style="background:#f0f4ff"><td style="padding:10px 14px;font-weight:600;color:#555;border:1px solid #ddd">🏪 Shop Name</td><td style="padding:10px 14px;font-weight:700;color:#1a237e;border:1px solid #ddd">{u.get('shop_name','')}</td></tr>
    <tr><td style="padding:10px 14px;font-weight:600;color:#555;border:1px solid #ddd">👤 Owner Name</td><td style="padding:10px 14px;color:#333;border:1px solid #ddd">{u.get('owner_name','')}</td></tr>
    <tr style="background:#f0f4ff"><td style="padding:10px 14px;font-weight:600;color:#555;border:1px solid #ddd">📱 Phone</td><td style="padding:10px 14px;color:#333;border:1px solid #ddd">{u.get('phone','')}</td></tr>
    <tr><td style="padding:10px 14px;font-weight:600;color:#555;border:1px solid #ddd">📧 Email</td><td style="padding:10px 14px;color:#1a73e8;border:1px solid #ddd">{email}</td></tr>
    <tr style="background:#f0f4ff"><td style="padding:10px 14px;font-weight:600;color:#555;border:1px solid #ddd">📍 Address</td><td style="padding:10px 14px;color:#333;border:1px solid #ddd">{u.get('address','N/A')}</td></tr>
    <tr><td style="padding:10px 14px;font-weight:600;color:#555;border:1px solid #ddd">🪪 License No</td><td style="padding:10px 14px;color:#333;border:1px solid #ddd">{u.get('shop_number','N/A')}</td></tr>
    <tr style="background:#fff3cd"><td style="padding:10px 14px;font-weight:600;color:#555;border:1px solid #ddd">🕐 Joined On</td><td style="padding:10px 14px;font-weight:700;color:#e65100;border:1px solid #ddd">{join_time}</td></tr>
  </table>
  <div style="background:#fff3cd;border:1px solid #ffc107;border-radius:8px;padding:14px;text-align:center;margin:16px 0">
    <p style="color:#856404;font-weight:700;margin:0">⚠️ Action Required</p>
    <p style="color:#856404;font-size:13px;margin:4px 0 0">Login to Owner Admin Panel to Approve or Reject this registration.</p>
  </div>
  <p style="color:#888;font-size:12px">Sanjana Software — Admin Notification</p>
</div>"""
            notify_msg.attach(MT(notify_text, 'plain'))
            notify_msg.attach(MT(notify_html, 'html'))

            try:
                s = smtplib.SMTP_SSL('smtp.gmail.com', 465, timeout=15)
                s.login(NOTIFY_SENDER, NOTIFY_PASSWORD)
                s.sendmail(NOTIFY_SENDER, ADMIN_EMAIL, notify_msg.as_string())
                s.quit()
                print(f"[SANJANA] 📧 Admin notified about new registration: {email}")
            except:
                s2 = smtplib.SMTP('smtp.gmail.com', 587, timeout=15)
                s2.ehlo(); s2.starttls(); s2.ehlo()
                s2.login(NOTIFY_SENDER, NOTIFY_PASSWORD)
                s2.sendmail(NOTIFY_SENDER, ADMIN_EMAIL, notify_msg.as_string())
                s2.quit()
                print(f"[SANJANA] 📧 Admin notified (TLS) about new registration: {email}")
        except Exception as notify_err:
            print(f"[SANJANA] ⚠️ Admin notification email failed: {notify_err}")

        return jsonify({'message': 'Registration successful! Waiting for admin approval.', 'pending_approval': True}), 201

    except Exception as e:
        db.session.rollback()
        print(f"[SANJANA] Registration error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/login', methods=['POST'])
@limiter.limit("20 per minute")
def login():
    data = request.json
    login_id = (data.get('email') or data.get('username') or data.get('login_id', '')).strip()
    if not login_id:
        return jsonify({'error': 'Email or Username is required'}), 400
    shop = Shop.query.filter(
        (db.func.lower(Shop.email) == login_id.lower()) |
        (db.func.lower(Shop.username) == login_id.lower())
    ).first()
    if not shop or not check_password_hash(shop.password_hash, data.get('password','')):
        return jsonify({'error':'Invalid Login'}), 401
    if not shop.approved:
        return jsonify({'error':'Your registration is pending admin approval. Please wait for approval.','pending_approval':True}), 403
    if getattr(shop, 'is_stopped', False):
        return jsonify({
            'error': 'Your shop software access has been temporarily stopped by Admin.\n\nPlease contact Sanjana Software Team:\nPhone: 9025422389\nEmail: sanjanasoftware03@gmail.com',
            'is_stopped': True,
            'contact_phone': '9025422389',
            'contact_email': 'sanjanasoftware03@gmail.com'
        }), 403
    token = create_access_token(identity=str(shop.id))
    now_dt = datetime.now()
    shop.last_online = now_dt
    db.session.commit()
    active_sessions[shop.id] = {
        'shop_name': shop.shop_name,
        'owner_name': shop.owner_name,
        'email': shop.email,
        'username': shop.username or '',
        'phone': shop.phone,
        'login_time': now_dt.strftime("%Y-%m-%d %H:%M:%S")
    }
    return jsonify({
        'token': token,
        'shop_name': shop.shop_name,
        'owner_name': shop.owner_name,
        'username': shop.username or '',
        'phone': shop.phone,
        'address': shop.address,
        'bill_number': shop.bill_number,
        'shop_number': shop.shop_number or ''
    })

@app.route('/api/heartbeat', methods=['POST'])
@jwt_required()
def heartbeat():
    """Called every 2 minutes from dashboard to keep last_online fresh."""
    try:
        shop = Shop.query.get(int(get_jwt_identity()))
        if shop:
            shop.last_online = datetime.now()
            db.session.commit()
        return jsonify({'ok': True})
    except Exception:
        return jsonify({'ok': False}), 500

@app.route('/api/logout', methods=['POST'])
@jwt_required()
def logout():
    active_sessions.pop(int(get_jwt_identity()),None)
    return jsonify({'message':'Logged out'})

@app.route('/api/license-status', methods=['GET'])
@jwt_required()
def license_status():
    shop  = Shop.query.get(int(get_jwt_identity()))
    start = shop.license_start or shop.created_at or datetime.now()
    end   = shop.license_end
    if not end:
        end = start.replace(year=start.year+1)
    days_left = (end - datetime.now()).days
    is_stopped = bool(getattr(shop, 'is_stopped', False))
    return jsonify({
        'days_left':     days_left,
        'expired':       days_left < 0,
        'warning':       0 <= days_left <= 10,
        'blocked':       (days_left < 0) or is_stopped,
        'is_stopped':    is_stopped,
        'contact_phone': '9025422389',
        'contact_email': 'sanjanasoftware03@gmail.com',
        'license_start': start.strftime('%Y-%m-%d'),
        'license_end':   end.strftime('%Y-%m-%d'),
        'shop_name':     shop.shop_name,
        'owner_name':    shop.owner_name
    })

@app.route('/api/admin/renew-license', methods=['POST'])
def renew_license():
    data = request.json
    if data.get('password') != ADMIN_PASSWORD:
        return jsonify({'error':'Unauthorized'}), 401
    shop = Shop.query.filter_by(email=data.get('email')).first()
    if not shop:
        return jsonify({'error':'Shop not found'}), 404
    now         = datetime.now()
    current_end = shop.license_end or now
    new_end     = current_end.replace(year=current_end.year+1) if current_end >= now \
                  else now.replace(year=now.year+1)
    shop.license_end = new_end
    db.session.commit()
    return jsonify({'message':f'License renewed for {shop.shop_name}',
                    'new_end':new_end.strftime('%Y-%m-%d')})

@app.route('/api/profile', methods=['GET','PUT'])
@jwt_required()
def manage_profile():
    shop=Shop.query.get(int(get_jwt_identity()))
    if request.method=='GET':
        # Return license info too
        end   = shop.license_end or (shop.created_at.replace(year=shop.created_at.year+1) if shop.created_at else datetime.now())
        days_left=(end-datetime.now()).days
        
        import os
        logo_path = os.path.join(STATIC_FOLDER, f'logo_{shop.id}.png')
        logo_url = f'/static/logo_{shop.id}.png' if os.path.exists(logo_path) else None
        
        latest_req = LicenseRenewalRequest.query.filter_by(shop_id=shop.id).order_by(LicenseRenewalRequest.id.desc()).first()
        renewal_status = latest_req.status if latest_req else None
        renewal_requested_at = latest_req.requested_at.strftime('%Y-%m-%d %H:%M') if (latest_req and latest_req.requested_at) else None

        return jsonify({
            'shop_name':            shop.shop_name,
            'owner_name':           shop.owner_name,
            'phone':                shop.phone,
            'address':              shop.address,
            'email':                shop.email,
            'username':             shop.username or '',
            'shop_number':          shop.shop_number or '',
            'bill_number':          shop.bill_number or '1001',
            'gst_number':           shop.gst_number or '',
            'license_end':          end.strftime('%Y-%m-%d'),
            'days_left':            days_left,
            'warning':              0 <= days_left <= 10,
            'expired':              days_left < 0,
            'logo_url':             logo_url,
            'renewal_status':       renewal_status,
            'renewal_requested_at': renewal_requested_at
        })
    d=request.json
    shop.shop_name  = d.get('shop_name',  shop.shop_name)
    shop.owner_name = d.get('owner_name', shop.owner_name)
    shop.phone      = d.get('phone',      shop.phone)
    shop.address    = d.get('address',    shop.address)
    shop.shop_number= d.get('shop_number',shop.shop_number)
    shop.gst_number = d.get('gst_number',  shop.gst_number)

    # Handle Username update
    if 'username' in d:
        new_un = str(d.get('username','')).strip()
        if new_un != (shop.username or ''):
            if new_un:
                import re
                if not re.match(r'^[a-zA-Z0-9_-]{3,30}$', new_un):
                    return jsonify({'error': 'Username must be 3-30 characters (letters, numbers, _ or -)'}), 400
                existing = Shop.query.filter(db.func.lower(Shop.username) == new_un.lower(), Shop.id != shop.id).first()
                if existing:
                    return jsonify({'error': 'Username is already taken by another user'}), 409
                shop.username = new_un
            else:
                shop.username = None

    # Allow updating bill number
    if d.get('bill_number'):
        shop.bill_number = str(d['bill_number'])
    db.session.commit()
    return jsonify({'message':'Updated'})

@app.route('/api/send-license-renewal-otp', methods=['POST'])
@jwt_required()
@limiter.limit("10 per minute")
def send_license_renewal_otp():
    try:
        shop = Shop.query.get(int(get_jwt_identity()))
        if not shop:
            return jsonify({'error': 'Shop not found'}), 404
        
        email = shop.email
        otp = str(random.randint(100000, 999999))
        temp_license_renewal_otps[shop.id] = {'otp': otp, 'expires': time.time() + 600}
        
        print(f"\n{'='*40}")
        print(f"  LICENSE RENEWAL OTP FOR: {email} ({shop.shop_name})")
        print(f"  OTP CODE: {otp}")
        print(f"{'='*40}\n")
        
        # 1. Send OTP Email to User
        user_subject = f"Your License Renewal OTP: {otp} - Sanjana Software"
        user_text = f"Hello {shop.owner_name},\n\nYour OTP for license renewal is: {otp}\nValid for 10 minutes.\n\nSanjana Software"
        user_html = f"""<div style="font-family:Arial,sans-serif;max-width:450px;margin:0 auto;padding:20px;border:1px solid #e0e0e0;border-radius:10px">
  <h2 style="color:#1a237e">🔑 License Renewal OTP</h2>
  <p>Hello <b>{shop.owner_name}</b> ({shop.shop_name}),</p>
  <p>You have requested to renew your license for <b>Sanjana Software</b>.</p>
  <p>Your 6-digit verification code is:</p>
  <div style="background:#f0f4ff;border:2px solid #1a237e;border-radius:10px;padding:20px;text-align:center;margin:16px 0">
    <span style="font-size:32px;font-weight:900;letter-spacing:8px;color:#1a237e">{otp}</span>
  </div>
  <p style="color:#888;font-size:12px">Valid for 10 minutes. Do not share this OTP.</p>
</div>"""
        user_email_sent = send_email_notification(email, user_subject, user_text, user_html)
        
        # 2. Send Notification Email to Admin (sanjanasoftware03@gmail.com)
        admin_email = "sanjanasoftware03@gmail.com"
        admin_subject = f"🔔 License Renewal OTP Requested: {shop.shop_name}"
        admin_text = f"License Renewal requested by shop!\n\nShop Name: {shop.shop_name}\nOwner: {shop.owner_name}\nEmail: {shop.email}\nPhone: {shop.phone}\nCurrent Expiry: {shop.license_end}\nRequest Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        admin_html = f"""<div style="font-family:Arial,sans-serif;max-width:500px;margin:0 auto;padding:20px;border:1px solid #30363d;border-radius:10px;background:#161b22;color:#ffffff">
  <h2 style="color:#58a6ff">⚙️ License Renewal Notification</h2>
  <p>A user has generated an OTP to renew their license:</p>
  <table style="width:100%;border-collapse:collapse;color:#ffffff;margin:15px 0">
    <tr><td style="padding:8px;border-bottom:1px solid #30363d;color:#8b949e">Shop Name</td><td style="padding:8px;border-bottom:1px solid #30363d"><b>{shop.shop_name}</b></td></tr>
    <tr><td style="padding:8px;border-bottom:1px solid #30363d;color:#8b949e">Owner Name</td><td style="padding:8px;border-bottom:1px solid #30363d">{shop.owner_name}</td></tr>
    <tr><td style="padding:8px;border-bottom:1px solid #30363d;color:#8b949e">Email</td><td style="padding:8px;border-bottom:1px solid #30363d">{shop.email}</td></tr>
    <tr><td style="padding:8px;border-bottom:1px solid #30363d;color:#8b949e">Phone</td><td style="padding:8px;border-bottom:1px solid #30363d">{shop.phone}</td></tr>
    <tr><td style="padding:8px;border-bottom:1px solid #30363d;color:#8b949e">Current Expiry</td><td style="padding:8px;border-bottom:1px solid #30363d">{shop.license_end.strftime('%Y-%m-%d') if shop.license_end else 'N/A'}</td></tr>
    <tr><td style="padding:8px;color:#8b949e">Request Time</td><td style="padding:8px">{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</td></tr>
  </table>
  <p style="color:#8b949e;font-size:12px">Log into Owner Control Panel to approve or reject requests after OTP verification.</p>
</div>"""
        threading.Thread(target=send_email_notification, args=(admin_email, admin_subject, admin_text, admin_html)).start()

        if user_email_sent:
            return jsonify({'message': f'OTP sent to your email ({email}) and admin notified!'}), 200
        else:
            return jsonify({
                'message': 'OTP generated! Email delivery failed.',
                'otp_console': True,
                'otp_hint': otp,
                'error_detail': 'Could not deliver email to SMTP server'
            }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/profile/submit-renewal-request', methods=['POST'])
@jwt_required()
def submit_renewal_request():
    try:
        shop = Shop.query.get(int(get_jwt_identity()))
        if not shop:
            return jsonify({'error': 'Shop not found'}), 404
        
        data = request.json or {}
        otp = str(data.get('otp', '')).strip()
        
        stored = temp_license_renewal_otps.get(shop.id)
        if not stored or stored.get('otp') != otp or time.time() > stored.get('expires', 0):
            return jsonify({'error': 'Invalid or expired OTP'}), 400
        
        # Clear OTP
        temp_license_renewal_otps.pop(shop.id, None)
        
        # Check existing pending request or create new
        existing_req = LicenseRenewalRequest.query.filter_by(shop_id=shop.id, status='pending').first()
        if existing_req:
            existing_req.requested_at = datetime.now()
        else:
            new_req = LicenseRenewalRequest(shop_id=shop.id, status='pending', requested_at=datetime.now())
            db.session.add(new_req)
        
        db.session.commit()
        
        # Notify Admin via email that OTP was verified and request is pending approval
        admin_email = "sanjanasoftware03@gmail.com"
        admin_subject = f"⏳ License Renewal Application Pending Approval: {shop.shop_name}"
        admin_text = f"License Renewal request verified via OTP by shop!\n\nShop Name: {shop.shop_name}\nOwner: {shop.owner_name}\nEmail: {shop.email}\nPhone: {shop.phone}\n\nPlease review in Owner Control Panel."
        admin_html = f"""<div style="font-family:Arial,sans-serif;max-width:500px;margin:0 auto;padding:20px;border:1px solid #30363d;border-radius:10px;background:#161b22;color:#ffffff">
  <h2 style="color:#ffc107">⏳ License Renewal Application Submitted</h2>
  <p>Shop <b>{shop.shop_name}</b> has successfully verified their OTP and submitted a License Renewal Request.</p>
  <p><b>Owner:</b> {shop.owner_name}<br><b>Email:</b> {shop.email}<br><b>Phone:</b> {shop.phone}</p>
</div>"""
        threading.Thread(target=send_email_notification, args=(admin_email, admin_subject, admin_text, admin_html)).start()
        
        return jsonify({'message': '✅ OTP Verified! License renewal request submitted to Admin for approval.'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/send-password-otp', methods=['POST'])
@jwt_required()
@limiter.limit("10 per minute")
def send_password_otp():
    try:
        shop = Shop.query.get(int(get_jwt_identity()))
        if not shop:
            return jsonify({'error': 'Shop not found'}), 404
        
        email = shop.email
        otp = str(random.randint(100000, 999999))
        temp_password_otps[shop.id] = {'otp': otp, 'expires': time.time() + 600}
        
        print(f"\n{'='*40}")
        print(f"  PASSWORD CHANGE OTP FOR: {email}")
        print(f"  OTP CODE: {otp}")
        print(f"{'='*40}\n")
        
        SENDER_EMAIL    = 'sanjanasoftware03@gmail.com'
        SENDER_PASSWORD = 'xqgylrjhthhnxfmd'
        
        email_sent = False
        err_msg = ''
        try:
            from email.mime.multipart import MIMEMultipart
            from email.mime.text import MIMEText as MT
            
            msg_root = MIMEMultipart('alternative')
            msg_root['Subject'] = f'Your Password Change OTP: {otp}'
            msg_root['From']    = f'Sanjana Software <{SENDER_EMAIL}>'
            msg_root['To']      = email
            
            text_body = f"Hello {shop.owner_name},\n\nYour OTP to change your password is: {otp}\n\nValid for 10 minutes.\n\nSanjana Software"
            html_body = f"""<div style="font-family:Arial,sans-serif;max-width:400px;margin:0 auto;padding:20px">
  <h2 style="color:#1a237e">Sanjana Software</h2>
  <p>Hello <b>{shop.owner_name}</b>,</p>
  <p>Your password change OTP is:</p>
  <div style="background:#f0f4ff;border:2px solid #1a237e;border-radius:10px;padding:20px;text-align:center;margin:16px 0">
    <span style="font-size:32px;font-weight:900;letter-spacing:8px;color:#1a237e">{otp}</span>
  </div>
  <p style="color:#888;font-size:12px">Valid for 10 minutes. Do not share this OTP.</p>
</div>"""
            msg_root.attach(MT(text_body, 'plain'))
            msg_root.attach(MT(html_body, 'html'))
            
            try:
                s = smtplib.SMTP_SSL('smtp.gmail.com', 465, timeout=15)
                s.login(SENDER_EMAIL, SENDER_PASSWORD)
                s.sendmail(SENDER_EMAIL, email, msg_root.as_string())
                s.quit()
                email_sent = True
            except Exception as ssl_err:
                s2 = smtplib.SMTP('smtp.gmail.com', 587, timeout=15)
                s2.ehlo(); s2.starttls(); s2.ehlo()
                s2.login(SENDER_EMAIL, SENDER_PASSWORD)
                s2.sendmail(SENDER_EMAIL, email, msg_root.as_string())
                s2.quit()
                email_sent = True
        except Exception as mail_err:
            err_msg = str(mail_err)
            print(f"[PASSWORD OTP] ❌ Email failed: {err_msg}")

        if email_sent:
            return jsonify({'message': f'OTP sent to registered email ({email})!'}), 200
        else:
            return jsonify({
                'message': 'OTP generated! Email could not be sent.',
                'otp_console': True,
                'otp_hint': otp,
                'error_detail': err_msg[:120]
            }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/change-password-otp', methods=['POST'])
@jwt_required()
def change_password_otp():
    try:
        shop = Shop.query.get(int(get_jwt_identity()))
        data = request.json
        otp = str(data.get('otp', '')).strip()
        new_pw = data.get('new_password', '')
        
        if not otp:
            return jsonify({'error': 'OTP is required'}), 400
        if shop.id not in temp_password_otps:
            return jsonify({'error': 'OTP expired or not requested. Please click Send OTP.'}), 400
        
        otp_info = temp_password_otps[shop.id]
        if time.time() > otp_info.get('expires', 0):
            del temp_password_otps[shop.id]
            return jsonify({'error': 'OTP expired. Please request a new one.'}), 400
        
        if otp_info.get('otp') != otp:
            return jsonify({'error': 'Wrong OTP. Please check and try again.'}), 400
        
        if len(new_pw) < 6:
            return jsonify({'error': 'New password must be at least 6 characters'}), 400
        
        shop.password_hash = generate_password_hash(new_pw)
        db.session.commit()
        del temp_password_otps[shop.id]
        return jsonify({'message': '✅ Password changed successfully with OTP!'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/send-email-change-otp', methods=['POST'])
@jwt_required()
@limiter.limit("10 per minute")
def send_email_change_otp():
    try:
        shop = Shop.query.get(int(get_jwt_identity()))
        if not shop:
            return jsonify({'error': 'Shop not found'}), 404
        
        data = request.json or {}
        new_email = data.get('new_email', '').strip()

        if not new_email or '@' not in new_email or '.' not in new_email:
            return jsonify({'error': 'Please enter a valid new email address'}), 400

        if new_email.lower() == shop.email.lower():
            return jsonify({'error': 'New email address must be different from your current registered email'}), 400

        existing = Shop.query.filter(db.func.lower(Shop.email) == new_email.lower(), Shop.id != shop.id).first()
        if existing:
            return jsonify({'error': 'This email address is already registered with another shop account'}), 409

        otp = str(random.randint(100000, 999999))
        temp_email_change_otps[shop.id] = {'new_email': new_email, 'otp': otp, 'expires': time.time() + 600}

        print(f"\n{'='*40}")
        print(f"  EMAIL CHANGE OTP FOR SHOP ID {shop.id}")
        print(f"  NEW EMAIL: {new_email}")
        print(f"  OTP CODE : {otp}")
        print(f"{'='*40}\n")

        SENDER_EMAIL    = 'sanjanasoftware03@gmail.com'
        SENDER_PASSWORD = 'xqgylrjhthhnxfmd'

        email_sent = False
        err_msg = ''
        try:
            from email.mime.multipart import MIMEMultipart
            from email.mime.text import MIMEText as MT

            msg_root = MIMEMultipart('alternative')
            msg_root['Subject'] = f'Your Email Change OTP: {otp} - Sanjana Software'
            msg_root['From']    = f'Sanjana Software <{SENDER_EMAIL}>'
            msg_root['To']      = new_email

            text_body = f"Hello {shop.owner_name},\n\nYour OTP to change your registered email for {shop.shop_name} is: {otp}\n\nValid for 10 minutes.\n\nSanjana Software"
            html_body = f"""<div style="font-family:Arial,sans-serif;max-width:400px;margin:0 auto;padding:20px;border:1px solid #e0e0e0;border-radius:10px">
  <h2 style="color:#1a237e">Sanjana Software</h2>
  <p>Hello <b>{shop.owner_name}</b>,</p>
  <p>You requested to update your registered email address for <b>{shop.shop_name}</b> to: <b style="color:#1a73e8">{new_email}</b>.</p>
  <p>Your verification OTP is:</p>
  <div style="background:#f0f4ff;border:2px solid #1a237e;border-radius:10px;padding:20px;text-align:center;margin:16px 0">
    <span style="font-size:32px;font-weight:900;letter-spacing:8px;color:#1a237e">{otp}</span>
  </div>
  <p style="color:#888;font-size:12px">Valid for 10 minutes. Do not share this OTP with anyone.</p>
</div>"""
            msg_root.attach(MT(text_body, 'plain'))
            msg_root.attach(MT(html_body, 'html'))

            try:
                s = smtplib.SMTP_SSL('smtp.gmail.com', 465, timeout=15)
                s.login(SENDER_EMAIL, SENDER_PASSWORD)
                s.sendmail(SENDER_EMAIL, new_email, msg_root.as_string())
                s.quit()
                email_sent = True
            except Exception as ssl_err:
                s2 = smtplib.SMTP('smtp.gmail.com', 587, timeout=15)
                s2.ehlo(); s2.starttls(); s2.ehlo()
                s2.login(SENDER_EMAIL, SENDER_PASSWORD)
                s2.sendmail(SENDER_EMAIL, new_email, msg_root.as_string())
                s2.quit()
                email_sent = True
        except Exception as mail_err:
            err_msg = str(mail_err)
            print(f"[EMAIL CHANGE OTP] ❌ Email failed: {err_msg}")

        if email_sent:
            return jsonify({'message': f'OTP sent to your new email ({new_email})!'}), 200
        else:
            return jsonify({
                'message': 'OTP generated! Email delivery failed.',
                'otp_console': True,
                'otp_hint': otp,
                'error_detail': err_msg[:120]
            }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/verify-email-change-otp', methods=['POST'])
@jwt_required()
def verify_email_change_otp():
    try:
        shop = Shop.query.get(int(get_jwt_identity()))
        data = request.json or {}
        otp = str(data.get('otp', '')).strip()
        new_email = data.get('new_email', '').strip()

        if not otp:
            return jsonify({'error': 'OTP is required'}), 400

        if shop.id not in temp_email_change_otps:
            return jsonify({'error': 'OTP expired or not requested. Please click Send OTP.'}), 400

        otp_info = temp_email_change_otps[shop.id]
        if time.time() > otp_info.get('expires', 0):
            del temp_email_change_otps[shop.id]
            return jsonify({'error': 'OTP expired. Please request a new one.'}), 400

        if otp_info.get('otp') != otp:
            return jsonify({'error': 'Wrong OTP. Please check and try again.'}), 400

        target_email = otp_info.get('new_email', new_email).strip()
        if not target_email or '@' not in target_email:
            return jsonify({'error': 'Invalid new email address'}), 400

        existing = Shop.query.filter(db.func.lower(Shop.email) == target_email.lower(), Shop.id != shop.id).first()
        if existing:
            return jsonify({'error': 'This email address is already registered with another account'}), 409

        shop.email = target_email
        if shop.id in active_sessions:
            active_sessions[shop.id]['email'] = target_email

        db.session.commit()
        backup_bg()
        del temp_email_change_otps[shop.id]

        print(f"[SANJANA] ✅ Email changed for shop ID {shop.id} -> {target_email}")
        return jsonify({'message': f'✅ Registered email updated to {target_email}!', 'new_email': target_email}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/change-password', methods=['POST'])
@jwt_required()
def change_password():
    shop=Shop.query.get(int(get_jwt_identity()))
    d=request.json
    old_pw = d.get('old_password','')
    new_pw = d.get('new_password','')
    if not check_password_hash(shop.password_hash, old_pw):
        return jsonify({'error':'Current password is incorrect'}),400
    if len(new_pw)<6:
        return jsonify({'error':'New password must be at least 6 characters'}),400
    shop.password_hash=generate_password_hash(new_pw)
    db.session.commit()
    return jsonify({'message':'Password changed successfully!'})

@app.route('/api/upload-logo', methods=['POST'])
@jwt_required()
def upload_logo():
    """Upload shop logo — saved as static/logo_<shop_id>.png"""
    import base64, os
    shop=Shop.query.get(int(get_jwt_identity()))
    d=request.json
    img_data=d.get('image_data','')  # base64 string
    if not img_data:
        return jsonify({'error':'No image data'}),400
    try:
        # Strip data URL prefix if present
        if ',' in img_data:
            img_data=img_data.split(',',1)[1]
        logo_path=os.path.join(STATIC_FOLDER, f'logo_{shop.id}.png')
        with open(logo_path,'wb') as f:
            f.write(base64.b64decode(img_data))
        # Also copy as default logo.png if not exists
        default_logo=os.path.join(STATIC_FOLDER,'logo.png')
        if not os.path.exists(default_logo):
            import shutil
            shutil.copy(logo_path, default_logo)
        return jsonify({'message':'Logo uploaded!','logo_url':f'/static/logo_{shop.id}.png'})
    except Exception as e:
        return jsonify({'error':str(e)}),500

@app.route('/api/sales-chart', methods=['GET'])
@jwt_required()
def sales_chart():
    """Returns daily/weekly/monthly sales data for bar chart"""
    sid=int(get_jwt_identity())
    period=request.args.get('period','week')  # day / week / month
    try:
        if period=='day':
            # Last 24 hours by hour
            rows=db.session.execute(text(
                "SELECT strftime('%H:00',bill_date) as label, "
                "COUNT(*) as bills, SUM(total_amount) as total "
                "FROM bill WHERE shop_id=:sid "
                "AND bill_date >= datetime('now','-1 day') "
                "GROUP BY label ORDER BY label"
            ),{'sid':sid}).fetchall()
        elif period=='week':
            # Last 7 days
            rows=db.session.execute(text(
                "SELECT strftime('%d/%m',bill_date) as label, "
                "COUNT(*) as bills, SUM(total_amount) as total "
                "FROM bill WHERE shop_id=:sid "
                "AND bill_date >= datetime('now','-7 days') "
                "GROUP BY label ORDER BY label"
            ),{'sid':sid}).fetchall()
        else:  # month
            # Last 30 days grouped by week
            rows=db.session.execute(text(
                "SELECT strftime('%d/%m',bill_date) as label, "
                "COUNT(*) as bills, SUM(total_amount) as total "
                "FROM bill WHERE shop_id=:sid "
                "AND bill_date >= datetime('now','-30 days') "
                "GROUP BY label ORDER BY label"
            ),{'sid':sid}).fetchall()

        data=[{'label':r[0],'bills':r[1],'total':round(r[2] or 0,2)} for r in rows]
        total_sum=sum(r['total'] for r in data)
        return jsonify({'data':data,'period':period,'total':total_sum})
    except Exception as e:
        return jsonify({'error':str(e),'data':[],'total':0})

# ══════════════════════════════════════════════════════════════════════
#  DOCTORS
# ══════════════════════════════════════════════════════════════════════
@app.route('/api/doctors', methods=['GET','POST'])
@jwt_required()
def handle_doctors():
    sid=int(get_jwt_identity())
    if request.method=='GET':
        docs=Doctor.query.filter_by(shop_id=sid).all()
        return jsonify([{'id':d.id,'name':d.name,'phone':d.phone,'spec':d.spec} for d in docs])
    d=request.json
    db.session.add(Doctor(shop_id=sid,name=d['name'],phone=d.get('phone',''),spec=d.get('spec','')))
    db.session.commit()
    return jsonify({'message':'Doctor added'})

@app.route('/api/doctors/<int:did>', methods=['DELETE','PUT'])
@jwt_required()
def modify_doctor(did):
    doc=Doctor.query.get_or_404(did)
    if request.method=='DELETE':
        db.session.delete(doc)
    else:
        d=request.json
        doc.name=d.get('name',doc.name); doc.phone=d.get('phone',doc.phone); doc.spec=d.get('spec',doc.spec)
    db.session.commit()
    return jsonify({'message':'Done'})

# ══════════════════════════════════════════════════════════════════════
#  PATIENTS
# ══════════════════════════════════════════════════════════════════════
@app.route('/api/patients', methods=['GET','POST'])
@jwt_required()
def handle_patients():
    sid=int(get_jwt_identity())
    if request.method=='GET':
        pats=Patient.query.filter_by(shop_id=sid).all()
        return jsonify([{'id':p.id,'name':p.name,'phone':p.phone,'address':p.address} for p in pats])
    d=request.json
    db.session.add(Patient(shop_id=sid,name=d['name'],phone=d.get('phone',''),address=d.get('address','')))
    db.session.commit()
    return jsonify({'message':'Patient added'})

@app.route('/api/patients/upload', methods=['POST'])
@jwt_required()
def upload_patients_excel():
    sid = int(get_jwt_identity())
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    try:
        import pandas as pd
        df = pd.read_excel(file)
        df.columns = df.columns.str.strip().str.lower()
        
        added_count = 0
        for idx, row in df.iterrows():
            name = ''
            phone = ''
            address = ''
            
            if 'name' in df.columns and not pd.isna(row['name']):
                name = str(row['name']).strip()
            elif len(df.columns) > 0 and not pd.isna(row.iloc[0]):
                name = str(row.iloc[0]).strip()
                
            if 'phone' in df.columns and not pd.isna(row['phone']):
                # Some phones might be parsed as floats, so remove .0
                phone = str(row['phone']).strip()
                if phone.endswith('.0'): phone = phone[:-2]
            elif len(df.columns) > 1 and not pd.isna(row.iloc[1]):
                phone = str(row.iloc[1]).strip()
                if phone.endswith('.0'): phone = phone[:-2]
                
            if 'address' in df.columns and not pd.isna(row['address']):
                address = str(row['address']).strip()
            elif len(df.columns) > 2 and not pd.isna(row.iloc[2]):
                address = str(row.iloc[2]).strip()
                
            if name:
                db.session.add(Patient(shop_id=sid, name=name, phone=phone, address=address))
                added_count += 1
                
        db.session.commit()
        return jsonify({'message': f'Successfully imported {added_count} patients'})
    except Exception as e:
        print("Excel Import Error:", e)
        return jsonify({'error': str(e)}), 500

@app.route('/api/patients/<int:pid>', methods=['DELETE','PUT'])
@jwt_required()
def modify_patient(pid):
    pat=Patient.query.get_or_404(pid)
    if request.method=='DELETE':
        db.session.delete(pat)
    else:
        d=request.json
        pat.name=d.get('name',pat.name); pat.phone=d.get('phone',pat.phone); pat.address=d.get('address',pat.address)
    db.session.commit()
    return jsonify({'message':'Done'})

# ══════════════════════════════════════════════════════════════════════
#  SUPPLIERS  (saved exactly like Doctor list)
# ══════════════════════════════════════════════════════════════════════
@app.route('/api/suppliers', methods=['GET','POST'])
@jwt_required()
def handle_suppliers():
    sid=int(get_jwt_identity())
    if request.method=='GET':
        sups=Supplier.query.filter_by(shop_id=sid).all()
        return jsonify([{
            'id':s.id,'name':s.name,'phone':s.phone,
            'address':s.address,'gst_number':s.gst_number,'company':s.company
        } for s in sups])
    d=request.json
    if not d.get('name'):
        return jsonify({'error':'Supplier name is required'}),400
    db.session.add(Supplier(
        shop_id=sid, name=d['name'],
        phone=d.get('phone',''), address=d.get('address',''),
        gst_number=d.get('gst_number',''), company=d.get('company','')
    ))
    db.session.commit()
    return jsonify({'message':'Supplier added'})

@app.route('/api/suppliers/<int:sid2>', methods=['DELETE','PUT'])
@jwt_required()
def modify_supplier(sid2):
    sup=Supplier.query.get_or_404(sid2)
    if request.method=='DELETE':
        db.session.delete(sup)
    else:
        d=request.json
        sup.name       = d.get('name',       sup.name)
        sup.phone      = d.get('phone',      sup.phone)
        sup.address    = d.get('address',    sup.address)
        sup.gst_number = d.get('gst_number', sup.gst_number)
        sup.company    = d.get('company',    sup.company)
    db.session.commit()
    return jsonify({'message':'Done'})

# ══════════════════════════════════════════════════════════════════════
#  PURCHASE ENTRY  (like the screen in the photo)
# ══════════════════════════════════════════════════════════════════════
@app.route('/api/purchases', methods=['GET','POST'])
@jwt_required()
def handle_purchases():
    import json
    sid=int(get_jwt_identity())
    if request.method=='GET':
        entries=PurchaseEntry.query.filter_by(shop_id=sid).order_by(PurchaseEntry.created_at.desc()).limit(100).all()
        return jsonify([{
            'id':e.id,'entry_number':e.entry_number,'supplier_name':e.supplier_name,
            'party_number':e.party_number,'entry_date':e.entry_date,'entry_type':e.entry_type,
            'value_of_goods':e.value_of_goods,'discount':e.discount,'gst':e.gst,
            'net_amount':e.net_amount,
            'created_at':e.created_at.strftime('%Y-%m-%d %H:%M') if e.created_at else '',
            'items':json.loads(e.items_json or '[]')
        } for e in entries])

    d=request.json
    try:
        items=d.get('items',[])

        for item in items:
            med_id = item.get('medicine_id')
            name   = item.get('name','').strip()
            if not name: continue   # skip empty rows

            if med_id:
                # ── Existing medicine → update stock ──────────────────
                med = Medicine.query.get(med_id)
                if med:
                    med.quantity     += int(item.get('qty', 0))
                    med.price         = float(item.get('p_rate', med.price))
                    med.mrp           = float(item.get('mrp', med.mrp))
                    med.gst           = float(item.get('gst_pct', med.gst))
                    med.batch         = item.get('batch', med.batch) or med.batch
                    med.expiry_date   = parse_expiry_to_ym(item.get('expiry', med.expiry_date)) or med.expiry_date
                    med.category      = item.get('category', med.category) or med.category
                    if item.get('company'): med.company_name = item['company']
                    if item.get('discount_pct', 0):
                        med.sale_discount = float(item['discount_pct'])
                    if item.get('pack_size'):
                        med.pack_size = str(item['pack_size'])
            else:
                # ── New medicine → CREATE in stock ────────────────────
                # Check if same name already exists for this shop
                existing = Medicine.query.filter_by(shop_id=sid, name=name).first()
                if existing:
                    # Update the existing one
                    existing.quantity     += int(item.get('qty', 0))
                    existing.price         = float(item.get('p_rate', existing.price))
                    existing.mrp           = float(item.get('mrp', existing.mrp))
                    existing.gst           = float(item.get('gst_pct', existing.gst))
                    existing.batch         = item.get('batch', existing.batch) or existing.batch
                    existing.expiry_date   = parse_expiry_to_ym(item.get('expiry', existing.expiry_date)) or existing.expiry_date
                    existing.category      = item.get('category', existing.category) or existing.category
                    if item.get('company'): existing.company_name = item['company']
                    if item.get('discount_pct', 0):
                        existing.sale_discount = float(item['discount_pct'])
                    if item.get('pack_size'):
                        existing.pack_size = str(item['pack_size'])
                else:
                    # Create brand new medicine record
                    p_rate = float(item.get('p_rate', 0)) or 1.0
                    new_med = Medicine(
                        shop_id       = sid,
                        name          = name,
                        category      = item.get('category', 'General'),
                        batch         = item.get('batch', ''),
                        quantity      = int(item.get('qty', 0)),
                        price         = p_rate,
                        mrp           = float(item.get('mrp', p_rate)),
                        gst           = float(item.get('gst_pct', 0)),
                        expiry_date   = parse_expiry_to_ym(item.get('expiry', '')),
                        supplier_name = d.get('supplier_name', ''),
                        company_name  = item.get('company', ''),
                        pack_size     = str(item.get('pack_size', '10')),
                        sale_discount = float(item.get('discount_pct', 0))
                    )
                    db.session.add(new_med)

        # Get next purchase entry number
        last     = PurchaseEntry.query.filter_by(shop_id=sid).order_by(PurchaseEntry.id.desc()).first()
        entry_no = str((int(last.entry_number) if last and str(last.entry_number).isdigit() else 0)+1).zfill(4)

        # GST calculations for purchase
        val_goods = float(d.get('value_of_goods', 0))
        discount = float(d.get('discount', 0))
        net_amount = float(d.get('net_amount', 0))
        
        # Calculate details from items
        taxable_sum = 0.0
        gst_sum = 0.0
        hsn_list = []
        gst_rates = []
        
        for item in items:
            item_hsn = str(item.get('hsn_code') or item.get('hsn') or '').strip()
            if item_hsn and item_hsn not in hsn_list:
                hsn_list.append(item_hsn)
                
            gst_pct = float(item.get('gst_pct', 0) or item.get('gst', 0))
            if gst_pct not in gst_rates:
                gst_rates.append(gst_pct)
                
            qty = float(item.get('qty', 0))
            prate = float(item.get('p_rate', 0))
            dis_pct = float(item.get('discount_pct', 0))
            
            base = qty * prate
            disc_amt = (base * dis_pct) / 100.0
            item_taxable = base - disc_amt
            item_gst = item_taxable * (gst_pct / 100.0)
            
            taxable_sum += item_taxable
            gst_sum += item_gst
            
        predominant_gst = gst_rates[0] if gst_rates else 0.0
        
        sup_gstin = str(d.get('supplier_gstin', '')).strip().upper()
        place_supply = str(d.get('place_of_supply', '')).strip()
        shop = Shop.query.get(sid)
        
        is_interstate = False
        if place_supply and shop.gst_number and len(shop.gst_number) >= 2:
            is_interstate = (place_supply[:2] != shop.gst_number[:2])
        elif sup_gstin and len(sup_gstin) >= 2 and shop.gst_number and len(shop.gst_number) >= 2:
            is_interstate = (sup_gstin[:2] != shop.gst_number[:2])
            
        if is_interstate:
            igst_rate = predominant_gst
            igst_amount = gst_sum
            cgst_rate = 0.0
            cgst_amount = 0.0
            sgst_rate = 0.0
            sgst_amount = 0.0
        else:
            igst_rate = 0.0
            igst_amount = 0.0
            cgst_rate = predominant_gst / 2.0
            cgst_amount = gst_sum / 2.0
            sgst_rate = predominant_gst / 2.0
            sgst_amount = gst_sum / 2.0

        fys = FinancialYear.query.all()
        fy_id = None
        entry_date_str = d.get('entry_date', '')
        for f in fys:
            try:
                f_start = datetime.strptime(f.start_date, "%Y-%m-%d").date()
                f_end = datetime.strptime(f.end_date, "%Y-%m-%d").date()
                dt = datetime.strptime(str(entry_date_str)[:10], "%Y-%m-%d").date()
                if f_start <= dt <= f_end:
                    fy_id = f.id
                    break
            except:
                pass
        if not fy_id:
            active = next((f for f in fys if f.is_active), None)
            fy_id = active.id if active else None

        entry = PurchaseEntry(
            shop_id       = sid,
            entry_number  = entry_no,
            supplier_name = d.get('supplier_name',''),
            party_number  = d.get('party_number',''),
            entry_date    = entry_date_str,
            entry_type    = d.get('entry_type','Purchase'),
            value_of_goods= val_goods,
            discount      = discount,
            gst           = gst_sum,
            net_amount    = net_amount,
            items_json    = json.dumps(items),
            supplier_gstin= sup_gstin,
            place_of_supply= place_supply,
            hsn_code      = ', '.join(hsn_list),
            gst_rate      = predominant_gst,
            taxable_amount= round(taxable_sum, 2),
            cgst_rate     = cgst_rate,
            cgst_amount   = round(cgst_amount, 2),
            sgst_rate     = sgst_rate,
            sgst_amount   = round(sgst_amount, 2),
            igst_rate     = igst_rate,
            igst_amount   = round(igst_amount, 2),
            total_gst     = round(gst_sum, 2),
            grand_total   = net_amount,
            financial_year_id = fy_id
        )
        db.session.add(entry)
        db.session.commit()
        backup_bg()
        return jsonify({'message':'Purchase entry saved','entry_number':entry_no}),201
    except Exception as e:
        db.session.rollback()
        print(f"[PURCHASE ERROR] {e}")
        return jsonify({'error':str(e)}),500

@app.route('/api/purchases/<int:eid>', methods=['DELETE'])
@jwt_required()
def delete_purchase(eid):
    import json
    sid = int(get_jwt_identity())
    entry = PurchaseEntry.query.filter_by(id=eid, shop_id=sid).first_or_404()
    
    # Reverse stock automatically
    if entry.items_json:
        try:
            items = json.loads(entry.items_json)
            for item in items:
                name = item.get('name', '').strip()
                qty = int(item.get('qty', 0))
                med_id = item.get('medicine_id')
                
                if med_id:
                    med = Medicine.query.filter_by(id=med_id, shop_id=sid).first()
                else:
                    med = Medicine.query.filter_by(name=name, shop_id=sid).first()
                
                if med:
                    med.quantity -= qty
        except Exception as e:
            print("Error reversing stock on delete:", e)
            
    db.session.delete(entry)
    db.session.commit()
    return jsonify({'message':'Deleted and stock reversed'})

# ══════════════════════════════════════════════════════════════════════
#  MEDICINES
# ══════════════════════════════════════════════════════════════════════
def parse_expiry_to_ym(val):
    if not val:
        return ""
    import re
    val = str(val).strip()
    if val.lower() in ['nan', 'none', '']:
        return ""
    
    # Try YYYY-MM-DD -> YYYY-MM
    m = re.match(r"^(\d{4})-(\d{2})-\d{2}$", val)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
        
    # Try YYYY-MM
    if re.match(r"^\d{4}-\d{2}$", val):
        return val
        
    # Try MM/YYYY
    m = re.match(r"^(\d{1,2})/(\d{4})$", val)
    if m:
        return f"{int(m.group(2)):04d}-{int(m.group(1)):02d}"
        
    # Try MM/YY
    m = re.match(r"^(\d{1,2})/(\d{2})$", val)
    if m:
        return f"{int(m.group(2))+2000:04d}-{int(m.group(1)):02d}"

    # Try MM-YYYY
    m = re.match(r"^(\d{1,2})-(\d{4})$", val)
    if m:
        return f"{int(m.group(2)):04d}-{int(m.group(1)):02d}"
        
    # Try MM-YY
    m = re.match(r"^(\d{1,2})-(\d{2})$", val)
    if m:
        return f"{int(m.group(2))+2000:04d}-{int(m.group(1)):02d}"

    # Try DD-MM-YYYY or DD/MM/YYYY
    m = re.match(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{4})$", val)
    if m:
        return f"{int(m.group(3)):04d}-{int(m.group(2)):02d}"

    # Try digits only: MMYY or MMYYYY
    if val.isdigit():
        if len(val) == 4:
            m_part = int(val[:2])
            y_part = int(val[2:]) + 2000
            if 1 <= m_part <= 12:
                return f"{y_part:04d}-{m_part:02d}"
        elif len(val) == 6:
            m_part = int(val[:2])
            y_part = int(val[2:])
            if 1 <= m_part <= 12:
                return f"{y_part:04d}-{m_part:02d}"
        elif len(val) == 3:
            m_part = int(val[:1])
            y_part = int(val[1:]) + 2000
            if 1 <= m_part <= 9:
                return f"{y_part:04d}-{m_part:02d}"

    return val[:7]

def med_dict(m):
    return {'id':m.id,'name':m.name,'category':m.category,'batch':m.batch or '',
            'quantity':m.quantity,'price':m.price,'mrp':m.mrp or 0,'gst':m.gst or 0,
            'expiry_date':m.expiry_date or '','supplier_name':m.supplier_name or '',
            'company_name':m.company_name or '',
            'pack_size':   m.pack_size    if m.pack_size    else '10',
            'sale_discount': m.sale_discount if m.sale_discount else 0.0}

@app.route('/api/medicines', methods=['GET','POST'])
@jwt_required()
def handle_medicines():
    sid=int(get_jwt_identity())
    if request.method=='GET':
        return jsonify([med_dict(m) for m in Medicine.query.filter_by(shop_id=sid).all()])
    d=request.json
    try:
        db.session.add(Medicine(shop_id=sid,name=d['name'],category=d.get('category','General'),
            batch=d.get('batch',''),price=float(d['price']),mrp=float(d.get('mrp',d['price'])),
            quantity=int(d['quantity']),gst=float(d.get('gst',0)),
            expiry_date=parse_expiry_to_ym(d.get('expiry_date','')),supplier_name=d.get('supplier_name',''),
            company_name=d.get('company_name',''),
            pack_size=str(d.get('pack_size','10'))))  # FEATURE 3
        db.session.commit(); backup_bg()
        return jsonify({'message':'Added'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error':str(e)}),500

@app.route('/api/medicines/<int:mid>', methods=['DELETE','PUT'])
@jwt_required()
def modify_med(mid):
    med=Medicine.query.get_or_404(mid)
    try:
        if request.method=='DELETE':
            db.session.delete(med)
        else:
            d=request.json
            for k in ['name','category','batch','expiry_date','supplier_name','company_name']:
                if k in d:
                    val = d[k]
                    if k == 'expiry_date':
                        val = parse_expiry_to_ym(val)
                    setattr(med,k,val)
            for k in ['price','mrp','gst','sale_discount']:
                if k in d: setattr(med,k,float(d[k]))
            if 'quantity'  in d: med.quantity   = int(d['quantity'])
            if 'pack_size' in d: med.pack_size   = str(d['pack_size'])
        db.session.commit(); backup_bg()
        return jsonify({'message':'Done'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error':str(e)}),500

@app.route('/api/import-excel', methods=['POST'])
@jwt_required()
def import_excel():
    try:
        import pandas as pd
        sid=int(get_jwt_identity())
        df=pd.read_excel(request.files['file'])
        
        col_map = {}
        for c in df.columns:
            clean_c = str(c).strip().upper()
            if clean_c in ['NAME', 'ITEM NAME', 'MEDICINE NAME']: col_map[c] = 'Name'
            elif clean_c in ['CATEGORY', 'TYPE']: col_map[c] = 'Category'
            elif clean_c in ['BATCH', 'BATCH NO', 'BATCH NUMBER']: col_map[c] = 'Batch'
            elif clean_c in ['PRICE', 'RATE', 'PTR']: col_map[c] = 'Price'
            elif clean_c in ['MRP']: col_map[c] = 'MRP'
            elif clean_c in ['QTY', 'QUANTITY', 'STOCK']: col_map[c] = 'Quantity'
            elif clean_c in ['GST', 'GST%', 'TAX']: col_map[c] = 'GST'
            elif clean_c in ['EXPIRY', 'EXP', 'EXP DATE', 'EXPIRY DATE']: col_map[c] = 'Expiry'
            elif clean_c in ['SUPPLIER', 'DISTRIBUTOR', 'VENDOR']: col_map[c] = 'Supplier'
            elif clean_c in ['COMPANY', 'MFG', 'MANUFACTURER']: col_map[c] = 'Company'
            elif clean_c in ['PACK', 'PACK SIZE', 'PACKING']: col_map[c] = 'Pack Size'
        df.rename(columns=col_map, inplace=True)

        if 'Name' not in df.columns:
            return jsonify({'error': 'Missing required column: Name (or NAME)'}), 400

        import re
        def safe_float(val, default=0.0):
            if pd.isna(val): return default
            try:
                return float(str(val).replace('%','').strip())
            except:
                return default

        def safe_int(val, default=0):
            if pd.isna(val): return default
            try:
                nums = re.findall(r'\d+', str(val))
                return int(nums[0]) if nums else default
            except:
                return default

        import json as _json

        purchase_items = []
        total_taxable  = 0.0
        total_gst_amt  = 0.0
        supplier_set   = set()

        for _,row in df.iterrows():
            if pd.isna(row.get('Name')) or str(row.get('Name')).strip() == '': continue
            exp=row.get('Expiry',''); exp='' if pd.isna(exp) or str(exp)=='nan' else str(exp)[:10]

            # Clean string fields
            def clean_str(val, default=''):
                return default if pd.isna(val) else str(val).strip()

            name     = clean_str(row['Name'])
            qty      = safe_int(row.get('Quantity', 0))
            price    = safe_float(row.get('Price', 0))
            mrp      = safe_float(row.get('MRP', row.get('Price', 0)))
            gst_pct  = safe_float(row.get('GST', 0))
            batch    = clean_str(row.get('Batch', ''))
            supplier = clean_str(row.get('Supplier', ''))
            company  = clean_str(row.get('Company', ''))
            pack_sz  = str(row.get('Pack Size', '10'))

            if supplier:
                supplier_set.add(supplier)

            # Compute financial totals for this item
            base_val    = qty * price
            item_gst    = base_val * (gst_pct / 100.0)
            total_taxable += base_val
            total_gst_amt += item_gst

            purchase_items.append({
                'name':        name,
                'qty':         qty,
                'p_rate':      price,
                'mrp':         mrp,
                'gst_pct':     gst_pct,
                'batch':       batch,
                'expiry':      parse_expiry_to_ym(exp) or '',
                'company':     company,
                'pack_size':   pack_sz,
                'discount_pct': 0,
            })

            db.session.add(Medicine(
                shop_id=sid,
                name=name,
                category=clean_str(row.get('Category', 'General'), 'General') or 'General',
                batch=batch,
                price=price,
                mrp=mrp,
                quantity=qty,
                gst=gst_pct,
                expiry_date=parse_expiry_to_ym(exp),
                supplier_name=supplier,
                company_name=company,
                pack_size=pack_sz
            ))   # FEATURE 3

        # ── Create a Purchase History entry for this Excel import ──────────
        if purchase_items:
            last     = PurchaseEntry.query.filter_by(shop_id=sid).order_by(PurchaseEntry.id.desc()).first()
            entry_no = str((int(last.entry_number) if last and str(last.entry_number).isdigit() else 0)+1).zfill(4)

            net_amount  = round(total_taxable + total_gst_amt, 2)
            cgst_amount = round(total_gst_amt / 2.0, 2)
            sgst_amount = round(total_gst_amt / 2.0, 2)
            supplier_name = ', '.join(sorted(supplier_set)) if supplier_set else 'Excel Import'

            # Find active financial year
            fys = FinancialYear.query.all()
            fy_id = None
            today_str = datetime.now().strftime('%Y-%m-%d')
            for f in fys:
                try:
                    f_start = datetime.strptime(f.start_date, "%Y-%m-%d").date()
                    f_end   = datetime.strptime(f.end_date,   "%Y-%m-%d").date()
                    dt_today = datetime.strptime(today_str, "%Y-%m-%d").date()
                    if f_start <= dt_today <= f_end:
                        fy_id = f.id; break
                except: pass
            if not fy_id:
                active = next((f for f in fys if f.is_active), None)
                fy_id = active.id if active else None

            purchase_entry = PurchaseEntry(
                shop_id        = sid,
                entry_number   = entry_no,
                supplier_name  = supplier_name,
                party_number   = '',
                entry_date     = today_str,
                entry_type     = 'Excel Import',
                value_of_goods = round(total_taxable, 2),
                discount       = 0.0,
                gst            = round(total_gst_amt, 2),
                net_amount     = net_amount,
                items_json     = _json.dumps(purchase_items),
                supplier_gstin = '',
                place_of_supply= '',
                hsn_code       = '',
                gst_rate       = 0.0,
                taxable_amount = round(total_taxable, 2),
                cgst_rate      = 0.0,
                cgst_amount    = cgst_amount,
                sgst_rate      = 0.0,
                sgst_amount    = sgst_amount,
                igst_rate      = 0.0,
                igst_amount    = 0.0,
                total_gst      = round(total_gst_amt, 2),
                grand_total    = net_amount,
                financial_year_id = fy_id
            )
            db.session.add(purchase_entry)

        db.session.commit(); backup_bg()
        return jsonify({'message':f'Imported {len(df)} items'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error':str(e)}),500

# ══════════════════════════════════════════════════════════════════════
#  DASHBOARD
# ══════════════════════════════════════════════════════════════════════
@app.route('/api/dashboard-stats', methods=['GET'])
@jwt_required()
def get_stats():
    sid=int(get_jwt_identity()); today=datetime.now().date()
    meds=Medicine.query.filter_by(shop_id=sid).all()
    today_bills=Bill.query.filter(Bill.shop_id==sid,db.func.date(Bill.bill_date)==today).all()
    week_bills=Bill.query.filter(Bill.shop_id==sid,
        Bill.bill_date>=datetime.now()-timedelta(days=7)).all()
    from datetime import datetime as dt
    expiring=[]
    for m in meds:
        if not m.expiry_date: continue
        try:
            d_str = m.expiry_date[:10]
            fmt = '%Y-%m-%d' if len(d_str)>=10 else '%Y-%m'
            if (dt.strptime(d_str, fmt) - dt.now()).days <= 90: expiring.append(m)
        except: pass
    today_returns_list = BillReturn.query.filter(BillReturn.shop_id==sid, db.func.date(BillReturn.return_date)==today).all()
    today_returns = sum(r.refund_amount for r in today_returns_list)
    return jsonify({'total_meds':len(meds),
        'today_sales':round(sum(b.total_amount for b in today_bills),2),
        'today_returns':round(today_returns,2),
        'today_net_sales':round(sum(b.total_amount for b in today_bills)-today_returns,2),
        'week_sales':round(sum(b.total_amount for b in week_bills),2),
        'today_bills':len(today_bills),
        'low_stock':len([m for m in meds if m.quantity<10]),
        'expiring_soon':len(expiring),
        'medicines':[med_dict(m) for m in meds]})

# ══════════════════════════════════════════════════════════════════════
#  BILLING
# ══════════════════════════════════════════════════════════════════════
def parse_qty_to_strips(qty_val, pack_size):
    try:
        qty_str = str(qty_val).strip()
        if '.' in qty_str:
            parts = qty_str.split('.')
            S = int(parts[0]) if (parts[0] and parts[0].isdigit()) else 0
            T = int(parts[1]) if (parts[1] and parts[1].isdigit()) else 0
            return S + (T / pack_size)
        return float(qty_str) if qty_str else 0.0
    except Exception:
        return 0.0

def parse_qty_to_tablets(qty_val, pack_size):
    try:
        qty_str = str(qty_val).strip()
        if '.' in qty_str:
            parts = qty_str.split('.')
            S = int(parts[0]) if (parts[0] and parts[0].isdigit()) else 0
            T = int(parts[1]) if (parts[1] and parts[1].isdigit()) else 0
            return (S * pack_size) + T
        return int(float(qty_str) * pack_size)
    except Exception:
        return 0

@app.route('/api/bills', methods=['POST'])
@jwt_required()
def save_bill():
    import json
    try:
        shop=Shop.query.get(int(get_jwt_identity()))
        d=request.json; curr_no=int(shop.bill_number)
        items=d.get('items',[])

        # FEATURE 1: Strip/Pack — validate and enrich each item
        for item in items:
            item['pack_size']     = str(item.get('pack_size', '1'))
            num_str = ''.join(filter(str.isdigit, item['pack_size']))
            p_val = int(num_str) if num_str else 1
            qty_val = item.get('qty', '1')
            item['total_tablets'] = parse_qty_to_tablets(qty_val, p_val)

        # Deduct stock — prevent negative
        for item in items:
            med=Medicine.query.get(item.get('id'))
            if med:
                num_str = ''.join(filter(str.isdigit, str(med.pack_size or '10')))
                p_val = int(num_str) if num_str else 10
                qty_val = item.get('qty', 0)
                actual_qty = parse_qty_to_strips(qty_val, p_val)
                med.quantity = max(0.0, med.quantity - actual_qty)

        # FEATURE 2: Custom date — use if provided, else now
        custom_date_str = d.get('custom_date', '').strip()
        if custom_date_str:
            try:
                bill_dt = datetime.strptime(custom_date_str, '%Y-%m-%d')
                now_time = datetime.now().time()
                bill_dt = datetime.combine(bill_dt.date(), now_time)
            except ValueError:
                bill_dt = datetime.now()
                custom_date_str = bill_dt.strftime('%Y-%m-%d')
        else:
            bill_dt = datetime.now()
            custom_date_str = bill_dt.strftime('%Y-%m-%d')

        # GST calculations
        subtotal = float(d.get('subtotal', 0))
        discount = float(d.get('discount', 0))
        total_amount = float(d.get('total_amount', 0))
        
        # Calculate discount ratio for each item's taxable split
        discount_factor = 1.0 - (discount / subtotal) if subtotal > 0 else 1.0
        
        taxable_sum = 0.0
        gst_sum = 0.0
        hsn_list = []
        gst_rates = []
        
        for item in items:
            item_hsn = str(item.get('hsn_code') or item.get('hsn') or '').strip()
            if item_hsn and item_hsn not in hsn_list:
                hsn_list.append(item_hsn)
                
            gst_pct = float(item.get('gst', 0) or item.get('gst_pct', 0))
            if gst_pct not in gst_rates:
                gst_rates.append(gst_pct)
                
            item_qty = float(item.get('qty', 1))
            item_price = float(item.get('price', 0))
            item_amount = float(item.get('amount', item_qty * item_price))
            
            discounted_amount = item_amount * discount_factor
            
            tax_factor = 1.0 + (gst_pct / 100.0)
            item_taxable = discounted_amount / tax_factor
            item_gst = discounted_amount - item_taxable
            
            taxable_sum += item_taxable
            gst_sum += item_gst

        predominant_gst = gst_rates[0] if gst_rates else 0.0
        
        cust_gstin = str(d.get('customer_gstin', '')).strip().upper()
        place_supply = str(d.get('place_of_supply', '')).strip()
        
        is_interstate = False
        if place_supply and shop.gst_number and len(shop.gst_number) >= 2:
            is_interstate = (place_supply[:2] != shop.gst_number[:2])
        elif cust_gstin and len(cust_gstin) >= 2 and shop.gst_number and len(shop.gst_number) >= 2:
            is_interstate = (cust_gstin[:2] != shop.gst_number[:2])
            
        if is_interstate:
            igst_rate = predominant_gst
            igst_amount = gst_sum
            cgst_rate = 0.0
            cgst_amount = 0.0
            sgst_rate = 0.0
            sgst_amount = 0.0
        else:
            igst_rate = 0.0
            igst_amount = 0.0
            cgst_rate = predominant_gst / 2.0
            cgst_amount = gst_sum / 2.0
            sgst_rate = predominant_gst / 2.0
            sgst_amount = gst_sum / 2.0

        fys = FinancialYear.query.all()
        fy_id = None
        for f in fys:
            try:
                f_start = datetime.strptime(f.start_date, "%Y-%m-%d").date()
                f_end = datetime.strptime(f.end_date, "%Y-%m-%d").date()
                if f_start <= bill_dt.date() <= f_end:
                    fy_id = f.id
                    break
            except:
                pass
        if not fy_id:
            active = next((f for f in fys if f.is_active), None)
            fy_id = active.id if active else None

        new_bill=Bill(
            shop_id       = shop.id,
            bill_number   = str(curr_no),
            customer_name = d.get('customer_name','Walk-in'),
            customer_phone= d.get('customer_phone',''),
            doctor_name   = d.get('doctor_name',''),
            subtotal      = subtotal,
            cgst          = cgst_amount,
            sgst          = sgst_amount,
            discount      = discount,
            total_amount  = total_amount,
            bill_date     = bill_dt,
            custom_date   = custom_date_str,
            items_json    = json.dumps(items),
            customer_gstin= cust_gstin,
            place_of_supply= place_supply,
            hsn_code      = ', '.join(hsn_list),
            gst_rate      = predominant_gst,
            taxable_amount= round(taxable_sum, 2),
            cgst_rate     = cgst_rate,
            cgst_amount   = round(cgst_amount, 2),
            sgst_rate     = sgst_rate,
            sgst_amount   = round(sgst_amount, 2),
            igst_rate     = igst_rate,
            igst_amount   = round(igst_amount, 2),
            total_gst     = round(gst_sum, 2),
            grand_total   = total_amount,
            financial_year_id = fy_id
        )
        shop.bill_number=str(curr_no+1)
        db.session.add(new_bill)
        db.session.commit()
        backup_bg()
        return jsonify({'message':'Saved','bill_no':curr_no,'bill_id':new_bill.id}),201

    except Exception as e:
        db.session.rollback()
        return jsonify({'error':str(e)}),500

@app.route('/api/bills', methods=['GET'])
@jwt_required()
def get_bills():
    import json
    sid=int(get_jwt_identity())
    bills=Bill.query.filter_by(shop_id=sid).order_by(Bill.bill_date.desc()).limit(200).all()
    return jsonify([{
        'id':b.id,'bill_number':b.bill_number,
        'customer_name':b.customer_name,'customer_phone':b.customer_phone or '',
        'doctor_name':b.doctor_name,
        'subtotal':b.subtotal or 0,'cgst':b.cgst or 0,
        'sgst':b.sgst or 0,'discount':b.discount or 0,
        'total_amount':b.total_amount,
        'bill_date':b.bill_date.strftime('%Y-%m-%d %H:%M') if b.bill_date else '',
        'custom_date':b.custom_date or '',       # FEATURE 2
        'status':getattr(b,'status','active') or 'active',
        'returned_amount':float(getattr(b,'returned_amount',0.0) or 0.0),
        'items':json.loads(b.items_json or '[]')
    } for b in bills])

@app.route('/api/bills/<int:bid>', methods=['DELETE'])
@jwt_required()
def delete_bill(bid):
    import json
    sid = int(get_jwt_identity())
    bill = Bill.query.filter_by(id=bid, shop_id=sid).first_or_404()
    
    # Reverse stock automatically (sale reversed -> stock goes up)
    if bill.items_json:
        try:
            items = json.loads(bill.items_json)
            for item in items:
                med_id = item.get('id')
                qty = int(item.get('qty', 0))
                if med_id:
                    med = Medicine.query.filter_by(id=med_id, shop_id=sid).first()
                    if med:
                        med.quantity += qty
        except Exception as e:
            print("Error reversing stock on bill delete:", e)
            
    db.session.delete(bill)
    db.session.commit()
    return jsonify({'message':'Bill deleted and stock reversed'})

# ══════════════════════════════════════════════════════════════════════


# ══════════════════════════════════════════════════════════════════════
#  BILL RETURN (SALES RETURN) ENDPOINTS
# ══════════════════════════════════════════════════════════════════════

@app.route('/api/bills/<int:bid>/return-info', methods=['GET'])
@jwt_required()
def get_bill_return_info(bid):
    import json
    sid = int(get_jwt_identity())
    bill = Bill.query.filter_by(id=bid, shop_id=sid).first_or_404()
    
    returns = BillReturn.query.filter_by(bill_id=bid, shop_id=sid).all()
    returned_qty_map = {}
    for ret in returns:
        try:
            r_items = json.loads(ret.items_json or '[]')
            for r_item in r_items:
                m_id = r_item.get('id')
                r_qty = float(r_item.get('return_qty', 0))
                if m_id:
                    returned_qty_map[m_id] = returned_qty_map.get(m_id, 0.0) + r_qty
        except Exception as e:
            print("Error parsing return items:", e)
            
    sold_items = json.loads(bill.items_json or '[]')
    enriched_items = []
    for item in sold_items:
        m_id = item.get('id')
        qty_sold = float(item.get('qty', 0))
        prev_ret = returned_qty_map.get(m_id, 0.0) if m_id else 0.0
        max_ret = max(0.0, qty_sold - prev_ret)
        enriched_items.append({
            'id': m_id,
            'name': item.get('name', 'Unknown'),
            'batch': item.get('batch', ''),
            'expiry': item.get('expiry', ''),
            'mrp': float(item.get('mrp', 0)),
            'rate': float(item.get('rate', item.get('unit_price', 0))),
            'pack_size': str(item.get('pack_size', '10')),
            'sold_qty': qty_sold,
            'previously_returned_qty': prev_ret,
            'max_returnable_qty': max_ret,
            'gst': float(item.get('gst', 0)),
            'discount': float(item.get('discount', 0))
        })
        
    return jsonify({
        'bill_id': bill.id,
        'bill_number': bill.bill_number,
        'customer_name': bill.customer_name,
        'customer_phone': bill.customer_phone or '',
        'doctor_name': bill.doctor_name or '',
        'bill_date': bill.bill_date.strftime('%Y-%m-%d %H:%M') if bill.bill_date else '',
        'total_amount': bill.total_amount,
        'status': getattr(bill, 'status', 'active') or 'active',
        'returned_amount': getattr(bill, 'returned_amount', 0.0) or 0.0,
        'items': enriched_items
    })


@app.route('/api/bill-returns', methods=['POST'])
@jwt_required()
def create_bill_return():
    import json
    try:
        sid = int(get_jwt_identity())
        shop = Shop.query.get(sid)
        data = request.json or {}
        bill_id = data.get('bill_id')
        if not bill_id:
            return jsonify({'error': 'bill_id is required'}), 400
            
        bill = Bill.query.filter_by(id=bill_id, shop_id=sid).first()
        if not bill:
            return jsonify({'error': 'Bill not found'}), 404
            
        items_to_return = data.get('items', [])
        if not items_to_return:
            return jsonify({'error': 'No items selected for return'}), 400
            
        prior_returns = BillReturn.query.filter_by(bill_id=bill_id, shop_id=sid).all()
        prev_returned_map = {}
        for ret in prior_returns:
            try:
                r_items = json.loads(ret.items_json or '[]')
                for r_item in r_items:
                    m_id = r_item.get('id')
                    if m_id:
                        prev_returned_map[m_id] = prev_returned_map.get(m_id, 0.0) + float(r_item.get('return_qty', 0))
            except Exception as e:
                print("Error parsing prior returns:", e)
                
        sold_items = json.loads(bill.items_json or '[]')
        sold_map = {item.get('id'): item for item in sold_items if item.get('id')}
        
        valid_return_items = []
        total_subtotal_refund = 0.0
        total_tax_refund = 0.0
        total_refund = 0.0
        
        for r_item in items_to_return:
            med_id = r_item.get('id')
            ret_qty = float(r_item.get('return_qty', 0))
            if ret_qty <= 0:
                continue
                
            orig_item = sold_map.get(med_id)
            if not orig_item:
                return jsonify({'error': f'Item ID {med_id} not found in original bill'}), 400
                
            sold_qty = float(orig_item.get('qty', 0))
            already_ret = prev_returned_map.get(med_id, 0.0)
            max_ret = max(0.0, sold_qty - already_ret)
            
            if ret_qty > max_ret + 0.001:
                return jsonify({'error': f'Cannot return {ret_qty} units of {orig_item.get("name")}. Max returnable: {max_ret}'}), 400
                
            rate = float(r_item.get('rate', orig_item.get('rate', orig_item.get('unit_price', 0))))
            gst_pct = float(r_item.get('gst', orig_item.get('gst', 0)))
            
            item_subtotal = ret_qty * rate
            item_tax = item_subtotal * (gst_pct / 100.0) if gst_pct else 0.0
            item_refund = item_subtotal + item_tax
            
            total_subtotal_refund += item_subtotal
            total_tax_refund += item_tax
            total_refund += item_refund
            
            med = Medicine.query.filter_by(id=med_id, shop_id=sid).first()
            if med:
                num_str = ''.join(filter(str.isdigit, str(med.pack_size or '10')))
                p_val = int(num_str) if num_str else 10
                actual_strips = parse_qty_to_strips(ret_qty, p_val)
                med.quantity += actual_strips
                
            valid_return_items.append({
                'id': med_id,
                'name': orig_item.get('name', 'Unknown'),
                'batch': orig_item.get('batch', ''),
                'expiry': orig_item.get('expiry', ''),
                'pack_size': orig_item.get('pack_size', '10'),
                'rate': rate,
                'gst': gst_pct,
                'return_qty': ret_qty,
                'refund_amount': round(item_refund, 2),
                'reason': r_item.get('reason', 'Customer Return')
            })
            
        if not valid_return_items:
            return jsonify({'error': 'No valid items to return (quantity must be > 0)'}), 400
            
        curr_ret_no = int(getattr(shop, 'return_number', '1001') or '1001')
        ret_num_str = f"RET-{curr_ret_no}"
        
        new_return = BillReturn(
            shop_id=sid,
            return_number=ret_num_str,
            bill_id=bill.id,
            bill_number=bill.bill_number,
            customer_name=bill.customer_name,
            customer_phone=bill.customer_phone or '',
            return_date=datetime.now(),
            items_json=json.dumps(valid_return_items),
            subtotal=round(total_subtotal_refund, 2),
            tax_amount=round(total_tax_refund, 2),
            refund_amount=round(total_refund, 2),
            reason=data.get('reason', 'Bill Return'),
            notes=data.get('notes', '')
        )
        
        shop.return_number = str(curr_ret_no + 1)
        bill.returned_amount = (getattr(bill, 'returned_amount', 0.0) or 0.0) + total_refund
        
        all_returned = True
        for orig_item in sold_items:
            m_id = orig_item.get('id')
            s_qty = float(orig_item.get('qty', 0))
            tot_ret = prev_returned_map.get(m_id, 0.0)
            for v_item in valid_return_items:
                if v_item['id'] == m_id:
                    tot_ret += v_item['return_qty']
            if tot_ret < s_qty - 0.001:
                all_returned = False
                break
                
        bill.status = 'returned' if all_returned else 'partially_returned'
        
        db.session.add(new_return)
        db.session.commit()
        backup_bg()
        
        return jsonify({
            'message': 'Bill return processed successfully',
            'return_id': new_return.id,
            'return_number': new_return.return_number,
            'bill_number': bill.bill_number,
            'refund_amount': new_return.refund_amount,
            'return_date': new_return.return_date.strftime('%Y-%m-%d %H:%M')
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@app.route('/api/bill-returns', methods=['GET'])
@jwt_required()
def get_bill_returns():
    import json
    sid = int(get_jwt_identity())
    returns = BillReturn.query.filter_by(shop_id=sid).order_by(BillReturn.return_date.desc()).limit(200).all()
    return jsonify([{
        'id': r.id,
        'return_number': r.return_number,
        'bill_id': r.bill_id,
        'bill_number': r.bill_number,
        'customer_name': r.customer_name,
        'customer_phone': r.customer_phone or '',
        'return_date': r.return_date.strftime('%Y-%m-%d %H:%M') if r.return_date else '',
        'subtotal': r.subtotal or 0.0,
        'tax_amount': r.tax_amount or 0.0,
        'refund_amount': r.refund_amount or 0.0,
        'reason': r.reason or '',
        'notes': r.notes or '',
        'items': json.loads(r.items_json or '[]')
    } for r in returns])


@app.route('/api/bill-returns/<int:rid>', methods=['GET'])
@jwt_required()
def get_single_bill_return(rid):
    import json
    sid = int(get_jwt_identity())
    r = BillReturn.query.filter_by(id=rid, shop_id=sid).first_or_404()
    return jsonify({
        'id': r.id,
        'return_number': r.return_number,
        'bill_id': r.bill_id,
        'bill_number': r.bill_number,
        'customer_name': r.customer_name,
        'customer_phone': r.customer_phone or '',
        'return_date': r.return_date.strftime('%Y-%m-%d %H:%M') if r.return_date else '',
        'subtotal': r.subtotal or 0.0,
        'tax_amount': r.tax_amount or 0.0,
        'refund_amount': r.refund_amount or 0.0,
        'reason': r.reason or '',
        'notes': r.notes or '',
        'items': json.loads(r.items_json or '[]')
    })


#  SCHEDULE H REPORT
# ══════════════════════════════════════════════════════════════════════
@app.route('/api/reports/schedule-h', methods=['GET'])
@jwt_required()
def get_schedule_h_report():
    import json
    sid = int(get_jwt_identity())
    date_str = request.args.get('date', '')
    from_date = request.args.get('from', '')
    to_date = request.args.get('to', '')
    
    if not date_str and not from_date and not to_date:
        return jsonify({'error':'Date is required'}), 400
        
    query = Bill.query.filter(Bill.shop_id == sid)
    if date_str:
        query = query.filter(db.or_(
            Bill.custom_date.like(f"{date_str}%"),
            db.cast(Bill.bill_date, db.String).like(f"{date_str}%")
        ))
    else:
        if from_date:
            query = query.filter(db.or_(
                Bill.custom_date >= from_date,
                db.cast(Bill.bill_date, db.String) >= from_date
            ))
        if to_date:
            query = query.filter(db.or_(
                Bill.custom_date <= to_date,
                db.cast(Bill.bill_date, db.String) <= f"{to_date} 23:59:59"
            ))
            
    bills = query.order_by(Bill.id.asc()).all()

    report_items = []
    # Cache patient records for this shop to avoid repeated queries in loop
    patients = Patient.query.filter_by(shop_id=sid).all()
    patient_address_map = {p.name.strip().lower(): (p.address or '') for p in patients if p.name}
    patient_phone_map = {p.name.strip().lower(): (p.phone or '') for p in patients if p.name}

    # Build a lookup of company_name from the Medicine table for this shop
    all_meds = Medicine.query.filter_by(shop_id=sid).all()
    # Lookup by id
    med_company_by_id = {m.id: (m.company_name or '') for m in all_meds}
    # Fallback lookup by name (lowercase)
    med_company_by_name = {}
    for m in all_meds:
        if m.name:
            key = m.name.strip().lower()
            if key not in med_company_by_name or m.company_name:
                med_company_by_name[key] = (m.company_name or '')

    for b in bills:
        try:
            items = json.loads(b.items_json or '[]')
        except:
            items = []
            
        p_name_lower = (b.customer_name or '').strip().lower()
        p_phone = b.customer_phone or patient_phone_map.get(p_name_lower, '')
        p_address = patient_address_map.get(p_name_lower, '')
            
        for it in items:
            name = it.get('name', '')
            ntype = 'TAB'
            ln = name.upper()
            if 'SYP' in ln or 'SYRUP' in ln or 'SUSP' in ln or 'ML' in ln: ntype = 'SYP'
            elif 'INJ' in ln: ntype = 'INJ'
            elif 'DROP' in ln: ntype = 'DRP'
            elif 'OINT' in ln or 'CREAM' in ln or 'GEL' in ln: ntype = 'OINT'
            elif 'CAP' in ln: ntype = 'CAP'
            
            # Resolve company/manufacturer from item data or Medicine table
            company = it.get('company_name') or it.get('mfg_by') or it.get('mfg') or ''
            if not company:
                med_id = it.get('id')
                if med_id and med_id in med_company_by_id:
                    company = med_company_by_id[med_id]
                if not company:
                    company = med_company_by_name.get(name.strip().lower(), '')
            
            b_date = date_str if date_str else (b.custom_date if b.custom_date else (b.bill_date.strftime('%Y-%m-%d') if b.bill_date else ''))
            report_items.append({
                'bill_no': b.bill_number,
                'date': b_date,
                'doctor': b.doctor_name or '',
                'patient': b.customer_name or '',
                'phone': p_phone,
                'address': p_address,
                'medicine': name,
                'type': ntype,
                'qty': it.get('qty', 0),
                'batch': it.get('batch') or it.get('batch_no') or '',
                'mfg': company,
                'expiry': it.get('expiry') or it.get('expiry_date') or ''
            })
    return jsonify(report_items)

# ══════════════════════════════════════════════════════════════════════
#  AUTO BILLING  (Hidden feature — triple-click on Billing nav)
# ══════════════════════════════════════════════════════════════════════
@app.route('/api/auto-bill', methods=['POST'])
@jwt_required()
def auto_bill():
    """
    Hidden auto-billing feature.
    Generates N bills automatically using random medicines from stock.
    Each bill is saved to DB with a custom date and sequential bill numbers.

    Request JSON:
      { "count": 32, "date": "2024-06-15" }

    Each generated bill:
      - Picks 1–4 random in-stock medicines
      - Random tablet qty (1–3 strips) per medicine
      - Customer name: "Patient-<random 3 digits>"
      - Calculates subtotal, GST (CGST+SGST), discount (0–10%), total
      - Saves to DB and deducts stock
    """
    import json, random as rnd
    try:
        shop  = Shop.query.get(int(get_jwt_identity()))
        d     = request.json
        count = int(d.get('count', 1))
        date_str = d.get('date', '').strip()
        max_amount = float(d.get('max_amount', 0))

        if count < 1 or count > 500:
            return jsonify({'error': 'Count must be between 1 and 500'}), 400

        # Validate / parse date
        if date_str:
            try:
                bill_dt = datetime.strptime(date_str, '%Y-%m-%d')
            except ValueError:
                return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400
        else:
            bill_dt = datetime.now()

        # Load all in-stock medicines for this shop
        meds = Medicine.query.filter(
            Medicine.shop_id == shop.id,
            Medicine.quantity > 0
        ).all()
        doctors = Doctor.query.filter_by(shop_id=shop.id).all()
        patients = Patient.query.filter_by(shop_id=shop.id).all()

        if not meds:
            return jsonify({'error': 'No medicines in stock. Please add stock first.'}), 400

        curr_no   = int(shop.bill_number)
        bills_saved = 0
        errors    = []
        generated_bills_list = []

        for i in range(count):
            try:
                # Pick 1–4 random medicines (or fewer if stock is limited)
                pick_count = min(rnd.randint(1, 4), len(meds))
                chosen     = rnd.sample(meds, pick_count)

                items      = []
                subtotal   = 0.0
                total_gst  = 0.0

                # Vary the time slightly so bills are distinguishable
                minutes_offset = rnd.randint(0, 59)
                hours_offset   = rnd.randint(0, 10)
                this_dt = bill_dt.replace(
                    hour   = min(8 + hours_offset, 20),
                    minute = minutes_offset,
                    second = rnd.randint(0, 59)
                )

                for med in chosen:
                    qty      = rnd.randint(1, 3)        # 1–3 strips
                    actual_q = min(qty, med.quantity)   # don't exceed stock
                    if actual_q < 1:
                        continue

                    price    = med.mrp if med.mrp else med.price
                    gst_pct  = med.gst or 0.0
                    item_sub = round(price * actual_q, 2)
                    
                    if max_amount > 0 and (subtotal + item_sub) > max_amount:
                        break
                    
                    item_gst = round(item_sub * gst_pct / 100, 2)

                    items.append({
                        'id':          med.id,
                        'name':        med.name,
                        'batch':       med.batch or '',
                        'expiry_date': med.expiry_date or '',
                        'qty':         actual_q,
                        'price':       price,
                        'mrp':         med.mrp or price,
                        'gst':         gst_pct,
                        'pack_size':   med.pack_size or '10',
                        'total_tablets': actual_q * (int(''.join(filter(str.isdigit, str(med.pack_size))) or 1) if med.pack_size else 10),
                        'subtotal':    item_sub,
                        'gst_amount':  item_gst
                    })
                    subtotal   += item_sub
                    total_gst  += item_gst

                    # Deduct stock
                    med.quantity = max(0, med.quantity - actual_q)

                if not items:
                    continue  # skip if no items could be added

                discount_pct = rnd.choice([0, 0, 0, 2, 5, 10])  # mostly no discount
                discount_amt = round(subtotal * discount_pct / 100, 2)
                cgst         = round(total_gst / 2, 2)
                sgst         = round(total_gst / 2, 2)
                total        = round(subtotal + cgst + sgst - discount_amt, 2)

                doc_name = rnd.choice(doctors).name if doctors else ''
                
                if patients:
                    pat = rnd.choice(patients)
                    cust_name = pat.name
                    cust_phone = pat.phone
                else:
                    cust_num = rnd.randint(100, 999)
                    cust_name = f'Patient-{cust_num}'
                    cust_phone = ''

                new_bill = Bill(
                    shop_id        = shop.id,
                    bill_number    = str(curr_no),
                    customer_name  = cust_name,
                    customer_phone = cust_phone,
                    doctor_name    = doc_name,
                    subtotal       = round(subtotal, 2),
                    cgst           = cgst,
                    sgst           = sgst,
                    discount       = discount_amt,
                    total_amount   = total,
                    bill_date      = this_dt,
                    custom_date    = date_str,
                    items_json     = json.dumps(items)
                )
                db.session.add(new_bill)
                generated_bills_list.append(new_bill)
                curr_no    += 1
                bills_saved += 1

            except Exception as item_err:
                errors.append(f'Bill {i+1}: {str(item_err)}')
                continue

        # Update shop bill number counter
        shop.bill_number = str(curr_no)
        db.session.commit()
        backup_bg()

        bill_data = [{
            'id': b.id,
            'bill_number': b.bill_number,
            'items': json.loads(b.items_json),
            'customer_name': b.customer_name,
            'customer_phone': b.customer_phone,
            'doctor_name': b.doctor_name,
            'subtotal': b.subtotal,
            'cgst': b.cgst,
            'sgst': b.sgst,
            'discount': b.discount,
            'total_amount': b.total_amount,
            'bill_date': b.bill_date.isoformat(),
            'custom_date': b.custom_date
        } for b in generated_bills_list]

        result = {
            'message':     f'✅ {bills_saved} bills generated successfully!',
            'bills_saved': bills_saved,
            'date_used':   date_str or datetime.now().strftime('%Y-%m-%d'),
            'next_bill_no': curr_no,
            'bills': bill_data
        }
        if errors:
            result['warnings'] = errors[:10]  # cap at 10 error messages
        return jsonify(result), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


# ══════════════════════════════════════════════════════════════════════
#  BACKUP APIs
# ══════════════════════════════════════════════════════════════════════
@app.route('/api/backup', methods=['POST'])
@jwt_required()
def manual_backup():
    ok,res=create_excel_backup()
    if ok: return jsonify({'message':'Backup created!','path':res,'time':datetime.now().strftime('%Y-%m-%d %H:%M:%S')})
    return jsonify({'error':res}),500

@app.route('/api/backup-info', methods=['GET'])
@jwt_required()
def backup_info():
    ex=os.path.exists(BACKUP_PATH)
    return jsonify({'backup_path':BACKUP_PATH,'data_folder':DATA_FOLDER,'db_path':DB_PATH,
        'backup_exists':ex,
        'file_size_kb':round(os.path.getsize(BACKUP_PATH)/1024,1) if ex else 0,
        'last_backup':datetime.fromtimestamp(os.path.getmtime(BACKUP_PATH)).strftime('%Y-%m-%d %H:%M:%S') if ex else 'Never'})

# ══════════════════════════════════════════════════════════════════════
#  ADMIN
# ══════════════════════════════════════════════════════════════════════
@app.route('/api/admin/approve-shop', methods=['POST'])
def approve_shop():
    data = request.json
    if data.get('password') != ADMIN_PASSWORD:
        return jsonify({'error': 'Unauthorized'}), 401
    shop = Shop.query.get(data.get('shop_id'))
    if not shop:
        return jsonify({'error': 'Shop not found'}), 404
    shop.approved = True
    db.session.commit()
    print(f"[SANJANA] ✅ Admin APPROVED: {shop.shop_name} ({shop.email})")
    return jsonify({'message': f'{shop.shop_name} has been approved!'})

@app.route('/api/admin/reject-shop', methods=['POST'])
def reject_shop():
    data = request.json
    if data.get('password') != ADMIN_PASSWORD:
        return jsonify({'error': 'Unauthorized'}), 401
    shop = Shop.query.get(data.get('shop_id'))
    if not shop:
        return jsonify({'error': 'Shop not found'}), 404
    shop_name = shop.shop_name
    shop_email = shop.email
    db.session.delete(shop)
    db.session.commit()
    print(f"[SANJANA] ❌ Admin REJECTED: {shop_name} ({shop_email})")
    return jsonify({'message': f'{shop_name} has been rejected and removed.'})

@app.route('/api/admin/approve-license-renewal', methods=['POST'])
def approve_license_renewal():
    data = request.json or {}
    if data.get('password') != ADMIN_PASSWORD:
        return jsonify({'error': 'Unauthorized'}), 401
    
    req_id = data.get('request_id')
    shop_id = data.get('shop_id')
    
    req = None
    if req_id:
        req = LicenseRenewalRequest.query.get(req_id)
    elif shop_id:
        req = LicenseRenewalRequest.query.filter_by(shop_id=shop_id, status='pending').first()
        
    if not req:
        return jsonify({'error': 'Renewal request not found'}), 404
        
    shop = Shop.query.get(req.shop_id)
    if not shop:
        return jsonify({'error': 'Shop not found'}), 404
        
    now = datetime.now()
    current_end = shop.license_end or now
    new_end = current_end.replace(year=current_end.year+1) if current_end >= now else now.replace(year=now.year+1)
    
    shop.license_end = new_end
    req.status = 'approved'
    req.processed_at = now
    
    db.session.commit()
    
    # Send Approval Email to User
    user_subject = "✅ License Renewal Approved - Sanjana Software"
    user_text = f"Dear {shop.owner_name},\n\nYour license renewal for {shop.shop_name} has been APPROVED by Admin!\nYour new license expiry date is: {new_end.strftime('%Y-%m-%d')}.\n\nThank you for choosing Sanjana Software!"
    user_html = f"""<div style="font-family:Arial,sans-serif;max-width:500px;margin:0 auto;padding:20px;border:1px solid #c8e6c9;border-radius:10px;background:#f1f8e9">
  <h2 style="color:#2e7d32">🎉 License Renewal Approved!</h2>
  <p>Dear <b>{shop.owner_name}</b>,</p>
  <p>Great news! Your license renewal request for <b>{shop.shop_name}</b> has been <b>APPROVED</b> by Admin.</p>
  <div style="background:#e8f5e9;border:1px solid #a5d6a7;padding:15px;border-radius:8px;margin:15px 0">
    <p style="margin:0;font-size:14px;color:#1b5e20"><b>New License Expiry Date:</b> {new_end.strftime('%Y-%m-%d')}</p>
  </div>
  <p style="font-size:13px;color:#555">Thank you for using Sanjana Software.</p>
  <p style="font-size:12px;color:#888">Sanjana Software Team</p>
</div>"""
    threading.Thread(target=send_email_notification, args=(shop.email, user_subject, user_text, user_html)).start()
    
    print(f"[SANJANA] ✅ Admin APPROVED License Renewal for: {shop.shop_name} ({shop.email})")
    return jsonify({'message': f'License renewed for {shop.shop_name} until {new_end.strftime("%Y-%m-%d")}!'}), 200

@app.route('/api/admin/reject-license-renewal', methods=['POST'])
def reject_license_renewal():
    data = request.json or {}
    if data.get('password') != ADMIN_PASSWORD:
        return jsonify({'error': 'Unauthorized'}), 401
    
    req_id = data.get('request_id')
    shop_id = data.get('shop_id')
    
    req = None
    if req_id:
        req = LicenseRenewalRequest.query.get(req_id)
    elif shop_id:
        req = LicenseRenewalRequest.query.filter_by(shop_id=shop_id, status='pending').first()
        
    if not req:
        return jsonify({'error': 'Renewal request not found'}), 404
        
    shop = Shop.query.get(req.shop_id)
    if not shop:
        return jsonify({'error': 'Shop not found'}), 404
        
    now = datetime.now()
    req.status = 'rejected'
    req.processed_at = now
    
    db.session.commit()
    
    # Send Rejection Email to User
    user_subject = "❌ License Renewal Application Rejected - Sanjana Software"
    user_text = (f"Dear {shop.owner_name},\n\n"
                 f"Your license renewal application for {shop.shop_name} has been REJECTED.\n\n"
                 f"Please make payment and send screenshot to sanjanasoftware03@gmail.com with shop name ({shop.shop_name}) and ph.no ({shop.phone}).\n\n"
                 f"Sanjana Software Team")
    
    user_html = f"""<div style="font-family:Arial,sans-serif;max-width:500px;margin:0 auto;padding:20px;border:1px solid #ffcdd2;border-radius:10px;background:#fff5f5">
  <h2 style="color:#c62828">Sanjana Software - License Renewal Status</h2>
  <p>Dear <b>{shop.owner_name}</b>,</p>
  <p>Your license renewal application for <b>{shop.shop_name}</b> has been <b>REJECTED</b>.</p>
  <div style="background:#ffebee;border-left:4px solid #c62828;padding:15px;margin:15px 0">
    <p style="margin:0;color:#b71c1c;font-size:14px;font-weight:bold">Required Action:</p>
    <p style="margin:6px 0 0 0;color:#333;font-size:13px">
      Please make payment and send screenshot to 
      <a href="mailto:sanjanasoftware03@gmail.com" style="color:#c62828;font-weight:bold">sanjanasoftware03@gmail.com</a> 
      with your shop name (<b>{shop.shop_name}</b>) and ph.no (<b>{shop.phone}</b>).
    </p>
  </div>
  <p style="font-size:12px;color:#777">If you have any questions, please reply to this email or contact support.</p>
  <p style="font-size:12px;color:#777">Sanjana Software Team</p>
</div>"""
    threading.Thread(target=send_email_notification, args=(shop.email, user_subject, user_text, user_html)).start()
    
    print(f"[SANJANA] ❌ Admin REJECTED License Renewal for: {shop.shop_name} ({shop.email})")
    return jsonify({'message': f'Renewal request for {shop.shop_name} rejected. Email sent to user.'}), 200

@app.route('/api/admin/delete-shop', methods=['POST'])
def delete_shop():
    data = request.json
    if data.get('password') != ADMIN_PASSWORD:
        return jsonify({'error': 'Unauthorized'}), 401
    shop = Shop.query.get(data.get('shop_id'))
    if not shop:
        return jsonify({'error': 'Shop not found'}), 404
        
    shop_name = shop.shop_name
    sid = shop.id
    
    # Delete all associated data
    Doctor.query.filter_by(shop_id=sid).delete()
    Patient.query.filter_by(shop_id=sid).delete()
    Supplier.query.filter_by(shop_id=sid).delete()
    PurchaseEntry.query.filter_by(shop_id=sid).delete()
    Medicine.query.filter_by(shop_id=sid).delete()
    Bill.query.filter_by(shop_id=sid).delete()
    LicenseRenewalRequest.query.filter_by(shop_id=sid).delete()
    
    # Delete shop
    db.session.delete(shop)
    db.session.commit()
    
    print(f"[SANJANA] 🗑️ Admin DELETED account and data for: {shop_name}")
    return jsonify({'message': f'"{shop_name}" and ALL associated data (medicines, bills, patients) have been permanently deleted.'})

@app.route('/api/admin/toggle-shop-stop', methods=['POST'])
def toggle_shop_stop():
    data = request.json or {}
    if data.get('password') != ADMIN_PASSWORD:
        return jsonify({'error': 'Unauthorized'}), 401
    
    shop_id = data.get('shop_id')
    shop = Shop.query.get(shop_id)
    if not shop:
        return jsonify({'error': 'Shop not found'}), 404
        
    is_stopped = bool(data.get('is_stopped', True))
    shop.is_stopped = is_stopped
    
    if is_stopped:
        active_sessions.pop(shop.id, None)
        
    db.session.commit()
    backup_bg()
    
    action_str = "STOPPED" if is_stopped else "ALLOWED"
    print(f"[SANJANA] ⚙️ Admin {action_str} shop access for: {shop.shop_name} ({shop.email})")
    return jsonify({
        'message': f'Shop "{shop.shop_name}" access has been {action_str.lower()} successfully.',
        'is_stopped': is_stopped
    })

# ── Remote Access URL endpoint for Admin Panel (Render) ──
@app.route('/api/ngrok-url')
def get_remote_url():
    render_url = CLOUD_ADMIN_URL
    if render_url:
        return jsonify({
            'active': True,
            'admin_url': render_url + '/owner-admin',
            'public_url': render_url
        })
    return jsonify({'active': False})

@app.route('/api/admin/shops', methods=['POST'])
def get_all_shops():
    if request.json.get('password')!=ADMIN_PASSWORD:
        return jsonify({'error':'Unauthorized'}),401
    shops=Shop.query.all()
    result=[]
    today_str = date.today().strftime('%Y-%m-%d')
    for s in shops:
        latest_req = LicenseRenewalRequest.query.filter_by(shop_id=s.id).order_by(LicenseRenewalRequest.id.desc()).first()
        bill_count  = Bill.query.filter_by(shop_id=s.id).count()
        today_bills = Bill.query.filter(
            Bill.shop_id == s.id,
            db.or_(
                Bill.custom_date.like(today_str + '%'),
                db.cast(Bill.bill_date, db.String).like(today_str + '%')
            )
        ).count()
        rev_row = db.session.execute(
            text("SELECT COALESCE(SUM(total_amount),0) FROM bill WHERE shop_id=:sid"),
            {'sid': s.id}
        ).fetchone()
        total_revenue = float(rev_row[0]) if rev_row else 0.0
        today_rev_row = db.session.execute(
            text("SELECT COALESCE(SUM(total_amount),0) FROM bill WHERE shop_id=:sid AND (custom_date LIKE :td OR CAST(bill_date AS TEXT) LIKE :td)"),
            {'sid': s.id, 'td': today_str + '%'}
        ).fetchone()
        today_revenue = float(today_rev_row[0]) if today_rev_row else 0.0
        lic_end = s.license_end
        if lic_end:
            days_left = (lic_end.date() - date.today()).days
        else:
            days_left = 999
        result.append({
            'id':             s.id,
            'shop_name':      s.shop_name,
            'owner_name':     s.owner_name,
            'email':          s.email,
            'phone':          s.phone,
            'address':        s.address,
            'approved':       s.approved if s.approved is not None else True,
            'is_stopped':     bool(s.is_stopped if hasattr(s, 'is_stopped') and s.is_stopped is not None else False),
            'med_count':      Medicine.query.filter_by(shop_id=s.id).count(),
            'bill_count':     bill_count,
            'today_bills':    today_bills,
            'total_revenue':  round(total_revenue, 2),
            'today_revenue':  round(today_revenue, 2),
            'joined_on':      s.created_at.strftime('%Y-%m-%d %H:%M') if s.created_at else 'N/A',
            'license_end':    lic_end.strftime('%Y-%m-%d') if lic_end else 'N/A',
            'license_days':   days_left,
            'renewal_status': latest_req.status if latest_req else None,
            'is_active':      s.id in active_sessions,
            'login_time':     active_sessions.get(s.id,{}).get('login_time','-'),
            'last_online':    s.last_online.strftime('%Y-%m-%d %H:%M:%S') if s.last_online else 'Never',
            'last_online_ago': _time_ago(s.last_online)
        })
        
    pending_reqs = LicenseRenewalRequest.query.filter_by(status='pending').order_by(LicenseRenewalRequest.id.desc()).all()
    pending_renewals = []
    for req in pending_reqs:
        s = Shop.query.get(req.shop_id)
        if s:
            pending_renewals.append({
                'request_id':   req.id,
                'shop_id':      s.id,
                'shop_name':    s.shop_name,
                'owner_name':   s.owner_name,
                'email':        s.email,
                'phone':        s.phone,
                'license_end':  s.license_end.strftime('%Y-%m-%d') if s.license_end else 'N/A',
                'requested_at': req.requested_at.strftime('%Y-%m-%d %H:%M') if req.requested_at else 'N/A'
            })

    total_bills_all   = db.session.execute(text("SELECT COUNT(*) FROM bill")).fetchone()[0]
    today_bills_all   = db.session.execute(
        text("SELECT COUNT(*) FROM bill WHERE custom_date LIKE :td OR CAST(bill_date AS TEXT) LIKE :td"),
        {'td': today_str + '%'}
    ).fetchone()[0]
    total_revenue_all = db.session.execute(text("SELECT COALESCE(SUM(total_amount),0) FROM bill")).fetchone()[0]
    today_revenue_all = db.session.execute(
        text("SELECT COALESCE(SUM(total_amount),0) FROM bill WHERE custom_date LIKE :td OR CAST(bill_date AS TEXT) LIKE :td"),
        {'td': today_str + '%'}
    ).fetchone()[0]
    total_meds_all    = db.session.execute(text("SELECT COUNT(*) FROM medicine")).fetchone()[0]
    expiring_soon     = sum(1 for sh in result if 0 <= sh['license_days'] <= 30)

    return jsonify({
        'shops':              result,
        'active_count':       len(active_sessions),
        'total_count':        len(shops),
        'pending_renewals':   pending_renewals,
        'total_bills_all':    int(total_bills_all),
        'today_bills_all':    int(today_bills_all),
        'total_revenue_all':  round(float(total_revenue_all), 2),
        'today_revenue_all':  round(float(today_revenue_all), 2),
        'total_meds_all':     int(total_meds_all),
        'expiring_soon':      expiring_soon,
        'today_str':          today_str
    })


# ══════════════════════════════════════════════════════════════════════
#  ENTRY POINT  —  browser opens automatically
# ══════════════════════════════════════════════════════════════════════
@app.route('/api/admin/add-sample-data', methods=['POST'])
@jwt_required()
def add_sample_data():
    """Add 30 common Indian medicines as sample data for testing."""
    sid = int(get_jwt_identity())
    try:
        samples = [
            # name, category, batch, qty, price, mrp, gst, expiry, pack_size, company, supplier
            ("Paracetamol 500mg",      "TABLET",    "PCM2024A", 100, 1.50,  2.00,  5.0, "2026-12-31", 10, "Cipla Ltd",       "Cipla Distributor"),
            ("Amoxicillin 500mg",      "CAPSULE",   "AMX2024B", 50,  8.50,  12.00, 12.0,"2026-10-31", 10, "Sun Pharma",      "Sun Distributor"),
            ("Azithromycin 500mg",     "TABLET",    "AZI2024C", 30,  18.00, 24.00, 12.0,"2026-09-30", 3,  "Cipla Ltd",       "Cipla Distributor"),
            ("Metformin 500mg",        "TABLET",    "MET2024D", 200, 2.00,  3.50,  5.0, "2027-03-31", 10, "USV Pharma",      "USV Distributor"),
            ("Atorvastatin 10mg",      "TABLET",    "ATO2024E", 90,  4.50,  7.00,  12.0,"2027-01-31", 10, "Lupin Ltd",       "Lupin Distributor"),
            ("Omeprazole 20mg",        "CAPSULE",   "OMP2024F", 80,  3.00,  5.50,  12.0,"2026-11-30", 10, "Sun Pharma",      "Sun Distributor"),
            ("Cetirizine 10mg",        "TABLET",    "CET2024G", 120, 1.20,  2.00,  5.0, "2027-02-28", 10, "Cadila Health",   "Cadila Distributor"),
            ("Amlodipine 5mg",         "TABLET",    "AML2024H", 150, 2.50,  4.00,  12.0,"2027-04-30", 10, "Cipla Ltd",       "Cipla Distributor"),
            ("Cough Syrup 100ml",      "SYRUP",     "CSY2024I", 40,  55.00, 80.00, 12.0,"2026-08-31", 1,  "Benadryl",        "JnJ Distributor"),
            ("Dolo 650mg",             "TABLET",    "DOL2024J", 200, 2.50,  3.00,  5.0, "2026-12-31", 15, "Micro Labs",      "Micro Distributor"),
            ("Vitamin C 500mg",        "TABLET",    "VTC2024K", 100, 3.00,  5.00,  5.0, "2027-06-30", 10, "Himalaya",        "Himalaya Distributor"),
            ("Vitamin D3 60K",         "CAPSULE",   "VTD2024L", 60,  18.00, 28.00, 12.0,"2027-05-31", 4,  "Sun Pharma",      "Sun Distributor"),
            ("Ranitidine 150mg",       "TABLET",    "RAN2024M", 80,  1.50,  2.50,  5.0, "2026-07-31", 10, "Cipla Ltd",       "Cipla Distributor"),
            ("Pantoprazole 40mg",      "TABLET",    "PAN2024N", 90,  5.00,  8.00,  12.0,"2027-01-31", 10, "Alkem Labs",      "Alkem Distributor"),
            ("Metronidazole 400mg",    "TABLET",    "MTZ2024O", 60,  2.00,  3.50,  5.0, "2026-10-31", 10, "Cipla Ltd",       "Cipla Distributor"),
            ("Ibuprofen 400mg",        "TABLET",    "IBU2024P", 100, 2.50,  4.00,  5.0, "2026-09-30", 10, "Abbott India",    "Abbott Distributor"),
            ("Betadine Ointment 20g",  "CREAM",     "BET2024Q", 30,  40.00, 65.00, 12.0,"2027-03-31", 1,  "Win Medicare",    "Win Distributor"),
            ("Digene Syrup 200ml",     "SYRUP",     "DIG2024R", 25,  90.00, 130.00,12.0,"2026-11-30", 1,  "Abbott India",    "Abbott Distributor"),
            ("Combiflam Tablet",       "TABLET",    "CMB2024S", 80,  8.00,  12.00, 12.0,"2026-12-31", 10, "Sanofi India",    "Sanofi Distributor"),
            ("Aspirin 75mg",           "TABLET",    "ASP2024T", 200, 0.80,  1.50,  5.0, "2027-02-28", 14, "Bayer India",     "Bayer Distributor"),
            ("Clonazepam 0.5mg",       "TABLET",    "CLN2024U", 30,  3.50,  6.00,  12.0,"2026-08-31", 10, "Sun Pharma",      "Sun Distributor"),
            ("Montelukast 10mg",       "TABLET",    "MON2024V", 50,  6.00,  10.00, 12.0,"2027-01-31", 10, "Lupin Ltd",       "Lupin Distributor"),
            ("Levocetirizine 5mg",     "TABLET",    "LEV2024W", 90,  2.50,  4.50,  5.0, "2027-04-30", 10, "USV Pharma",      "USV Distributor"),
            ("Glimepiride 1mg",        "TABLET",    "GLM2024X", 60,  4.00,  7.00,  12.0,"2026-10-31", 10, "Sanofi India",    "Sanofi Distributor"),
            ("Telmisartan 40mg",       "TABLET",    "TEL2024Y", 90,  5.50,  9.00,  12.0,"2027-03-31", 10, "Cipla Ltd",       "Cipla Distributor"),
            ("Albendazole 400mg",      "TABLET",    "ALB2024Z", 40,  8.00,  14.00, 12.0,"2026-09-30", 1,  "GSK India",       "GSK Distributor"),
            ("Lactulose Syrup 100ml",  "SYRUP",     "LAC2024A", 20,  120.00,175.00,12.0,"2026-07-31", 1,  "Abbott India",    "Abbott Distributor"),
            ("Salbutamol Inhaler",     "INJECTION", "SAL2024B", 15,  95.00, 150.00,12.0,"2026-11-30", 1,  "Cipla Ltd",       "Cipla Distributor"),
            ("Betamethasone Cream",    "CREAM",     "BCR2024C", 25,  35.00, 58.00, 12.0,"2027-02-28", 1,  "GSK India",       "GSK Distributor"),
            ("Iron + Folic Acid",      "TABLET",    "IFA2024D", 150, 1.00,  2.00,  5.0, "2027-05-31", 10, "Himalaya",        "Himalaya Distributor"),
        ]
        added = 0
        skipped = 0
        for s in samples:
            name, cat, batch, qty, price, mrp, gst, expiry, pack_size, company, supplier = s
            if Medicine.query.filter_by(shop_id=sid, name=name).first():
                skipped += 1
                continue
            db.session.add(Medicine(
                shop_id=sid, name=name, category=cat, batch=batch,
                quantity=qty, price=price, mrp=mrp, gst=gst,
                expiry_date=expiry, pack_size=pack_size,
                company_name=company, supplier_name=supplier,
                sale_discount=0
            ))
            added += 1
        db.session.commit()
        backup_bg()
        return jsonify({
            'message': f'✅ Added {added} sample medicines! ({skipped} already existed)',
            'added': added, 'skipped': skipped
        }), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


# ══════════════════════════════════════════════════════════════════════
#  RAW TEXT PRINTING (DOT MATRIX / IMPACT PRINTERS)
#  Sends plain text directly to printer — bypasses browser graphics
# ══════════════════════════════════════════════════════════════════════

def get_default_printer():
    """Get the Windows default printer name using multiple fallback methods."""
    # Method 1: ctypes GetDefaultPrinterW (most reliable when it works)
    try:
        buf = ctypes.create_unicode_buffer(256)
        bufsize = ctypes.wintypes.DWORD(256)
        result = ctypes.windll.winspool.GetDefaultPrinterW(buf, ctypes.byref(bufsize))
        if result and buf.value:
            print(f"[PRINTER] Default printer (ctypes): {buf.value}")
            return buf.value
    except Exception as e:
        print(f"[PRINTER] ctypes method failed: {e}")

    # Method 2: PowerShell Get-CimInstance (works even when ctypes fails)
    try:
        import subprocess
        result = subprocess.run(
            ['powershell', '-NoProfile', '-Command',
             '(Get-CimInstance -ClassName Win32_Printer -Filter "Default=True").Name'],
            capture_output=True, text=True, timeout=10,
            creationflags=0x08000000
        )
        name = result.stdout.strip()
        if name:
            print(f"[PRINTER] Default printer (WMI): {name}")
            return name
    except Exception as e:
        print(f"[PRINTER] WMI method failed: {e}")

    # Method 3: Read from registry (fastest fallback)
    try:
        import winreg
        key = winreg.OpenKey(
            winreg.HKEY_CURRENT_USER,
            r"Software\Microsoft\Windows NT\CurrentVersion\Windows"
        )
        device, _ = winreg.QueryValueEx(key, "Device")
        winreg.CloseKey(key)
        # Value format is "PrinterName,winspool,Ne00:"
        name = device.split(",")[0]
        if name:
            print(f"[PRINTER] Default printer (registry): {name}")
            return name
    except Exception as e:
        print(f"[PRINTER] Registry method failed: {e}")

    # Method 4: If nothing worked, pick the first available printer
    try:
        printers = list_printers()
        if printers:
            print(f"[PRINTER] No default found, using first available: {printers[0]}")
            return printers[0]
    except:
        pass

    print("[PRINTER] ERROR: No printer found by any method!")
    return None

def list_printers():
    """List all available printers on Windows."""
    # Method 1: PowerShell Get-Printer cmdlet
    try:
        import subprocess
        result = subprocess.run(
            ['powershell', '-NoProfile', '-Command',
             'Get-Printer | Select-Object -ExpandProperty Name'],
            capture_output=True, text=True, timeout=10,
            creationflags=0x08000000
        )
        printers = [p.strip() for p in result.stdout.strip().split('\n') if p.strip()]
        if printers:
            return printers
    except:
        pass

    # Method 2: WMI fallback (works on older Windows / restricted environments)
    try:
        import subprocess
        result = subprocess.run(
            ['powershell', '-NoProfile', '-Command',
             '(Get-CimInstance -ClassName Win32_Printer).Name'],
            capture_output=True, text=True, timeout=10,
            creationflags=0x08000000
        )
        printers = [p.strip() for p in result.stdout.strip().split('\n') if p.strip()]
        if printers:
            return printers
    except:
        pass

    return []

def raw_print_to_printer(text, printer_name=None, cut=False, condensed=False):
    """Send raw text directly to a printer using Windows winspool API.
    This prints as TEXT characters (fast) instead of as a graphical bitmap (slow).
    """
    try:
        winspool = ctypes.WinDLL('winspool.drv')

        # Get default printer if not specified
        if not printer_name:
            printer_name = get_default_printer()
        print(f"[PRINTER] Attempting to print to: '{printer_name}'")
        if not printer_name:
            available = list_printers()
            return False, f"No printer found. Available printers: {available if available else 'None detected'}"

        # Open printer
        hPrinter = ctypes.wintypes.HANDLE()
        if not winspool.OpenPrinterW(printer_name, ctypes.byref(hPrinter), None):
            error_code = ctypes.GetLastError()
            available = list_printers()
            return False, f"Cannot open printer: {printer_name} (error {error_code}). Available: {available}"

        # DOC_INFO_1 structure
        class DOC_INFO_1(ctypes.Structure):
            _fields_ = [
                ("pDocName",    ctypes.wintypes.LPWSTR),
                ("pOutputFile", ctypes.wintypes.LPWSTR),
                ("pDatatype",   ctypes.wintypes.LPWSTR),
            ]

        doc_info = DOC_INFO_1()
        doc_info.pDocName    = "SANJANA BILL"
        doc_info.pOutputFile = None
        doc_info.pDatatype   = "RAW"  # RAW = send bytes directly, no graphics rendering

        # ── STEP 1: Normalise line endings ──────────────────────────────────────
        # Convert every newline variant to bare \n first, then to CRLF (dot-matrix
        # standard).  Also STRIP every Form Feed (\f) character — a stray \f tells
        # the printer to eject to the next page boundary, wasting a full page of
        # tractor paper.
        text = text.replace('\r\n', '\n').replace('\r', '\n')
        text = text.replace('\f', '')          # FIX: remove all form-feed characters
        text = text.replace('\n', '\r\n')      # convert to CRLF for dot-matrix

        # ── STEP 2: Strip trailing whitespace / blank lines ──────────────────────
        # rstrip() removes trailing CRLF pairs so the printer does not feed
        # blank lines after the last printed line.
        text = text.rstrip()

        # ── STEP 3: Detect multi-page vs single-page job ─────────────────────────
        # Standard 11-inch tractor paper at 6 LPI = 66 lines per page.
        # (Use 72 if your paper is 12-inch / 14-inch.)
        LINES_PER_PAGE = 66

        total_lines = len(text.split('\r\n'))
        is_multi_page = total_lines > LINES_PER_PAGE

        if is_multi_page:
            # Multi-page continuous job:
            # ESC C 0 = use the hardware page length (set via printer DIP switches).
            # Combined with ESC O (no skip-over-perf) the printer rolls continuously
            # across perforations without jumping or wasting paper.
            page_len_cmd = b'\x1B\x43\x00'   # ESC C 0 — hardware page length
        else:
            # Single-page bill:
            # ESC C n = exact line count so the printer stops feeding after the
            # last line instead of completing a full 66-line physical page.
            n = max(1, min(total_lines + 1, 127))  # +1 = one blank line for tear-off
            page_len_cmd = b'\x1B\x43' + bytes([n])  # ESC C n — exact page size

        # ── STEP 4: Add a single trailing blank line (tear-off gap) ──────────────
        text += '\r\n'

        # ── STEP 5: Build ESC/P control prefix ───────────────────────────────────
        esc_prefix  = b'\x1B\x40'          # ESC @  — hard reset / initialise printer
        if condensed:
            esc_prefix += b'\x0F'          # SI     — condensed mode (17 CPI)
        else:
            esc_prefix += b'\x1B\x4D'     # ESC M  — elite 12 CPI
        esc_prefix += b'\x1B\x4F'         # ESC O  — cancel skip-over-perforation
        esc_prefix += page_len_cmd         # ESC C n or ESC C 0 — page length
        
        # ── STEP 6: Start print job ───────────────────────────────────────────────
        # NOTE: StartPagePrinter / EndPagePrinter intentionally skipped.
        # On many dot-matrix drivers EndPagePrinter emits a Form Feed which ejects
        # to the next page boundary — the exact waste we are eliminating.
        # Flow: StartDocPrinter → WritePrinter → EndDocPrinter  (no page calls).
        job_id = winspool.StartDocPrinterW(hPrinter, 1, ctypes.byref(doc_info))
        if not job_id:
            winspool.ClosePrinter(hPrinter)
            return False, "StartDocPrinter failed"

        # ── STEP 7: Encode and send bytes ────────────────────────────────────────
        # cp437 is the native IBM/DOS code page for dot-matrix printers.
        # Characters outside cp437 are replaced with '?' to avoid garbage output.
        try:
            text_bytes = text.encode('cp437', errors='replace')
        except Exception:
            text_bytes = text.encode('utf-8', errors='replace')

        data = esc_prefix + text_bytes

        written = ctypes.wintypes.DWORD()
        winspool.WritePrinter(hPrinter, data, len(data), ctypes.byref(written))

        # ── STEP 8: End job — NO EndPagePrinter (avoids form-feed) ───────────────
        winspool.EndDocPrinter(hPrinter)
        winspool.ClosePrinter(hPrinter)

        return True, f"Printed OK ({written.value} bytes to {printer_name})"

    except Exception as e:
        return False, str(e)


@app.route('/api/printers', methods=['GET'])
@jwt_required()
def api_list_printers():
    """List all available printers."""
    printers = list_printers()
    default  = get_default_printer()
    return jsonify({'printers': printers, 'default': default})


@app.route('/api/raw-print', methods=['POST'])
@jwt_required()
def api_raw_print():
    """Send raw text to printer — used for fast dot matrix printing."""
    data = request.json
    text_content = data.get('text', '')
    printer_name = data.get('printer', None)  # optional, uses default if empty
    cut = bool(data.get('cut', False))

    if not text_content:
        return jsonify({'error': 'No text to print'}), 400

    success, msg = raw_print_to_printer(text_content, printer_name, cut=cut, condensed=data.get('condensed', False))

    if success:
        return jsonify({'message': msg, 'success': True})
    else:
        # Return exact error so the user knows RAW failed
        return jsonify({'error': msg, 'success': False}), 500
# ══════════════════════════════════════════════════════════════════════
#  OCR BILL EXTRACTION ENDPOINT
# ══════════════════════════════════════════════════════════════════════
def get_gemini_api_key():
    import os
    # 1. Check environment variable
    key = os.environ.get("GEMINI_API_KEY") or os.environ.get("GOOGLE_API_KEY")
    if key:
        return key
    
    # 2. Check root .env file
    root_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    root_env = os.path.join(root_dir, ".env")
    if os.path.exists(root_env):
        try:
            with open(root_env, "r", encoding="utf-8") as f:
                for line in f:
                    if line.strip().startswith("GEMINI_API_KEY="):
                        return line.split("=", 1)[1].strip().strip('"').strip("'")
                    if line.strip().startswith("GOOGLE_API_KEY="):
                        return line.split("=", 1)[1].strip().strip('"').strip("'")
        except Exception:
            pass
            
    # 3. Check current dir .env file
    local_env = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
    if os.path.exists(local_env):
        try:
            with open(local_env, "r", encoding="utf-8") as f:
                for line in f:
                    if line.strip().startswith("GEMINI_API_KEY="):
                        return line.split("=", 1)[1].strip().strip('"').strip("'")
                    if line.strip().startswith("GOOGLE_API_KEY="):
                        return line.split("=", 1)[1].strip().strip('"').strip("'")
        except Exception:
            pass
            
    return None

def extract_bill_with_gemini(file_bytes, mime_type, api_key):
    import base64, urllib.request, json
    base64_data = base64.b64encode(file_bytes).decode("utf-8")
    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={api_key}"
    
    prompt = """
    Analyze this purchase bill / invoice and extract details. 
    You must return a JSON object with the following fields:
    - supplier: Name of the supplier / company selling the medicines.
    - invoice_no: The invoice or bill number.
    - date: The invoice date in DD/MM/YYYY format.
    - items: A list of objects representing the items in the bill. Each item must have:
      * product_name: The name of the product / medicine.
      * qty: The quantity purchased (integer).
      * rate: The purchase rate per unit or strip (float).
      * amount: The total amount for this item (float).
      * batch: The batch number if visible, otherwise leave blank or guess a reasonable format.
      * expiry: The expiry date in MM/YYYY format if visible.
      * mrp: The MRP per unit/strip (float) if visible.
      * free: The free quantity (integer) if visible, otherwise 0.
      * dis: The discount percentage (float) if visible, otherwise 0.
      * gst: The GST percentage (float) if visible, otherwise 12.0 or 18.0 (guess from the bill if possible).
      * pack: The pack size (e.g. "10", "15", "1") if visible.
      * category: The category of the item, must be one of: TABLET, SYRUP, INJECTION, CREAM, CAPSULE, DROPS, GENERAL.

    Return ONLY the raw JSON object. Do not include any markdown formatting like ```json or ```.
    """
    
    payload = {
        "contents": [
            {
                "parts": [
                    {
                        "inlineData": {
                            "mimeType": mime_type,
                            "data": base64_data
                        }
                    },
                    {
                        "text": prompt
                    }
                ]
            }
        ],
        "generationConfig": {
            "responseMimeType": "application/json"
        }
    }
    
    req_data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(
        url,
        data=req_data,
        headers={"Content-Type": "application/json"},
        method="POST"
    )
    
    try:
        with urllib.request.urlopen(req, timeout=30) as response:
            res_body = response.read().decode("utf-8")
            res_json = json.loads(res_body)
            text_response = res_json["candidates"][0]["content"]["parts"][0]["text"]
            return json.loads(text_response.strip())
    except Exception as e:
        print("[Gemini API Error]", e)
        return None

def parse_bill_text_heuristics(text):
    import re, random
    from datetime import datetime
    supplier = "Unknown Supplier"
    invoice_no = f"INV-{random.randint(1000, 9999)}"
    date_str = datetime.now().strftime('%d/%m/%Y')
    items = []
    
    lines = [line.strip() for line in text.split('\n') if line.strip()]
    
    supplier_found = False
    for line in lines[:8]:
        lower_line = line.lower()
        if any(keyword in lower_line for keyword in ["distributor", "agency", "agencies", "pharmacy", "pharmaceutical", "ltd", "limited", "pvt", "corp", "medical", "healthcare", "labs", "laboratories"]):
            supplier = line.split("Supplier:")[-1].split("Name:")[-1].strip()
            supplier_found = True
            break
    if not supplier_found and lines:
        supplier = lines[0]
        
    inv_match = re.search(r'(?:invoice\s*no|inv\s*no|bill\s*no|invoice|inv|bill)[:.\-\s#]+([A-Za-z0-9\-]+)', text, re.IGNORECASE)
    if inv_match:
        invoice_no = inv_match.group(1)
        
    date_match = re.search(r'(?:date|dt)[:.\-\s]+(\d{1,2}[-./]\d{1,2}[-./]\d{2,4})', text, re.IGNORECASE)
    if date_match:
        date_str = date_match.group(1)
    else:
        any_date = re.search(r'(\d{1,2}[-./]\d{1,2}[-./]\d{2,4})', text)
        if any_date:
            date_str = any_date.group(1)
            
    for line in lines:
        if any(h in line.lower() for h in ["invoice", "supplier", "date", "total", "tax", "cgst", "sgst", "gst", "subtotal", "amount", "rate", "quantity"]):
            continue
            
        parts = line.split()
        if len(parts) >= 3:
            try:
                last_val = float(parts[-1].replace(',', ''))
                sec_last_val = float(parts[-2].replace(',', ''))
                
                qty = 1
                rate = sec_last_val
                amount = last_val
                
                if len(parts) >= 4:
                    try:
                        third_last_val = float(parts[-3].replace(',', ''))
                        if third_last_val.is_integer():
                            qty = int(third_last_val)
                        else:
                            qty = int(round(amount / rate)) if rate > 0 else 1
                    except ValueError:
                        pass
                
                name_parts = []
                for p in parts[:-2]:
                    try:
                        float(p.replace(',', ''))
                        if len(name_parts) >= 1:
                            break
                    except ValueError:
                        name_parts.append(p)
                
                product_name = " ".join(name_parts)
                if not product_name:
                    product_name = "Item " + parts[0]
                    
                if qty > 0 and rate >= 0 and amount >= 0 and len(product_name) > 2:
                    items.append({
                        "product_name": product_name,
                        "name": product_name,
                        "qty": qty,
                        "rate": rate,
                        "amount": amount,
                        "batch": "GEN_BATCH",
                        "expiry": "12/2029",
                        "mrp": round(rate * 1.2, 2),
                        "free": 0,
                        "dis": 0.0,
                        "gst": 12.0,
                        "pack": "10",
                        "category": "TABLET"
                    })
            except (ValueError, IndexError):
                continue
                
    return {
        "supplier": supplier,
        "supplier_name": supplier,
        "invoice_no": invoice_no,
        "date": date_str,
        "items": items
    }

@app.route('/api/extract_bill', methods=['POST'])
@jwt_required()
def extract_bill():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    file_bytes = file.read()
    filename_lower = file.filename.lower()
    
    # 1. Determine mime type
    mime_type = "application/octet-stream"
    if filename_lower.endswith('.pdf'):
        mime_type = "application/pdf"
    elif filename_lower.endswith(('.png', '.jpg', '.jpeg', '.webp')):
        if filename_lower.endswith('.png'):
            mime_type = "image/png"
        elif filename_lower.endswith(('.jpg', '.jpeg')):
            mime_type = "image/jpeg"
        elif filename_lower.endswith('.webp'):
            mime_type = "image/webp"
        
    # 2. Check for Gemini Key
    api_key = get_gemini_api_key()
    if api_key:
        print("[OCR] Using Gemini API key for extraction...")
        extracted = extract_bill_with_gemini(file_bytes, mime_type, api_key)
        if extracted:
            # Map raw keys to standard keys for frontend compatibility
            supplier = extracted.get("supplier") or extracted.get("supplier_name") or "Unknown Supplier"
            invoice_no = extracted.get("invoice_no") or "AI-INV"
            date_str = extracted.get("date") or datetime.now().strftime('%d/%m/%Y')
            raw_items = extracted.get("items") or []
            
            items = []
            for it in raw_items:
                name = it.get("product_name") or it.get("name") or "Unknown Medicine"
                items.append({
                    "product_name": name,
                    "name": name,
                    "qty": it.get("qty", 1),
                    "rate": it.get("rate", 0.0),
                    "amount": it.get("amount", 0.0),
                    "batch": it.get("batch") or "AI_BATCH",
                    "expiry": it.get("expiry") or "12/2029",
                    "mrp": it.get("mrp") or round(it.get("rate", 0.0) * 1.2, 2),
                    "free": it.get("free", 0),
                    "dis": it.get("dis", 0.0),
                    "gst": it.get("gst", 12.0),
                    "pack": it.get("pack") or "10",
                    "category": it.get("category") or "TABLET"
                })
            return jsonify({
                'success': True,
                'gemini_active': True,
                'supplier': supplier,
                'supplier_name': supplier,
                'invoice_no': invoice_no,
                'date': date_str,
                'items': items
            })
            
    # 3. Fallback to Local Parsers
    print("[OCR] Falling back to local/heuristic extraction...")
    if mime_type == "application/pdf":
        try:
            import io
            from pypdf import PdfReader
            reader = PdfReader(io.BytesIO(file_bytes))
            text = ""
            for page in reader.pages:
                text += page.extract_text() or ""
            print("[OCR] PDF extracted text length:", len(text))
            
            parsed = parse_bill_text_heuristics(text)
            parsed["success"] = True
            parsed["gemini_active"] = False
            return jsonify(parsed)
        except Exception as pdf_err:
            print("[OCR] PDF fallback parsing error:", pdf_err)
            
    # If image or fallback PDF, try pytesseract
    try:
        from PIL import Image
        import pytesseract
        import io
        img = Image.open(io.BytesIO(file_bytes))
        text = pytesseract.image_to_string(img)
        print("[OCR] Extracted text from image using Tesseract:", text[:100], "...")
        
        parsed = parse_bill_text_heuristics(text)
        parsed["success"] = True
        parsed["gemini_active"] = False
        return jsonify(parsed)
    except Exception as img_err:
        print("[OCR] Tesseract image parsing error/not found:", img_err)
        
    # Final fallback: Mock data with warning message
    return jsonify({
        'success': True,
        'gemini_active': False,
        'warning': 'No Gemini API key found, and local OCR is not configured. Showing sample data.',
        'supplier': 'Fallback Medicine Distributor (Local Demo)',
        'supplier_name': 'Fallback Medicine Distributor (Local Demo)',
        'date': datetime.now().strftime('%d/%m/%Y'),
        'invoice_no': f"DEMO-{random.randint(1000, 9999)}",
        'items': [
            {
                'product_name': 'Dolo 650 Tablet',
                'name': 'Dolo 650 Tablet',
                'qty': 100,
                'rate': 2.50,
                'amount': 250.0,
                'batch': 'DL559',
                'expiry': '08/2028',
                'mrp': 3.12,
                'free': 0,
                'dis': 0.0,
                'gst': 12.0,
                'pack': '15',
                'category': 'TABLET'
            },
            {
                'product_name': 'Cough Syrup 100ml',
                'name': 'Cough Syrup 100ml',
                'qty': 50,
                'rate': 45.00,
                'amount': 2250.0,
                'batch': 'CS712',
                'expiry': '11/2027',
                'mrp': 58.00,
                'free': 5,
                'dis': 5.0,
                'gst': 18.0,
                'pack': '1',
                'category': 'SYRUP'
            }
        ]
    })



# ══════════════════════════════════════════════════════════════════════
#  24/7 CLOUD OWNER PANEL SYNC WORKER
# ══════════════════════════════════════════════════════════════════════
# CLOUD_ADMIN_URL is defined at the top of the file

def _sync_to_cloud_admin():
    """Background worker that syncs shop stats to 24/7 Cloud Owner Panel."""
    import urllib.request, json
    while True:
        try:
            time.sleep(15) # Wait after boot, then loop every 60s
            with app.app_context():
                shops = Shop.query.all()
                today_str = date.today().strftime('%Y-%m-%d')
                for s in shops:
                    pending_renewal = LicenseRenewalRequest.query.filter_by(shop_id=s.id, status='pending').first()
                    bill_count  = Bill.query.filter_by(shop_id=s.id).count()
                    today_bills = Bill.query.filter(
                        Bill.shop_id == s.id,
                        db.or_(
                            Bill.custom_date.like(today_str + '%'),
                            db.cast(Bill.bill_date, db.String).like(today_str + '%')
                        )
                    ).count()
                    rev_row = db.session.execute(
                        text("SELECT COALESCE(SUM(total_amount),0) FROM bill WHERE shop_id=:sid"),
                        {'sid': s.id}
                    ).fetchone()
                    total_revenue = float(rev_row[0]) if rev_row else 0.0

                    today_rev_row = db.session.execute(
                        text("SELECT COALESCE(SUM(total_amount),0) FROM bill WHERE shop_id=:sid AND (custom_date LIKE :td OR CAST(bill_date AS TEXT) LIKE :td)"),
                        {'sid': s.id, 'td': today_str + '%'}
                    ).fetchone()
                    today_revenue = float(today_rev_row[0]) if today_rev_row else 0.0

                    # Monthly breakdown (YYYY-MM)
                    monthly_rows = db.session.execute(
                        text("""
                            SELECT 
                                COALESCE(NULLIF(SUBSTR(custom_date, 1, 7), ''), strftime('%Y-%m', bill_date)) AS ym,
                                COALESCE(SUM(total_amount), 0)
                            FROM bill
                            WHERE shop_id = :sid
                            GROUP BY ym
                            HAVING ym IS NOT NULL AND ym != ''
                            ORDER BY ym DESC
                        """), {'sid': s.id}
                    ).fetchall()
                    monthly_rev = {r[0]: round(float(r[1]), 2) for r in monthly_rows if r[0]}

                    # Yearly breakdown (YYYY)
                    yearly_rows = db.session.execute(
                        text("""
                            SELECT 
                                COALESCE(NULLIF(SUBSTR(custom_date, 1, 4), ''), strftime('%Y', bill_date)) AS yr,
                                COALESCE(SUM(total_amount), 0)
                            FROM bill
                            WHERE shop_id = :sid
                            GROUP BY yr
                            HAVING yr IS NOT NULL AND yr != ''
                            ORDER BY yr DESC
                        """), {'sid': s.id}
                    ).fetchall()
                    yearly_rev = {r[0]: round(float(r[1]), 2) for r in yearly_rows if r[0]}

                    payload = {
                        'shop_id':              s.id,
                        'shop_name':            s.shop_name,
                        'owner_name':           s.owner_name,
                        'email':                s.email,
                        'phone':                s.phone,
                        'address':              s.address,
                        'approved':             s.approved,
                        'med_count':            Medicine.query.filter_by(shop_id=s.id).count(),
                        'bill_count':           bill_count,
                        'today_bills':          today_bills,
                        'total_revenue':        round(total_revenue, 2),
                        'today_revenue':        round(today_revenue, 2),
                        'monthly_revenue':      monthly_rev,
                        'yearly_revenue':       yearly_rev,
                        'joined_on':            s.created_at.strftime('%Y-%m-%d %H:%M') if s.created_at else 'N/A',
                        'license_end':          s.license_end.strftime('%Y-%m-%d') if s.license_end else 'N/A',
                        'login_time':           active_sessions.get(s.id,{}).get('login_time','-'),
                        'has_renewal_request':  bool(pending_renewal),
                        'renewal_requested_at': pending_renewal.requested_at.strftime('%Y-%m-%d %H:%M') if pending_renewal and pending_renewal.requested_at else 'N/A'
                    }

                    req = urllib.request.Request(
                        f"{CLOUD_ADMIN_URL}/api/sync/heartbeat",
                        data=json.dumps(payload).encode('utf-8'),
                        headers={'Content-Type': 'application/json'}
                    )
                    try:
                        with urllib.request.urlopen(req, timeout=5) as resp:
                            res_data = json.loads(resp.read().decode('utf-8'))
                            # If admin remotely stopped shop, reflect locally
                            if res_data.get('is_stopped') and not s.is_stopped:
                                s.is_stopped = True
                                db.session.commit()
                            elif not res_data.get('is_stopped') and s.is_stopped:
                                s.is_stopped = False
                                db.session.commit()
                            
                            # If admin approved or rejected license renewal on Cloud Admin, update local request
                            renewal_status = res_data.get('renewal_status')
                            if renewal_status and pending_renewal:
                                if renewal_status == 'rejected' and pending_renewal.status != 'rejected':
                                    pending_renewal.status = 'rejected'
                                    pending_renewal.processed_at = datetime.now()
                                    db.session.commit()
                                elif renewal_status == 'approved' and pending_renewal.status != 'approved':
                                    pending_renewal.status = 'approved'
                                    pending_renewal.processed_at = datetime.now()
                                    db.session.commit()

                            # If admin approved license renewal on Cloud Admin, update local license
                            cloud_lic_str = res_data.get('license_end')
                            if cloud_lic_str and cloud_lic_str != 'N/A':
                                try:
                                    cloud_dt = datetime.strptime(cloud_lic_str, '%Y-%m-%d')
                                    if not s.license_end or cloud_dt.date() > s.license_end.date():
                                        s.license_end = cloud_dt
                                        if pending_renewal:
                                            pending_renewal.status = 'approved'
                                            pending_renewal.processed_at = datetime.now()
                                        db.session.commit()
                                except Exception:
                                    pass
                    except Exception:
                        pass
        except Exception as e:
            pass
        time.sleep(15)

# ══════════════════════════════════════════════════════════════════════
#  GST API ENDPOINTS
# ══════════════════════════════════════════════════════════════════════

@app.route('/api/gst/financial-years', methods=['GET', 'POST'])
@jwt_required()
def handle_gst_financial_years():
    if request.method == 'GET':
        fys = FinancialYear.query.order_by(FinancialYear.start_date.desc()).all()
        return jsonify([{
            'id': f.id,
            'fy_name': f.fy_name,
            'start_date': f.start_date,
            'end_date': f.end_date,
            'is_active': f.is_active
        } for f in fys])
    
    d = request.json
    if not d.get('fy_name') or not d.get('start_date') or not d.get('end_date'):
        return jsonify({'error': 'Name, start date, and end date are required'}), 400
    
    is_act = bool(d.get('is_active', False))
    if is_act:
        FinancialYear.query.update({FinancialYear.is_active: False})
        
    fy = FinancialYear(
        fy_name=d['fy_name'].strip(),
        start_date=d['start_date'].strip(),
        end_date=d['end_date'].strip(),
        is_active=is_act
    )
    db.session.add(fy)
    db.session.commit()
    return jsonify({'message': 'Financial Year added', 'id': fy.id}), 201

@app.route('/api/gst/financial-years/active', methods=['PUT'])
@jwt_required()
def set_active_financial_year():
    d = request.json
    fy_id = d.get('id')
    if not fy_id:
        return jsonify({'error': 'Financial Year ID is required'}), 400
    
    fy = FinancialYear.query.get(fy_id)
    if not fy:
        return jsonify({'error': 'Financial year not found'}), 404
        
    FinancialYear.query.update({FinancialYear.is_active: False})
    fy.is_active = True
    db.session.commit()
    return jsonify({'message': 'Active Financial Year updated'})

@app.route('/api/gst/dashboard-stats', methods=['GET'])
@jwt_required()
def get_gst_dashboard_stats():
    import json
    sid = int(get_jwt_identity())
    fy_id = request.args.get('financial_year_id', type=int)
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')
    month = request.args.get('month')
    gst_rate_filter = request.args.get('gst_rate', type=float)
    party = request.args.get('party')
    doc_no = request.args.get('doc_no')

    sb_query = Bill.query.filter_by(shop_id=sid)
    pe_query = PurchaseEntry.query.filter_by(shop_id=sid)

    if fy_id:
        sb_query = sb_query.filter_by(financial_year_id=fy_id)
        pe_query = pe_query.filter_by(financial_year_id=fy_id)
    else:
        active_fy = FinancialYear.query.filter_by(is_active=True).first()
        if active_fy:
            sb_query = sb_query.filter_by(financial_year_id=active_fy.id)
            pe_query = pe_query.filter_by(financial_year_id=active_fy.id)

    if from_date:
        sb_query = sb_query.filter(db.or_(Bill.custom_date >= from_date, Bill.bill_date >= from_date))
        pe_query = pe_query.filter(PurchaseEntry.entry_date >= from_date)
    if to_date:
        sb_query = sb_query.filter(db.or_(Bill.custom_date <= to_date, Bill.bill_date <= to_date))
        pe_query = pe_query.filter(PurchaseEntry.entry_date <= to_date)
    if month:
        sb_query = sb_query.filter(db.or_(Bill.custom_date.like(f"{month}%"), db.func.strftime('%Y-%m', Bill.bill_date) == month))
        pe_query = pe_query.filter(PurchaseEntry.entry_date.like(f"{month}%"))
    if gst_rate_filter is not None:
        sb_query = sb_query.filter(Bill.gst_rate == gst_rate_filter)
        pe_query = pe_query.filter(PurchaseEntry.gst_rate == gst_rate_filter)
    if party:
        sb_query = sb_query.filter(Bill.customer_name.ilike(f"%{party}%"))
        pe_query = pe_query.filter(PurchaseEntry.supplier_name.ilike(f"%{party}%"))
    if doc_no:
        sb_query = sb_query.filter(Bill.bill_number.like(f"%{doc_no}%"))
        pe_query = pe_query.filter(PurchaseEntry.entry_number.like(f"%{doc_no}%"))

    bills = sb_query.all()
    purchases = pe_query.all()

    taxable_sales = sum(b.taxable_amount or 0 for b in bills if b.status != 'returned')
    taxable_purchases = sum(p.taxable_amount or 0 for p in purchases)
    cgst_collected = sum(b.cgst_amount or 0 for b in bills if b.status != 'returned')
    sgst_collected = sum(b.sgst_amount or 0 for b in bills if b.status != 'returned')
    igst_collected = sum(b.igst_amount or 0 for b in bills if b.status != 'returned')
    total_gst_collected = sum(b.total_gst or 0 for b in bills if b.status != 'returned')

    cgst_paid = sum(p.cgst_amount or 0 for p in purchases)
    sgst_paid = sum(p.sgst_amount or 0 for p in purchases)
    igst_paid = sum(p.igst_amount or 0 for p in purchases)
    total_gst_paid = sum(p.total_gst or 0 for p in purchases)

    net_cgst = max(0.0, cgst_collected - cgst_paid)
    net_sgst = max(0.0, sgst_collected - sgst_paid)
    net_igst = max(0.0, igst_collected - igst_paid)
    net_gst_liability = net_cgst + net_sgst + net_igst

    return jsonify({
        'taxable_sales': taxable_sales,
        'taxable_purchases': taxable_purchases,
        'cgst_collected': cgst_collected,
        'sgst_collected': sgst_collected,
        'igst_collected': igst_collected,
        'total_gst_collected': total_gst_collected,
        'cgst_paid': cgst_paid,
        'sgst_paid': sgst_paid,
        'igst_paid': igst_paid,
        'total_gst_paid': total_gst_paid,
        'net_cgst': net_cgst,
        'net_sgst': net_sgst,
        'net_igst': net_igst,
        'net_gst_liability': net_gst_liability
    })

def filter_records_by_query_params(model_cls, sid, is_purchase=False):
    fy_id = request.args.get('financial_year_id', type=int)
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')
    month = request.args.get('month')
    gst_rate_filter = request.args.get('gst_rate', type=float)
    party = request.args.get('party')
    doc_no = request.args.get('doc_no')
    hsn = request.args.get('hsn_code')
    gstin = request.args.get('gstin')
    report_type = request.args.get('report_type', 'ALL')  # ALL, B2B, B2C, Export, SEZ, Nil, Exempted

    q = model_cls.query.filter_by(shop_id=sid)

    if fy_id:
        q = q.filter_by(financial_year_id=fy_id)
    else:
        active_fy = FinancialYear.query.filter_by(is_active=True).first()
        if active_fy:
            q = q.filter_by(financial_year_id=active_fy.id)

    if from_date:
        if is_purchase:
            q = q.filter(model_cls.entry_date >= from_date)
        else:
            q = q.filter(db.or_(model_cls.custom_date >= from_date, model_cls.bill_date >= from_date))
    if to_date:
        if is_purchase:
            q = q.filter(model_cls.entry_date <= to_date)
        else:
            q = q.filter(db.or_(model_cls.custom_date <= to_date, model_cls.bill_date <= to_date))

    if month:
        if is_purchase:
            q = q.filter(model_cls.entry_date.like(f"{month}%"))
        else:
            q = q.filter(db.or_(model_cls.custom_date.like(f"{month}%"), db.func.strftime('%Y-%m', model_cls.bill_date) == month))

    if gst_rate_filter is not None:
        q = q.filter(model_cls.gst_rate == gst_rate_filter)

    if party:
        if is_purchase:
            q = q.filter(model_cls.supplier_name.ilike(f"%{party}%"))
        else:
            q = q.filter(model_cls.customer_name.ilike(f"%{party}%"))

    if doc_no:
        if is_purchase:
            q = q.filter(model_cls.entry_number.like(f"%{doc_no}%"))
        else:
            q = q.filter(model_cls.bill_number.like(f"%{doc_no}%"))

    if hsn:
        q = q.filter(model_cls.hsn_code.like(f"%{hsn}%"))

    # GSTIN filter
    if gstin:
        if is_purchase:
            q = q.filter(model_cls.supplier_gstin.ilike(f"%{gstin}%"))
        else:
            q = q.filter(model_cls.customer_gstin.ilike(f"%{gstin}%"))

    # Report type filter (only for sales/bills)
    if not is_purchase and report_type and report_type != 'ALL':
        if report_type == 'B2B':
            q = q.filter(model_cls.customer_gstin != None, model_cls.customer_gstin != '')
        elif report_type == 'B2C':
            q = q.filter(db.or_(model_cls.customer_gstin == None, model_cls.customer_gstin == ''))
        elif report_type in ('Export', 'SEZ', 'Nil Rated', 'Exempted'):
            # For pharmacy context, these are specialty types stored in place_of_supply or hsn patterns
            # Nil Rated / Exempted: gst_rate = 0
            if report_type in ('Nil Rated', 'Exempted'):
                q = q.filter(model_cls.gst_rate == 0)
            elif report_type == 'Export':
                q = q.filter(model_cls.place_of_supply.in_(['97', '96']))
            elif report_type == 'SEZ':
                q = q.filter(model_cls.place_of_supply.like('SEZ%'))

    return q

@app.route('/api/gst/reports/sales-register', methods=['GET'])
@jwt_required()
def get_gst_sales_register():
    import json
    sid = int(get_jwt_identity())
    q = filter_records_by_query_params(Bill, sid, is_purchase=False)
    bills = q.order_by(Bill.bill_number.desc()).all()
    result = []
    for b in bills:
        date_str = b.custom_date or (b.bill_date.strftime('%Y-%m-%d') if b.bill_date else '')
        month_str = date_str[:7] if date_str else ''
        year_str = date_str[:4] if date_str else ''
        subtotal = float(b.subtotal or 0)
        taxable = float(b.taxable_amount or 0)
        cgst_a = float(b.cgst_amount or 0)
        sgst_a = float(b.sgst_amount or 0)
        igst_a = float(b.igst_amount or 0)
        total_gst = float(b.total_gst or 0)
        grand = float(b.grand_total or b.total_amount or 0)
        round_off = round(grand - (taxable + total_gst), 2)
        gstin = (b.customer_gstin or '').strip()
        doc_type = 'B2B' if (gstin and len(gstin) == 15) else 'B2C'
        result.append({
            'id': b.id,
            'document_type': doc_type,
            'bill_number': b.bill_number,
            'date': date_str,
            'month': month_str,
            'year': year_str,
            'customer_name': b.customer_name,
            'customer_gstin': gstin,
            'place_of_supply': b.place_of_supply or '',
            'hsn_code': b.hsn_code or '',
            'gst_rate': float(b.gst_rate or 0),
            'taxable_amount': taxable,
            'cgst_amount': cgst_a,
            'sgst_amount': sgst_a,
            'igst_amount': igst_a,
            'total_gst': total_gst,
            'grand_total': grand,
            'round_off': round_off,
            'status': b.status or 'active',
            'items': json.loads(b.items_json or '[]')
        })
    return jsonify(result)

@app.route('/api/gst/reports/purchase-register', methods=['GET'])
@jwt_required()
def get_gst_purchase_register():
    import json
    sid = int(get_jwt_identity())
    q = filter_records_by_query_params(PurchaseEntry, sid, is_purchase=True)
    purchases = q.order_by(PurchaseEntry.entry_number.desc()).all()
    result = []
    for p in purchases:
        date_str = p.entry_date or ''
        month_str = date_str[:7] if date_str else ''
        year_str = date_str[:4] if date_str else ''
        taxable = float(p.taxable_amount or 0)
        cgst_a = float(p.cgst_amount or 0)
        sgst_a = float(p.sgst_amount or 0)
        igst_a = float(p.igst_amount or 0)
        total_gst = float(p.total_gst or 0)
        grand = float(p.grand_total or p.net_amount or 0)
        round_off = round(grand - (taxable + total_gst), 2)
        gstin = (p.supplier_gstin or '').strip()
        doc_type = 'B2B' if (gstin and len(gstin) == 15) else 'B2C'
        result.append({
            'id': p.id,
            'document_type': doc_type,
            'entry_number': p.entry_number,
            'party_number': p.party_number or '',
            'date': date_str,
            'month': month_str,
            'year': year_str,
            'supplier_name': p.supplier_name,
            'supplier_gstin': gstin,
            'place_of_supply': p.place_of_supply or '',
            'hsn_code': p.hsn_code or '',
            'gst_rate': float(p.gst_rate or 0),
            'taxable_amount': taxable,
            'cgst_amount': cgst_a,
            'sgst_amount': sgst_a,
            'igst_amount': igst_a,
            'total_gst': total_gst,
            'grand_total': grand,
            'round_off': round_off,
            'items': json.loads(p.items_json or '[]')
        })
    return jsonify(result)

def _build_slab_breakdown(items_json_str, is_interstate=False):
    """Parse items JSON and return GST slab-wise breakdown."""
    import json as _json
    slabs = {0: {'taxable':0,'cgst':0,'sgst':0,'igst':0},
              5: {'taxable':0,'cgst':0,'sgst':0,'igst':0},
             12: {'taxable':0,'cgst':0,'sgst':0,'igst':0},
             18: {'taxable':0,'cgst':0,'sgst':0,'igst':0},
             28: {'taxable':0,'cgst':0,'sgst':0,'igst':0}}
    try:
        items = _json.loads(items_json_str or '[]')
        for it in items:
            gst_pct = float(it.get('gst', 0) or it.get('gst_pct', 0) or it.get('gst_rate', 0))
            qty = float(it.get('qty', 0))
            price = float(it.get('price', 0))
            mrp = float(it.get('mrp', price))
            disc_pct = float(it.get('discount_pct', 0) or it.get('disc', 0))
            amount = float(it.get('amount', qty * price * (1 - disc_pct/100)))
            # Get taxable value from amount (amount is inclusive or exclusive?)
            # In the quick entry: amount = qty * price * (1 - disc/100) (base amount, gst added on top)
            taxable = amount
            gst_amt = taxable * (gst_pct / 100)
            # Normalize to known slabs
            slab_key = 0
            for s in [0, 5, 12, 18, 28]:
                if abs(gst_pct - s) < 0.5:
                    slab_key = s
                    break
            if slab_key not in slabs:
                slabs[slab_key] = {'taxable':0,'cgst':0,'sgst':0,'igst':0}
            slabs[slab_key]['taxable'] += taxable
            if is_interstate:
                slabs[slab_key]['igst'] += gst_amt
            else:
                slabs[slab_key]['cgst'] += gst_amt / 2
                slabs[slab_key]['sgst'] += gst_amt / 2
    except Exception:
        pass
    return slabs

def _build_hsn_map_from_bills(bills, shop_gst_number):
    """Helper: build HSN aggregate from a list of bill or purchase records."""
    import json as _json
    hsn_map = {}
    shop_state_code = (shop_gst_number or '')[:2]
    for rec in bills:
        items_json_str = getattr(rec, 'items_json', '[]') or '[]'
        items = []
        try:
            items = _json.loads(items_json_str)
        except Exception:
            pass
        pos = (getattr(rec, 'place_of_supply', '') or '').strip()
        is_interstate = bool(pos and shop_state_code and pos[:2] != shop_state_code)
        for item in items:
            hsn = (item.get('hsn_code') or item.get('hsn') or 'N/A').strip()
            qty = float(item.get('qty', 0))
            price = float(item.get('price', 0))
            gst_pct = float(item.get('gst', 0) or item.get('gst_pct', 0) or item.get('gst_rate', 0))
            disc_pct = float(item.get('discount_pct', 0) or item.get('disc', 0))
            amount = float(item.get('amount', qty * price * (1 - disc_pct/100)))
            taxable = amount
            gst_amt = taxable * (gst_pct / 100)
            if hsn not in hsn_map:
                hsn_map[hsn] = {
                    'hsn_code': hsn,
                    'description': item.get('name', ''),
                    'gst_rate': gst_pct,
                    'quantity': 0,
                    'taxable_amount': 0.0,
                    'cgst_amount': 0.0,
                    'sgst_amount': 0.0,
                    'igst_amount': 0.0,
                    'total_gst': 0.0,
                    'total_amount': 0.0,
                    'invoice_count': 0,
                }
            hsn_map[hsn]['quantity'] += qty
            hsn_map[hsn]['taxable_amount'] += taxable
            hsn_map[hsn]['total_gst'] += gst_amt
            hsn_map[hsn]['total_amount'] += amount + gst_amt
            hsn_map[hsn]['invoice_count'] += 1
            if is_interstate:
                hsn_map[hsn]['igst_amount'] += gst_amt
            else:
                hsn_map[hsn]['cgst_amount'] += gst_amt / 2
                hsn_map[hsn]['sgst_amount'] += gst_amt / 2
    result = []
    for h in hsn_map.values():
        result.append({
            'hsn_code':      h['hsn_code'],
            'description':   h['description'],
            'gst_rate':      h['gst_rate'],
            'uqc':           'NOS',
            'total_qty':     round(h['quantity'], 3),
            'total_value':   round(h['total_amount'], 2),
            'taxable_value': round(h['taxable_amount'], 2),
            'cgst':          round(h['cgst_amount'], 2),
            'sgst':          round(h['sgst_amount'], 2),
            'igst':          round(h['igst_amount'], 2),
            'total_gst':     round(h['total_gst'], 2),
            'invoice_count': h['invoice_count'],
        })
    return result

@app.route('/api/gst/reports/gstr1', methods=['GET'])
@jwt_required()
def get_gstr1_report():
    import json
    sid = int(get_jwt_identity())
    shop = Shop.query.get(sid)
    shop_state_code = (shop.gst_number or '')[:2] if shop else ''
    q = filter_records_by_query_params(Bill, sid, is_purchase=False)
    bills = q.filter(Bill.status != 'returned').order_by(Bill.bill_number.asc()).all()
    result = []
    for b in bills:
        gstin = (b.customer_gstin or '').strip()
        pos = (b.place_of_supply or '').strip()
        is_interstate = bool(pos and shop_state_code and pos[:2] != shop_state_code)
        date_str = b.custom_date or (b.bill_date.strftime('%Y-%m-%d') if b.bill_date else '')
        month_str = date_str[:7] if date_str else ''
        year_str = date_str[:4] if date_str else ''
        doc_type = 'B2B' if (gstin and len(gstin) == 15) else 'B2C'
        taxable = float(b.taxable_amount or 0)
        cgst_a = float(b.cgst_amount or 0)
        sgst_a = float(b.sgst_amount or 0)
        igst_a = float(b.igst_amount or 0)
        total_gst = float(b.total_gst or 0)
        grand = float(b.grand_total or b.total_amount or 0)
        round_off = round(grand - (taxable + total_gst), 2)
        slabs = _build_slab_breakdown(b.items_json, is_interstate)
        record = {
            'document_type': doc_type,
            'customer_gstin': gstin,
            'bill_number': b.bill_number,
            'date': date_str,
            'month': month_str,
            'year': year_str,
            'customer_name': b.customer_name,
            'place_of_supply': pos,
            'taxable_amount': taxable,
            'cgst_amount': cgst_a,
            'sgst_amount': sgst_a,
            'igst_amount': igst_a,
            'total_gst': total_gst,
            'grand_total': grand,
            'round_off': round_off,
            'gst_rate': float(b.gst_rate or 0),
            'gst0_taxable':   round(slabs[0]['taxable'], 2),
            'cgst_2_5':  round(slabs[5]['cgst'], 2),
            'sgst_2_5':  round(slabs[5]['sgst'], 2),
            'igst_5':    round(slabs[5]['igst'], 2),
            'gst5_taxable':  round(slabs[5]['taxable'], 2),
            'cgst_6':    round(slabs[12]['cgst'], 2),
            'sgst_6':    round(slabs[12]['sgst'], 2),
            'igst_12':   round(slabs[12]['igst'], 2),
            'gst12_taxable': round(slabs[12]['taxable'], 2),
            'cgst_9':    round(slabs[18]['cgst'], 2),
            'sgst_9':    round(slabs[18]['sgst'], 2),
            'igst_18':   round(slabs[18]['igst'], 2),
            'gst18_taxable': round(slabs[18]['taxable'], 2),
            'cgst_14':   round(slabs[28]['cgst'], 2),
            'sgst_14':   round(slabs[28]['sgst'], 2),
            'igst_28':   round(slabs[28]['igst'], 2),
            'gst28_taxable': round(slabs[28]['taxable'], 2),
        }
        result.append(record)
    return jsonify(result)

@app.route('/api/gst/reports/gstr2', methods=['GET'])
@jwt_required()
def get_gstr2_report():
    """GSTR-2 — Inward supplies (purchase register) in GSTR-1 format."""
    import json
    sid = int(get_jwt_identity())
    shop = Shop.query.get(sid)
    shop_state_code = (shop.gst_number or '')[:2] if shop else ''
    q = filter_records_by_query_params(PurchaseEntry, sid, is_purchase=True)
    purchases = q.order_by(PurchaseEntry.entry_number.asc()).all()
    result = []
    for p in purchases:
        gstin = (p.supplier_gstin or '').strip()
        pos = (p.place_of_supply or '').strip()
        is_interstate = bool(pos and shop_state_code and pos[:2] != shop_state_code)
        date_str = p.entry_date or ''
        month_str = date_str[:7] if date_str else ''
        year_str = date_str[:4] if date_str else ''
        doc_type = 'B2B' if (gstin and len(gstin) == 15) else 'B2C'
        taxable = float(p.taxable_amount or 0)
        cgst_a = float(p.cgst_amount or 0)
        sgst_a = float(p.sgst_amount or 0)
        igst_a = float(p.igst_amount or 0)
        total_gst = float(p.total_gst or 0)
        grand = float(p.grand_total or p.net_amount or 0)
        round_off = round(grand - (taxable + total_gst), 2)
        slabs = _build_slab_breakdown(p.items_json, is_interstate)
        record = {
            'document_type': doc_type,
            'supplier_gstin': gstin,
            'entry_number': p.entry_number,
            'party_number': p.party_number or '',
            'date': date_str,
            'month': month_str,
            'year': year_str,
            'supplier_name': p.supplier_name,
            'place_of_supply': pos,
            'taxable_amount': taxable,
            'cgst_amount': cgst_a,
            'sgst_amount': sgst_a,
            'igst_amount': igst_a,
            'total_gst': total_gst,
            'grand_total': grand,
            'round_off': round_off,
            'gst_rate': float(p.gst_rate or 0),
            'gst0_taxable':   round(slabs[0]['taxable'], 2),
            'cgst_2_5':  round(slabs[5]['cgst'], 2),
            'sgst_2_5':  round(slabs[5]['sgst'], 2),
            'igst_5':    round(slabs[5]['igst'], 2),
            'gst5_taxable':  round(slabs[5]['taxable'], 2),
            'cgst_6':    round(slabs[12]['cgst'], 2),
            'sgst_6':    round(slabs[12]['sgst'], 2),
            'igst_12':   round(slabs[12]['igst'], 2),
            'gst12_taxable': round(slabs[12]['taxable'], 2),
            'cgst_9':    round(slabs[18]['cgst'], 2),
            'sgst_9':    round(slabs[18]['sgst'], 2),
            'igst_18':   round(slabs[18]['igst'], 2),
            'gst18_taxable': round(slabs[18]['taxable'], 2),
            'cgst_14':   round(slabs[28]['cgst'], 2),
            'sgst_14':   round(slabs[28]['sgst'], 2),
            'igst_28':   round(slabs[28]['igst'], 2),
            'gst28_taxable': round(slabs[28]['taxable'], 2),
        }
        result.append(record)
    return jsonify(result)

@app.route('/api/gst/reports/gstr3b', methods=['GET'])
@jwt_required()
def get_gstr3b_report():
    sid = int(get_jwt_identity())
    sales_q = filter_records_by_query_params(Bill, sid, is_purchase=False).filter(Bill.status != 'returned')
    purchases_q = filter_records_by_query_params(PurchaseEntry, sid, is_purchase=True)
    sales = sales_q.all()
    purchases = purchases_q.all()
    out_taxable = round(sum(float(b.taxable_amount or 0) for b in sales), 2)
    out_cgst = round(sum(float(b.cgst_amount or 0) for b in sales), 2)
    out_sgst = round(sum(float(b.sgst_amount or 0) for b in sales), 2)
    out_igst = round(sum(float(b.igst_amount or 0) for b in sales), 2)
    out_total = round(out_cgst + out_sgst + out_igst, 2)
    in_taxable = round(sum(float(p.taxable_amount or 0) for p in purchases), 2)
    in_cgst = round(sum(float(p.cgst_amount or 0) for p in purchases), 2)
    in_sgst = round(sum(float(p.sgst_amount or 0) for p in purchases), 2)
    in_igst = round(sum(float(p.igst_amount or 0) for p in purchases), 2)
    in_total = round(in_cgst + in_sgst + in_igst, 2)
    net_cgst = round(out_cgst - in_cgst, 2)
    net_sgst = round(out_sgst - in_sgst, 2)
    net_igst = round(out_igst - in_igst, 2)
    net_payable = round(max(0, net_cgst) + max(0, net_sgst) + max(0, net_igst), 2)
    return jsonify({
        'outward_supplies': {'taxable_amount': out_taxable, 'cgst_amount': out_cgst, 'sgst_amount': out_sgst, 'igst_amount': out_igst, 'total_gst': out_total},
        'eligible_itc':     {'taxable_amount': in_taxable,  'cgst_amount': in_cgst,  'sgst_amount': in_sgst,  'igst_amount': in_igst,  'total_gst': in_total},
        'net_liability':    {'cgst': net_cgst, 'sgst': net_sgst, 'igst': net_igst, 'total': net_payable}
    })

@app.route('/api/gst/reports/hsn-summary', methods=['GET'])
@jwt_required()
def get_hsn_summary_report():
    import json
    sid = int(get_jwt_identity())
    
    sales_q = filter_records_by_query_params(Bill, sid, is_purchase=False).filter(Bill.status != 'returned')
    bills = sales_q.all()
    
    shop = Shop.query.get(sid)
    shop_state_code = (shop.gst_number or '')[:2] if shop else ''
    hsn_map = {}
    for b in bills:
        items = json.loads(b.items_json or '[]')
        is_interstate = bool(
            b.place_of_supply and shop_state_code and
            b.place_of_supply[:2] != shop_state_code
        )
        for item in items:
            hsn = (item.get('hsn_code') or item.get('hsn') or 'N/A').strip()
            qty = float(item.get('qty', 0))
            price = float(item.get('price', 0))
            gst_pct = float(item.get('gst', 0) or item.get('gst_pct', 0))
            amount = float(item.get('amount', qty * price))
            tax_factor = 1 + (gst_pct / 100)
            taxable = amount / tax_factor if tax_factor else amount
            gst_amt = amount - taxable

            key = (hsn, gst_pct)  # group by HSN + rate
            if key not in hsn_map:
                hsn_map[key] = {
                    'hsn_code': hsn,
                    'description': item.get('name', ''),
                    'gst_rate': gst_pct,
                    'quantity': 0.0,
                    'taxable_amount': 0.0,
                    'cgst_amount': 0.0,
                    'sgst_amount': 0.0,
                    'igst_amount': 0.0,
                    'total_gst': 0.0,
                    'total_amount': 0.0,
                    'invoice_ids': set(),
                }
            hsn_map[key]['quantity'] += qty
            hsn_map[key]['taxable_amount'] += taxable
            hsn_map[key]['total_gst'] += gst_amt
            hsn_map[key]['total_amount'] += amount
            hsn_map[key]['invoice_ids'].add(b.id)
            if is_interstate:
                hsn_map[key]['igst_amount'] += gst_amt
            else:
                hsn_map[key]['cgst_amount'] += gst_amt / 2
                hsn_map[key]['sgst_amount'] += gst_amt / 2

    result = []
    for h in hsn_map.values():
        result.append({
            'hsn_code':      h['hsn_code'],
            'description':   h['description'],
            'gst_rate':      h['gst_rate'],
            'uqc':           'NOS',
            'total_qty':     round(h['quantity'], 3),
            'total_value':   round(h['total_amount'], 2),
            'taxable_value': round(h['taxable_amount'], 2),
            'cgst':          round(h['cgst_amount'], 2),
            'sgst':          round(h['sgst_amount'], 2),
            'igst':          round(h['igst_amount'], 2),
            'total_gst':     round(h['total_gst'], 2),
            'invoice_count': len(h['invoice_ids']),
        })
    result.sort(key=lambda x: (x['hsn_code'], x['gst_rate']))
    return jsonify(result)


@app.route('/api/gst/reports/purchase-hsn-summary', methods=['GET'])
@jwt_required()
def get_purchase_hsn_summary_report():
    """Purchase HSN/SAC summary grouped by HSN code and GST rate (ITC summary)."""
    import json
    sid = int(get_jwt_identity())
    shop = Shop.query.get(sid)
    shop_state_code = (shop.gst_number or '')[:2] if shop else ''
    q = filter_records_by_query_params(PurchaseEntry, sid, is_purchase=True)
    purchases = q.all()

    hsn_map = {}
    for p in purchases:
        items = json.loads(p.items_json or '[]')
        is_interstate = bool(
            p.place_of_supply and shop_state_code and
            p.place_of_supply[:2] != shop_state_code
        )
        for item in items:
            hsn = (item.get('hsn_code') or item.get('hsn') or 'N/A').strip()
            qty = float(item.get('qty', 0))
            price = float(item.get('price', 0))
            gst_pct = float(item.get('gst', 0) or item.get('gst_pct', 0))
            amount = float(item.get('amount', qty * price))
            tax_factor = 1 + (gst_pct / 100)
            taxable = amount / tax_factor if tax_factor else amount
            gst_amt = amount - taxable

            key = (hsn, gst_pct)
            if key not in hsn_map:
                hsn_map[key] = {
                    'hsn_code': hsn,
                    'description': item.get('name', ''),
                    'gst_rate': gst_pct,
                    'quantity': 0.0,
                    'taxable_amount': 0.0,
                    'cgst_amount': 0.0,
                    'sgst_amount': 0.0,
                    'igst_amount': 0.0,
                    'total_gst': 0.0,
                    'total_amount': 0.0,
                    'invoice_ids': set(),
                }
            hsn_map[key]['quantity'] += qty
            hsn_map[key]['taxable_amount'] += taxable
            hsn_map[key]['total_gst'] += gst_amt
            hsn_map[key]['total_amount'] += amount
            hsn_map[key]['invoice_ids'].add(p.id)
            if is_interstate:
                hsn_map[key]['igst_amount'] += gst_amt
            else:
                hsn_map[key]['cgst_amount'] += gst_amt / 2
                hsn_map[key]['sgst_amount'] += gst_amt / 2

    result = []
    for h in hsn_map.values():
        result.append({
            'hsn_code':      h['hsn_code'],
            'description':   h['description'],
            'gst_rate':      h['gst_rate'],
            'uqc':           'NOS',
            'total_qty':     round(h['quantity'], 3),
            'total_value':   round(h['total_amount'], 2),
            'taxable_value': round(h['taxable_amount'], 2),
            'cgst':          round(h['cgst_amount'], 2),
            'sgst':          round(h['sgst_amount'], 2),
            'igst':          round(h['igst_amount'], 2),
            'total_gst':     round(h['total_gst'], 2),
            'invoice_count': len(h['invoice_ids']),
        })
    result.sort(key=lambda x: (x['hsn_code'], x['gst_rate']))
    return jsonify(result)

@app.route('/api/gst/reports/summary', methods=['GET'])
@jwt_required()
def get_gst_summary():
    sid = int(get_jwt_identity())
    
    sales_q = filter_records_by_query_params(Bill, sid, is_purchase=False).filter(Bill.status != 'returned')
    purchases_q = filter_records_by_query_params(PurchaseEntry, sid, is_purchase=True)
    
    sales = sales_q.all()
    purchases = purchases_q.all()
    
    return jsonify({
        'collected': {
            'cgst': sum(b.cgst_amount or 0 for b in sales),
            'sgst': sum(b.sgst_amount or 0 for b in sales),
            'igst': sum(b.igst_amount or 0 for b in sales),
            'total': sum(b.total_gst or 0 for b in sales)
        },
        'paid': {
            'cgst': sum(p.cgst_amount or 0 for p in purchases),
            'sgst': sum(p.sgst_amount or 0 for p in purchases),
            'igst': sum(p.igst_amount or 0 for p in purchases),
            'total': sum(p.total_gst or 0 for p in purchases)
        }
    })

@app.route('/api/gst/reports/monthly', methods=['GET'])
@jwt_required()
def get_gst_monthly_report():
    sid = int(get_jwt_identity())

    sales = filter_records_by_query_params(Bill, sid, is_purchase=False).filter(Bill.status != 'returned').all()
    purchases = filter_records_by_query_params(PurchaseEntry, sid, is_purchase=True).all()

    months_data = {}

    def _empty_month(m):
        return {
            'month': m,
            'sales_taxable':    0.0,
            'sales_cgst':       0.0,
            'sales_sgst':       0.0,
            'sales_igst':       0.0,
            'purchase_taxable': 0.0,
            'purchase_cgst':    0.0,
            'purchase_sgst':    0.0,
            'purchase_igst':    0.0,
        }

    for b in sales:
        m = b.custom_date[:7] if b.custom_date else (b.bill_date.strftime('%Y-%m') if b.bill_date else '')
        if not m:
            continue
        if m not in months_data:
            months_data[m] = _empty_month(m)
        months_data[m]['sales_taxable'] += float(b.taxable_amount or 0)
        months_data[m]['sales_cgst']    += float(b.cgst_amount   or 0)
        months_data[m]['sales_sgst']    += float(b.sgst_amount   or 0)
        months_data[m]['sales_igst']    += float(b.igst_amount   or 0)

    for p in purchases:
        m = p.entry_date[:7] if p.entry_date else ''
        if not m:
            continue
        if m not in months_data:
            months_data[m] = _empty_month(m)
        months_data[m]['purchase_taxable'] += float(p.taxable_amount or 0)
        months_data[m]['purchase_cgst']    += float(p.cgst_amount   or 0)
        months_data[m]['purchase_sgst']    += float(p.sgst_amount   or 0)
        months_data[m]['purchase_igst']    += float(p.igst_amount   or 0)

    result = []
    for row in sorted(months_data.values(), key=lambda x: x['month'], reverse=True):
        sales_cgst   = round(row['sales_cgst'],       2)
        sales_sgst   = round(row['sales_sgst'],       2)
        sales_igst   = round(row['sales_igst'],       2)
        pur_cgst     = round(row['purchase_cgst'],    2)
        pur_sgst     = round(row['purchase_sgst'],    2)
        pur_igst     = round(row['purchase_igst'],    2)
        net_payable  = round(
            (sales_cgst - pur_cgst) +
            (sales_sgst - pur_sgst) +
            (sales_igst - pur_igst), 2
        )
        result.append({
            'month':            row['month'],
            'sales_taxable':    round(row['sales_taxable'],    2),
            'sales_cgst':       sales_cgst,
            'sales_sgst':       sales_sgst,
            'sales_igst':       sales_igst,
            'purchase_taxable': round(row['purchase_taxable'], 2),
            'purchase_cgst':    pur_cgst,
            'purchase_sgst':    pur_sgst,
            'purchase_igst':    pur_igst,
            'net_payable':      net_payable,
        })

    return jsonify(result)

@app.route('/api/gst/reports/financial-year', methods=['GET'])
@jwt_required()
def get_gst_financial_year_report():
    sid = int(get_jwt_identity())
    
    fys = FinancialYear.query.all()
    results = []
    
    for fy in fys:
        sales = Bill.query.filter_by(shop_id=sid, financial_year_id=fy.id).filter(Bill.status != 'returned').all()
        purchases = PurchaseEntry.query.filter_by(shop_id=sid, financial_year_id=fy.id).all()
        
        results.append({
            'fy_name': fy.fy_name,
            'sales_taxable': round(sum(b.taxable_amount or 0 for b in sales), 2),
            'sales_gst': round(sum(b.total_gst or 0 for b in sales), 2),
            'purchases_taxable': round(sum(p.taxable_amount or 0 for p in purchases), 2),
            'purchases_gst': round(sum(p.total_gst or 0 for p in purchases), 2)
        })
        
    return jsonify(results)

@app.route('/api/gst/reports/filing-status', methods=['GET', 'POST'])
@jwt_required()
def handle_gst_filing_status():
    sid = int(get_jwt_identity())
    fy_id = request.args.get('financial_year_id', type=int)
    if not fy_id:
        active_fy = FinancialYear.query.filter_by(is_active=True).first()
        fy_id = active_fy.id if active_fy else None
        
    if request.method == 'GET':
        if not fy_id:
            return jsonify([])
            
        fy = FinancialYear.query.get(fy_id)
        if not fy:
            return jsonify([])
            
        start_dt = datetime.strptime(fy.start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(fy.end_date, "%Y-%m-%d")
        
        months = []
        curr = start_dt
        while curr <= end_dt:
            months.append(curr.strftime("%Y-%m"))
            if curr.month == 12:
                curr = curr.replace(year=curr.year + 1, month=1)
            else:
                curr = curr.replace(month=curr.month + 1)
                
        statuses = GSTReturnStatus.query.filter_by(shop_id=sid, financial_year_id=fy_id).all()
        status_map = {s.month_val: s for s in statuses}
        
        results = []
        for m in months:
            db_status = status_map.get(m)
            results.append({
                'month_val': m,
                'status': db_status.status if db_status else 'Pending',
                'filed_date': db_status.filed_date if db_status else ''
            })
        return jsonify(results)
        
    d = request.json
    m_val = d.get('month_val')
    status = d.get('status', 'Pending')
    filed_date = d.get('filed_date', '')
    
    if not m_val or not fy_id:
        return jsonify({'error': 'Month value and Financial Year ID are required'}), 400
        
    db_status = GSTReturnStatus.query.filter_by(shop_id=sid, financial_year_id=fy_id, month_val=m_val).first()
    if not db_status:
        db_status = GSTReturnStatus(
            shop_id=sid,
            financial_year_id=fy_id,
            month_val=m_val,
            status=status,
            filed_date=filed_date
        )
        db.session.add(db_status)
    else:
        db_status.status = status
        db_status.filed_date = filed_date
        
    db.session.commit()
    return jsonify({'message': 'Return filing status updated'})

# ══════════════════════════════════════════════════════════════════════
#  PROFESSIONAL EXPORT — HELPERS
# ══════════════════════════════════════════════════════════════════════

def _get_export_meta(sid):
    """Return shop info + active FY + filter params for export headers."""
    shop = Shop.query.get(sid)
    fy_id = request.args.get('financial_year_id', type=int)
    from_date = request.args.get('from_date', '')
    to_date   = request.args.get('to_date',   '')
    active_fy = None
    if fy_id:
        active_fy = FinancialYear.query.get(fy_id)
    if not active_fy:
        active_fy = FinancialYear.query.filter_by(is_active=True).first()
    return {
        'shop':         shop,
        'fy':           active_fy,
        'from_date':    from_date,
        'to_date':      to_date,
        'generated_at': datetime.now().strftime('%d-%m-%Y  %H:%M:%S'),
    }


def _get_report_data(report_type, sid):
    """
    Return (title, headers, rows, currency_cols_set) for the given report_type.
    currency_cols_set contains 0-indexed column positions that hold INR amounts.
    """
    import json as _json

    # ── Sales Register ────────────────────────────────────────────────
    if report_type == 'sales':
        title = 'GST Sales Register'
        q = filter_records_by_query_params(Bill, sid, is_purchase=False)
        bills = q.order_by(Bill.bill_number.desc()).all()
        headers = ['#', 'Date', 'Bill No.', 'Customer Name', 'GSTIN',
                   'Place of Supply', 'HSN Code', 'GST Rate %',
                   'Taxable Amt', 'CGST', 'SGST', 'IGST', 'Total GST', 'Grand Total', 'Type']
        rows = []
        for i, b in enumerate(bills, 1):
            date_str = b.custom_date or (b.bill_date.strftime('%Y-%m-%d') if b.bill_date else '')
            gstin = (b.customer_gstin or '').strip()
            doc_type = 'B2B' if (gstin and len(gstin) == 15) else 'B2C'
            rows.append([
                i, date_str, b.bill_number or '', b.customer_name or '',
                gstin or 'Unregistered', b.place_of_supply or '',
                b.hsn_code or '', float(b.gst_rate or 0),
                float(b.taxable_amount or 0), float(b.cgst_amount or 0),
                float(b.sgst_amount or 0), float(b.igst_amount or 0),
                float(b.total_gst or 0),
                float(b.grand_total or b.total_amount or 0), doc_type,
            ])
        return title, headers, rows, {8, 9, 10, 11, 12, 13}

    # ── Purchase Register ─────────────────────────────────────────────
    elif report_type == 'purchase':
        title = 'GST Purchase Register'
        q = filter_records_by_query_params(PurchaseEntry, sid, is_purchase=True)
        purchases = q.order_by(PurchaseEntry.entry_number.desc()).all()
        headers = ['#', 'Date', 'Entry No.', 'Invoice No.', 'Supplier Name', 'GSTIN',
                   'Place of Supply', 'HSN Code', 'GST Rate %',
                   'Taxable Amt', 'CGST', 'SGST', 'IGST', 'Total GST', 'Grand Total', 'Type']
        rows = []
        for i, p in enumerate(purchases, 1):
            gstin = (p.supplier_gstin or '').strip()
            doc_type = 'B2B' if (gstin and len(gstin) == 15) else 'B2C'
            rows.append([
                i, p.entry_date or '', p.entry_number or '', p.party_number or '',
                p.supplier_name or '', gstin or 'Unregistered',
                p.place_of_supply or '', p.hsn_code or '',
                float(p.gst_rate or 0),
                float(p.taxable_amount or 0), float(p.cgst_amount or 0),
                float(p.sgst_amount or 0), float(p.igst_amount or 0),
                float(p.total_gst or 0),
                float(p.grand_total or p.net_amount or 0), doc_type,
            ])
        return title, headers, rows, {9, 10, 11, 12, 13, 14}

    # ── GSTR-1 ────────────────────────────────────────────────────────
    elif report_type == 'gstr1':
        title = 'GSTR-1 (Outward Supplies)'
        shop = Shop.query.get(sid)
        q = filter_records_by_query_params(Bill, sid, is_purchase=False)
        bills = q.filter(Bill.status != 'returned').order_by(Bill.bill_number.asc()).all()
        headers = ['#', 'Doc Type', 'Invoice No.', 'Date', 'Customer Name', 'GSTIN',
                   'Place of Supply', 'Taxable Amt', 'CGST', 'SGST', 'IGST',
                   'Total GST', 'Invoice Value']
        rows = []
        for i, b in enumerate(bills, 1):
            gstin = (b.customer_gstin or '').strip()
            doc_type = 'B2B' if (gstin and len(gstin) == 15) else 'B2C'
            date_str = b.custom_date or (b.bill_date.strftime('%Y-%m-%d') if b.bill_date else '')
            rows.append([
                i, doc_type, b.bill_number or '', date_str,
                b.customer_name or '', gstin or 'Unregistered',
                b.place_of_supply or '',
                float(b.taxable_amount or 0), float(b.cgst_amount or 0),
                float(b.sgst_amount or 0), float(b.igst_amount or 0),
                float(b.total_gst or 0),
                float(b.grand_total or b.total_amount or 0),
            ])
        return title, headers, rows, {7, 8, 9, 10, 11, 12}

    # ── GSTR-2 ────────────────────────────────────────────────────────
    elif report_type == 'gstr2':
        title = 'GSTR-2 (Inward Supplies)'
        q = filter_records_by_query_params(PurchaseEntry, sid, is_purchase=True)
        purchases = q.order_by(PurchaseEntry.entry_number.asc()).all()
        headers = ['#', 'Doc Type', 'Entry No.', 'Invoice No.', 'Date',
                   'Supplier Name', 'GSTIN', 'Place of Supply',
                   'Taxable Amt', 'CGST', 'SGST', 'IGST', 'Total GST', 'Invoice Value']
        rows = []
        for i, p in enumerate(purchases, 1):
            gstin = (p.supplier_gstin or '').strip()
            doc_type = 'B2B' if (gstin and len(gstin) == 15) else 'B2C'
            rows.append([
                i, doc_type, p.entry_number or '', p.party_number or '',
                p.entry_date or '', p.supplier_name or '',
                gstin or 'Unregistered', p.place_of_supply or '',
                float(p.taxable_amount or 0), float(p.cgst_amount or 0),
                float(p.sgst_amount or 0), float(p.igst_amount or 0),
                float(p.total_gst or 0),
                float(p.grand_total or p.net_amount or 0),
            ])
        return title, headers, rows, {8, 9, 10, 11, 12, 13}

    # ── GSTR-3B ───────────────────────────────────────────────────────
    elif report_type == 'gstr3b':
        title = 'GSTR-3B (Net Tax Liability)'
        sales_q = filter_records_by_query_params(Bill, sid, is_purchase=False).filter(Bill.status != 'returned')
        pur_q   = filter_records_by_query_params(PurchaseEntry, sid, is_purchase=True)
        sales     = sales_q.all()
        purchases = pur_q.all()
        out_taxable = round(sum(float(b.taxable_amount or 0) for b in sales), 2)
        out_cgst    = round(sum(float(b.cgst_amount or 0)   for b in sales), 2)
        out_sgst    = round(sum(float(b.sgst_amount or 0)   for b in sales), 2)
        out_igst    = round(sum(float(b.igst_amount or 0)   for b in sales), 2)
        out_total   = round(out_cgst + out_sgst + out_igst, 2)
        in_taxable  = round(sum(float(p.taxable_amount or 0) for p in purchases), 2)
        in_cgst     = round(sum(float(p.cgst_amount or 0)   for p in purchases), 2)
        in_sgst     = round(sum(float(p.sgst_amount or 0)   for p in purchases), 2)
        in_igst     = round(sum(float(p.igst_amount or 0)   for p in purchases), 2)
        in_total    = round(in_cgst + in_sgst + in_igst, 2)
        net_cgst    = round(out_cgst - in_cgst, 2)
        net_sgst    = round(out_sgst - in_sgst, 2)
        net_igst    = round(out_igst - in_igst, 2)
        net_pay     = round(max(0, net_cgst) + max(0, net_sgst) + max(0, net_igst), 2)
        headers = ['Section', 'Description', 'Taxable Amount', 'CGST', 'SGST', 'IGST', 'Total GST']
        rows = [
            ['3.1', 'Outward Taxable Supplies (Sales)',  out_taxable, out_cgst, out_sgst, out_igst, out_total],
            ['4',   'Eligible ITC (Inward Supplies)',    in_taxable,  in_cgst,  in_sgst,  in_igst,  in_total],
            ['NET', 'Net Tax Payable (3.1 minus 4)',     round(out_taxable - in_taxable, 2), net_cgst, net_sgst, net_igst, net_pay],
        ]
        return title, headers, rows, {2, 3, 4, 5, 6}

    # ── HSN Summary (Sales) ───────────────────────────────────────────
    elif report_type == 'hsn':
        title = 'HSN/SAC Summary – Sales'
        sales_q = filter_records_by_query_params(Bill, sid, is_purchase=False).filter(Bill.status != 'returned')
        bills   = sales_q.all()
        shop    = Shop.query.get(sid)
        sc      = (shop.gst_number or '')[:2] if shop else ''
        hsn_map = {}
        for b in bills:
            items = _json.loads(b.items_json or '[]')
            is_inter = bool(b.place_of_supply and sc and b.place_of_supply[:2] != sc)
            for it in items:
                hsn     = (it.get('hsn_code') or it.get('hsn') or 'N/A').strip()
                qty     = float(it.get('qty', 0))
                gst_pct = float(it.get('gst', 0) or it.get('gst_pct', 0))
                amount  = float(it.get('amount', qty * float(it.get('price', 0))))
                tf      = 1 + gst_pct / 100
                taxable = amount / tf if tf else amount
                gst_amt = amount - taxable
                key     = (hsn, gst_pct)
                if key not in hsn_map:
                    hsn_map[key] = {'hsn': hsn, 'desc': it.get('name', ''), 'rate': gst_pct,
                                    'qty': 0.0, 'taxable': 0.0, 'cgst': 0.0, 'sgst': 0.0,
                                    'igst': 0.0, 'gst': 0.0, 'total': 0.0, 'ids': set()}
                hsn_map[key]['qty']     += qty
                hsn_map[key]['taxable'] += taxable
                hsn_map[key]['gst']     += gst_amt
                hsn_map[key]['total']   += amount
                hsn_map[key]['ids'].add(b.id)
                if is_inter: hsn_map[key]['igst'] += gst_amt
                else:
                    hsn_map[key]['cgst'] += gst_amt / 2
                    hsn_map[key]['sgst'] += gst_amt / 2
        headers = ['#', 'HSN Code', 'Description', 'GST Rate %', 'UQC', 'Total Qty',
                   'Total Value', 'Taxable Value', 'CGST', 'SGST', 'IGST', 'Total GST', 'Invoice Count']
        rows = []
        for i, h in enumerate(sorted(hsn_map.values(), key=lambda x: (x['hsn'], x['rate'])), 1):
            rows.append([i, h['hsn'], h['desc'], h['rate'], 'NOS',
                         round(h['qty'], 3), round(h['total'], 2), round(h['taxable'], 2),
                         round(h['cgst'], 2), round(h['sgst'], 2), round(h['igst'], 2),
                         round(h['gst'],  2), len(h['ids'])])
        return title, headers, rows, {6, 7, 8, 9, 10, 11}

    # ── Purchase HSN Summary ──────────────────────────────────────────
    elif report_type == 'purchase_hsn':
        title = 'Purchase HSN/SAC Summary – ITC'
        shop  = Shop.query.get(sid)
        sc    = (shop.gst_number or '')[:2] if shop else ''
        q     = filter_records_by_query_params(PurchaseEntry, sid, is_purchase=True)
        purch = q.all()
        hsn_map = {}
        for p in purch:
            items = _json.loads(p.items_json or '[]')
            is_inter = bool(p.place_of_supply and sc and p.place_of_supply[:2] != sc)
            for it in items:
                hsn     = (it.get('hsn_code') or it.get('hsn') or 'N/A').strip()
                qty     = float(it.get('qty', 0))
                gst_pct = float(it.get('gst', 0) or it.get('gst_pct', 0))
                amount  = float(it.get('amount', qty * float(it.get('price', 0))))
                tf      = 1 + gst_pct / 100
                taxable = amount / tf if tf else amount
                gst_amt = amount - taxable
                key     = (hsn, gst_pct)
                if key not in hsn_map:
                    hsn_map[key] = {'hsn': hsn, 'desc': it.get('name', ''), 'rate': gst_pct,
                                    'qty': 0.0, 'taxable': 0.0, 'cgst': 0.0, 'sgst': 0.0,
                                    'igst': 0.0, 'gst': 0.0, 'total': 0.0, 'ids': set()}
                hsn_map[key]['qty']     += qty
                hsn_map[key]['taxable'] += taxable
                hsn_map[key]['gst']     += gst_amt
                hsn_map[key]['total']   += amount
                hsn_map[key]['ids'].add(p.id)
                if is_inter: hsn_map[key]['igst'] += gst_amt
                else:
                    hsn_map[key]['cgst'] += gst_amt / 2
                    hsn_map[key]['sgst'] += gst_amt / 2
        headers = ['#', 'HSN Code', 'Description', 'GST Rate %', 'UQC', 'Total Qty',
                   'Total Value', 'Taxable Value', 'CGST', 'SGST', 'IGST', 'Total GST', 'Invoice Count']
        rows = []
        for i, h in enumerate(sorted(hsn_map.values(), key=lambda x: (x['hsn'], x['rate'])), 1):
            rows.append([i, h['hsn'], h['desc'], h['rate'], 'NOS',
                         round(h['qty'], 3), round(h['total'], 2), round(h['taxable'], 2),
                         round(h['cgst'], 2), round(h['sgst'], 2), round(h['igst'], 2),
                         round(h['gst'],  2), len(h['ids'])])
        return title, headers, rows, {6, 7, 8, 9, 10, 11}

    # ── GST Summary ───────────────────────────────────────────────────
    elif report_type == 'summary':
        title = 'GST Summary Report'
        sales = filter_records_by_query_params(Bill, sid, is_purchase=False).filter(Bill.status != 'returned').all()
        purch = filter_records_by_query_params(PurchaseEntry, sid, is_purchase=True).all()
        sc = round(sum(float(b.cgst_amount or 0) for b in sales), 2)
        ss = round(sum(float(b.sgst_amount or 0) for b in sales), 2)
        si = round(sum(float(b.igst_amount or 0) for b in sales), 2)
        st = round(sum(float(b.total_gst   or 0) for b in sales), 2)
        pc = round(sum(float(p.cgst_amount or 0) for p in purch), 2)
        ps = round(sum(float(p.sgst_amount or 0) for p in purch), 2)
        pi = round(sum(float(p.igst_amount or 0) for p in purch), 2)
        pt = round(sum(float(p.total_gst   or 0) for p in purch), 2)
        headers = ['Description', 'CGST', 'SGST', 'IGST', 'Total GST']
        rows = [
            ['GST Collected on Sales',           sc, ss, si, st],
            ['GST Paid on Purchases (ITC)',       pc, ps, pi, pt],
            ['Net GST Payable (Sales – ITC)',     round(sc-pc,2), round(ss-ps,2), round(si-pi,2), round(st-pt,2)],
        ]
        return title, headers, rows, {1, 2, 3, 4}

    # ── Monthly GST Report ────────────────────────────────────────────
    elif report_type == 'monthly':
        title = 'Monthly GST Report'
        sales = filter_records_by_query_params(Bill, sid, is_purchase=False).filter(Bill.status != 'returned').all()
        purch = filter_records_by_query_params(PurchaseEntry, sid, is_purchase=True).all()
        md = {}
        for b in sales:
            m = b.custom_date[:7] if b.custom_date else (b.bill_date.strftime('%Y-%m') if b.bill_date else '')
            if not m: continue
            if m not in md:
                md[m] = {'m': m, 'st': 0, 'sc': 0, 'ss': 0, 'si': 0,
                         'pt': 0, 'pc': 0, 'ps': 0, 'pi': 0}
            md[m]['st'] += float(b.taxable_amount or 0)
            md[m]['sc'] += float(b.cgst_amount   or 0)
            md[m]['ss'] += float(b.sgst_amount   or 0)
            md[m]['si'] += float(b.igst_amount   or 0)
        for p in purch:
            m = p.entry_date[:7] if p.entry_date else ''
            if not m: continue
            if m not in md:
                md[m] = {'m': m, 'st': 0, 'sc': 0, 'ss': 0, 'si': 0,
                         'pt': 0, 'pc': 0, 'ps': 0, 'pi': 0}
            md[m]['pt'] += float(p.taxable_amount or 0)
            md[m]['pc'] += float(p.cgst_amount   or 0)
            md[m]['ps'] += float(p.sgst_amount   or 0)
            md[m]['pi'] += float(p.igst_amount   or 0)
        headers = ['#', 'Month', 'Sales Taxable', 'Sales CGST', 'Sales SGST', 'Sales IGST',
                   'Purchase Taxable', 'Purchase CGST', 'Purchase SGST', 'Purchase IGST', 'Net Payable']
        rows = []
        for i, row in enumerate(sorted(md.values(), key=lambda x: x['m']), 1):
            sc, ss, si = round(row['sc'],2), round(row['ss'],2), round(row['si'],2)
            pc, ps, pi = round(row['pc'],2), round(row['ps'],2), round(row['pi'],2)
            net = round((sc-pc)+(ss-ps)+(si-pi), 2)
            rows.append([i, row['m'], round(row['st'],2), sc, ss, si,
                         round(row['pt'],2), pc, ps, pi, net])
        return title, headers, rows, {2, 3, 4, 5, 6, 7, 8, 9, 10}

    return 'Unknown Report', [], [], set()


# ══════════════════════════════════════════════════════════════════════
#  PROFESSIONAL EXCEL EXPORT
# ══════════════════════════════════════════════════════════════════════
@app.route('/api/gst/export/excel', methods=['GET'])
@jwt_required()
def gst_export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    import io as _io

    sid         = int(get_jwt_identity())
    report_type = request.args.get('report_type', 'sales')
    meta        = _get_export_meta(sid)
    shop        = meta['shop']
    fy          = meta['fy']

    title, headers, rows, currency_cols = _get_report_data(report_type, sid)

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    # ── Styles ─────────────────────────────────────────────────────────
    HDR_COLOR  = 'FF1F4E79'   # dark navy blue
    TOT_COLOR  = 'FFFFF2CC'   # light yellow
    INFO_COLOR = 'FFEBF3FB'   # pale blue
    ALT_COLOR  = 'FFF5F9FF'   # very pale blue (alt row)
    thin    = Side(style='thin',   color='BFBFBF')
    medium  = Side(style='medium', color='1F4E79')
    thin_b  = Border(left=thin, right=thin, top=thin, bottom=thin)

    num_cols = len(headers)
    col_last = get_column_letter(num_cols)

    # ── Row 1 — Logo + Shop Name ────────────────────────────────────────
    ws.row_dimensions[1].height = 44
    logo_path = os.path.join(_BUNDLE, 'static', 'logo.png')
    if os.path.exists(logo_path):
        try:
            xl_img = XLImage(logo_path)
            xl_img.width  = 130
            xl_img.height = 52
            ws.add_image(xl_img, 'A1')
        except Exception:
            pass
    shop_name = shop.shop_name if shop else 'Sanjana Pro'
    c = ws['C1']
    c.value     = shop_name
    c.font      = Font(name='Calibri', bold=True, size=18, color='1F4E79')
    c.alignment = Alignment(vertical='center')
    ws.merge_cells(f'C1:{col_last}1')

    # ── Row 2 — Address ─────────────────────────────────────────────────
    ws.row_dimensions[2].height = 16
    addr = (shop.address or '') if shop else ''
    c = ws['C2']
    c.value = f'Address: {addr}'
    c.font  = Font(name='Calibri', size=10, color='444444')
    ws.merge_cells(f'C2:{col_last}2')

    # ── Row 3 — GSTIN + Phone ───────────────────────────────────────────
    ws.row_dimensions[3].height = 16
    gstin = (shop.gst_number or 'N/A') if shop else 'N/A'
    phone = (shop.phone or '')         if shop else ''
    c = ws['C3']
    c.value = f'GSTIN: {gstin}    |    Phone: {phone}'
    c.font  = Font(name='Calibri', size=10, color='444444')
    ws.merge_cells(f'C3:{col_last}3')

    # ── Row 4 — Spacer ──────────────────────────────────────────────────
    ws.row_dimensions[4].height = 6

    # ── Row 5 — Report Title ────────────────────────────────────────────
    ws.row_dimensions[5].height = 24
    ws.merge_cells(f'A5:{col_last}5')
    c = ws['A5']
    c.value     = title
    c.font      = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
    c.fill      = PatternFill('solid', fgColor=HDR_COLOR)
    c.alignment = Alignment(horizontal='center', vertical='center')

    # ── Row 6 — FY + Date range ─────────────────────────────────────────
    ws.row_dimensions[6].height = 16
    fy_name = fy.fy_name if fy else 'All Years'
    from_d  = meta['from_date'] or 'All'
    to_d    = meta['to_date']   or 'All'
    ws.merge_cells(f'A6:{col_last}6')
    c = ws['A6']
    c.value     = f'Financial Year: {fy_name}    |    Period: {from_d}  to  {to_d}'
    c.font      = Font(name='Calibri', size=10, color='1F4E79')
    c.fill      = PatternFill('solid', fgColor=INFO_COLOR)
    c.alignment = Alignment(horizontal='center')

    # ── Row 7 — Generated timestamp ─────────────────────────────────────
    ws.row_dimensions[7].height = 13
    ws.merge_cells(f'A7:{col_last}7')
    c = ws['A7']
    c.value     = f'Generated: {meta["generated_at"]}    |    Sanjana Pro – Medical Shop Management'
    c.font      = Font(name='Calibri', size=9, italic=True, color='666666')
    c.alignment = Alignment(horizontal='right')

    # ── Row 8 — Spacer ──────────────────────────────────────────────────
    ws.row_dimensions[8].height = 6

    # ── Row 9 — Column Headers ──────────────────────────────────────────
    HEADER_ROW = 9
    ws.row_dimensions[HEADER_ROW].height = 22
    for ci, hdr in enumerate(headers, 1):
        cell = ws.cell(row=HEADER_ROW, column=ci, value=hdr)
        cell.font      = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
        cell.fill      = PatternFill('solid', fgColor=HDR_COLOR)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = thin_b

    ws.freeze_panes = f'A{HEADER_ROW + 1}'

    # ── Data Rows ────────────────────────────────────────────────────────
    INR_FORMAT = '[$₹-hi-IN]#,##0.00'
    totals = [0.0] * len(headers)

    for ri, row_data in enumerate(rows):
        xl_row = HEADER_ROW + 1 + ri
        ws.row_dimensions[xl_row].height = 16
        for ci, val in enumerate(row_data):
            cell       = ws.cell(row=xl_row, column=ci + 1, value=val)
            cell.border = thin_b
            cell.font   = Font(name='Calibri', size=9)
            if ci in currency_cols:
                cell.number_format = INR_FORMAT
                cell.alignment     = Alignment(horizontal='right')
                if isinstance(val, (int, float)):
                    totals[ci] += val
            else:
                cell.alignment = Alignment(horizontal='left')
            if ri % 2 == 1:
                cell.fill = PatternFill('solid', fgColor=ALT_COLOR)

    # ── Grand Totals Row ─────────────────────────────────────────────────
    if rows and currency_cols:
        tot_xl = HEADER_ROW + 1 + len(rows)
        ws.row_dimensions[tot_xl].height = 20
        lbl = ws.cell(row=tot_xl, column=1, value='GRAND TOTAL')
        lbl.font      = Font(name='Calibri', bold=True, size=10, color='1F4E79')
        lbl.fill      = PatternFill('solid', fgColor=TOT_COLOR)
        lbl.border    = thin_b
        lbl.alignment = Alignment(horizontal='center')
        for ci in range(1, len(headers)):
            cell        = ws.cell(row=tot_xl, column=ci + 1)
            cell.fill   = PatternFill('solid', fgColor=TOT_COLOR)
            cell.border = thin_b
            cell.font   = Font(name='Calibri', bold=True, size=10)
            if ci in currency_cols:
                cell.value          = round(totals[ci], 2)
                cell.number_format  = INR_FORMAT
                cell.alignment      = Alignment(horizontal='right')

    # ── Auto-size Columns ────────────────────────────────────────────────
    for ci_idx, col_cells in enumerate(ws.columns):
        max_len   = 0
        col_letter = get_column_letter(ci_idx + 1)
        for cell in col_cells:
            try:
                val_str = str(cell.value or '')
                max_len = max(max_len, len(val_str))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 45)

    # ── Page Setup ───────────────────────────────────────────────────────
    ws.page_setup.orientation  = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize    = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage    = True
    ws.page_setup.fitToWidth   = 1
    ws.page_setup.fitToHeight  = 0
    ws.print_options.horizontalCentered = True
    ws.oddHeader.center.text   = f'&B{title}'
    ws.oddFooter.left.text     = f'Generated by Sanjana Pro  |  {meta["generated_at"]}'
    ws.oddFooter.center.text   = 'Page &P of &N'
    ws.oddFooter.right.text    = shop_name

    # ── Stream ───────────────────────────────────────────────────────────
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_title = title.replace(' ', '_').replace('/', '_').replace('–', '-')
    fname      = f'{safe_title}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    from flask import send_file as _send_file
    return _send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=fname,
    )


# ══════════════════════════════════════════════════════════════════════
#  PROFESSIONAL PDF EXPORT
# ══════════════════════════════════════════════════════════════════════
@app.route('/api/gst/export/pdf', methods=['GET'])
@jwt_required()
def gst_export_pdf():
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer)
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    import io as _io

    sid         = int(get_jwt_identity())
    report_type = request.args.get('report_type', 'sales')
    meta        = _get_export_meta(sid)
    shop        = meta['shop']
    fy          = meta['fy']

    title, headers, rows, currency_cols = _get_report_data(report_type, sid)

    PAGE_SIZE = landscape(A4)
    PAGE_W, PAGE_H = PAGE_SIZE
    MARGIN = 1.5 * cm

    buf = _io.BytesIO()

    # ── Page decorators ──────────────────────────────────────────────────
    def _draw_page(canvas, doc):
        canvas.saveState()
        W, H = PAGE_SIZE

        # Header bar
        canvas.setFillColorRGB(0.122, 0.306, 0.475)   # #1F4E79
        canvas.rect(0, H - 72, W, 72, fill=1, stroke=0)

        # Logo
        logo_path = os.path.join(_BUNDLE, 'static', 'logo.png')
        if os.path.exists(logo_path):
            try:
                canvas.drawImage(logo_path, MARGIN, H - 67,
                                  width=90, height=54,
                                  preserveAspectRatio=True, mask='auto')
            except Exception:
                pass

        # Shop Name
        canvas.setFillColorRGB(1, 1, 1)
        canvas.setFont('Helvetica-Bold', 13)
        sname = shop.shop_name if shop else 'Sanjana Pro'
        canvas.drawString(MARGIN + 100, H - 25, sname)

        # GSTIN / Phone
        canvas.setFont('Helvetica', 8.5)
        sgstin = (shop.gst_number or 'N/A') if shop else 'N/A'
        sphone = (shop.phone or '')          if shop else ''
        canvas.drawString(MARGIN + 100, H - 40, f'GSTIN: {sgstin}    Phone: {sphone}')

        # Address
        saddr = (shop.address or '') if shop else ''
        canvas.setFont('Helvetica', 8)
        canvas.drawString(MARGIN + 100, H - 53, saddr[:90])

        # Report title — centered
        canvas.setFont('Helvetica-Bold', 15)
        canvas.drawCentredString(W / 2, H - 25, title)

        # FY + date range — right
        fy_name = fy.fy_name if fy else 'All Years'
        from_d  = meta['from_date'] or 'All'
        to_d    = meta['to_date']   or 'All'
        canvas.setFont('Helvetica', 8)
        canvas.drawRightString(W - MARGIN, H - 22, f'FY: {fy_name}')
        canvas.drawRightString(W - MARGIN, H - 35, f'Period: {from_d}  to  {to_d}')
        canvas.drawRightString(W - MARGIN, H - 48, f'Generated: {meta["generated_at"]}')

        # Footer bar
        canvas.setFillColorRGB(0.122, 0.306, 0.475)
        canvas.rect(0, 0, W, 22, fill=1, stroke=0)
        canvas.setFillColorRGB(1, 1, 1)
        canvas.setFont('Helvetica', 7.5)
        canvas.drawString(MARGIN, 7, 'Confidential – For Internal Use Only')
        canvas.drawCentredString(W / 2, 7, f'Page {doc.page}')
        canvas.drawRightString(W - MARGIN, 7, 'Generated by Sanjana Pro')

        canvas.restoreState()

    doc = SimpleDocTemplate(
        buf,
        pagesize=PAGE_SIZE,
        leftMargin=MARGIN, rightMargin=MARGIN,
        topMargin=82,      bottomMargin=30,
    )

    # ── Paragraph styles ─────────────────────────────────────────────────
    s_hdr  = ParagraphStyle('Hdr',  fontName='Helvetica-Bold', fontSize=7.5,
                             alignment=TA_CENTER, leading=9)
    s_left = ParagraphStyle('Left', fontName='Helvetica',      fontSize=7.5,
                             alignment=TA_LEFT,   leading=9)
    s_right= ParagraphStyle('Right',fontName='Helvetica',      fontSize=7.5,
                             alignment=TA_RIGHT,  leading=9)
    s_ctr  = ParagraphStyle('Ctr',  fontName='Helvetica',      fontSize=7.5,
                             alignment=TA_CENTER, leading=9)
    s_tot  = ParagraphStyle('Tot',  fontName='Helvetica-Bold', fontSize=7.5,
                             alignment=TA_RIGHT,  leading=9)
    s_totl = ParagraphStyle('TotL', fontName='Helvetica-Bold', fontSize=7.5,
                             alignment=TA_CENTER, leading=9)

    # ── Table data ───────────────────────────────────────────────────────
    INR = lambda v: f'\u20B9{v:,.2f}' if isinstance(v, (int, float)) else str(v)

    tbl_headers = [Paragraph(f'<b>{h}</b>', s_hdr) for h in headers]
    tbl_data    = [tbl_headers]
    totals      = [0.0] * len(headers)

    for row_data in rows:
        tbl_row = []
        for ci, val in enumerate(row_data):
            if ci in currency_cols and isinstance(val, (int, float)):
                totals[ci] += val
                tbl_row.append(Paragraph(INR(val), s_right))
            elif ci == 0:
                tbl_row.append(Paragraph(str(val), s_ctr))
            else:
                tbl_row.append(Paragraph(str(val), s_left))
        tbl_data.append(tbl_row)

    # Grand totals row
    if rows and currency_cols:
        tot_row = []
        for ci in range(len(headers)):
            if ci == 0:
                tot_row.append(Paragraph('<b>TOTAL</b>', s_totl))
            elif ci in currency_cols:
                tot_row.append(Paragraph(f'<b>{INR(round(totals[ci], 2))}</b>', s_tot))
            else:
                tot_row.append(Paragraph('', s_left))
        tbl_data.append(tot_row)

    # ── Column widths ─────────────────────────────────────────────────────
    usable_w = PAGE_W - 2 * MARGIN
    col_w    = []
    for hdr in headers:
        h_low = hdr.lower()
        if hdr == '#':                   col_w.append(0.55 * cm)
        elif 'name' in h_low or 'desc' in h_low: col_w.append(3.4 * cm)
        elif 'gstin' in h_low:           col_w.append(3.0 * cm)
        elif 'date'  in h_low or 'month' in h_low: col_w.append(1.9 * cm)
        elif 'no.'   in h_low or 'number' in h_low: col_w.append(2.0 * cm)
        elif hdr in ('Type', 'Doc Type', 'UQC', 'Section'): col_w.append(1.4 * cm)
        elif 'supply' in h_low:          col_w.append(1.8 * cm)
        else:                            col_w.append(2.3 * cm)
    total_w = sum(col_w)
    if total_w > usable_w:
        scale = usable_w / total_w
        col_w = [w * scale for w in col_w]

    # ── Table style ───────────────────────────────────────────────────────
    HDR_C = colors.HexColor('#1F4E79')
    ALT_C = colors.HexColor('#EBF3FB')
    TOT_C = colors.HexColor('#FFF2CC')
    BDR_C = colors.HexColor('#BFBFBF')

    n_data = len(tbl_data)   # includes header + data + totals

    style_cmds = [
        # Header row
        ('BACKGROUND',   (0, 0), (-1, 0), HDR_C),
        ('TEXTCOLOR',    (0, 0), (-1, 0), colors.white),
        ('FONTNAME',     (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',     (0, 0), (-1, 0), 7.5),
        # Grid
        ('GRID',         (0, 0), (-1, -1), 0.4, BDR_C),
        ('BOX',          (0, 0), (-1, -1), 0.9, HDR_C),
        # Alignment
        ('VALIGN',       (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING',   (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING',(0, 0), (-1, -1), 3),
        ('LEFTPADDING',  (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
    ]
    # Totals row (last)
    if rows and currency_cols:
        style_cmds += [
            ('BACKGROUND', (0, -1), (-1, -1), TOT_C),
            ('FONTNAME',   (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('LINEABOVE',  (0, -1), (-1, -1), 1.0, HDR_C),
        ]
    # Alternating row colors (only data rows, not header or totals)
    for ri in range(1, n_data - (1 if rows and currency_cols else 0)):
        if ri % 2 == 0:
            style_cmds.append(('BACKGROUND', (0, ri), (-1, ri), ALT_C))

    tbl = Table(tbl_data, colWidths=col_w, repeatRows=1, splitByRow=1)
    tbl.setStyle(TableStyle(style_cmds))

    doc.build([tbl], onFirstPage=_draw_page, onLaterPages=_draw_page)

    buf.seek(0)
    safe_title = title.replace(' ', '_').replace('/', '_').replace('–', '-')
    fname      = f'{safe_title}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    from flask import send_file as _send_file
    return _send_file(buf, mimetype='application/pdf',
                      as_attachment=True, download_name=fname)



# ══════════════════════════════════════════════════════════════════════
#  STOCK INVENTORY EXCEL EXPORT
# ══════════════════════════════════════════════════════════════════════
@app.route('/api/export/stock/excel', methods=['GET'])
@jwt_required()
def export_stock_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    import io as _io

    sid  = int(get_jwt_identity())
    shop = Shop.query.get(sid)

    # Fetch all stock for this shop
    medicines = Medicine.query.filter_by(shop_id=sid).order_by(Medicine.name).all()

    HDR_COLOR = 'FF1F4E79'
    TOT_COLOR = 'FFFFF2CC'
    ALT_COLOR = 'FFF5F9FF'
    thin  = Side(style='thin', color='BFBFBF')
    thin_b = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ['#', 'Medicine Name', 'Category', 'Batch', 'Qty', 'Pack Size',
               'Price (₹)', 'MRP (₹)', 'GST %', 'Expiry Date', 'Company', 'Supplier']
    currency_cols = {6, 7}
    num_cols  = len(headers)
    col_last  = get_column_letter(num_cols)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Stock Inventory'

    shop_name = shop.shop_name if shop else 'Sanjana Pro'
    gen_at    = datetime.now().strftime('%d-%m-%Y  %H:%M:%S')

    # ── Header rows ────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 44
    logo_path = os.path.join(_BUNDLE, 'static', 'logo.png')
    if os.path.exists(logo_path):
        try:
            xl_img = XLImage(logo_path)
            xl_img.width = 130; xl_img.height = 52
            ws.add_image(xl_img, 'A1')
        except Exception:
            pass

    c = ws['C1']
    c.value = shop_name
    c.font  = Font(name='Calibri', bold=True, size=18, color='1F4E79')
    c.alignment = Alignment(vertical='center')
    ws.merge_cells(f'C1:{col_last}1')

    ws.row_dimensions[2].height = 14
    c = ws['C2']
    c.value = f'Address: {shop.address or ""}' if shop else ''
    c.font  = Font(name='Calibri', size=10, color='444444')
    ws.merge_cells(f'C2:{col_last}2')

    ws.row_dimensions[3].height = 14
    c = ws['C3']
    c.value = f'GSTIN: {shop.gst_number or "N/A"}    |    Phone: {shop.phone or ""}' if shop else ''
    c.font  = Font(name='Calibri', size=10, color='444444')
    ws.merge_cells(f'C3:{col_last}3')

    ws.row_dimensions[4].height = 6

    ws.row_dimensions[5].height = 24
    ws.merge_cells(f'A5:{col_last}5')
    c = ws['A5']
    c.value = 'Stock Inventory Report'
    c.font  = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
    c.fill  = PatternFill('solid', fgColor=HDR_COLOR)
    c.alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[6].height = 14
    ws.merge_cells(f'A6:{col_last}6')
    c = ws['A6']
    c.value = f'Total Items: {len(medicines)}    |    Generated: {gen_at}'
    c.font  = Font(name='Calibri', size=10, italic=True, color='666666')
    c.alignment = Alignment(horizontal='right')

    ws.row_dimensions[7].height = 6

    # ── Column header row ──────────────────────────────────────────────
    HEADER_ROW = 8
    ws.row_dimensions[HEADER_ROW].height = 22
    for ci, hdr in enumerate(headers, 1):
        cell = ws.cell(row=HEADER_ROW, column=ci, value=hdr)
        cell.font      = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
        cell.fill      = PatternFill('solid', fgColor=HDR_COLOR)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = thin_b
    ws.freeze_panes = f'A{HEADER_ROW + 1}'

    # ── Data rows ──────────────────────────────────────────────────────
    INR_FORMAT = '[$₹-hi-IN]#,##0.00'
    totals = {'qty': 0.0, 'price': 0.0, 'mrp': 0.0}

    for ri, m in enumerate(medicines):
        xl_row = HEADER_ROW + 1 + ri
        ws.row_dimensions[xl_row].height = 16
        row_data = [
            ri + 1, m.name or '', m.category or '', m.batch or '',
            float(m.quantity or 0), int(m.pack_size or 10),
            float(m.price or 0), float(m.mrp or 0),
            float(m.gst or 0), m.expiry_date or '',
            m.company_name or '', m.supplier_name or '',
        ]
        totals['qty']   += float(m.quantity or 0)
        totals['price'] += float(m.price or 0)
        totals['mrp']   += float(m.mrp or 0)

        for ci, val in enumerate(row_data):
            cell = ws.cell(row=xl_row, column=ci + 1, value=val)
            cell.border = thin_b
            cell.font   = Font(name='Calibri', size=9)
            if ci in currency_cols:
                cell.number_format = INR_FORMAT
                cell.alignment     = Alignment(horizontal='right')
            elif ci == 4:   # Qty — right-align numbers
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left')
            if ri % 2 == 1:
                cell.fill = PatternFill('solid', fgColor=ALT_COLOR)

    # ── Totals row ─────────────────────────────────────────────────────
    tot_xl = HEADER_ROW + 1 + len(medicines)
    ws.row_dimensions[tot_xl].height = 20
    lbl = ws.cell(row=tot_xl, column=1, value='TOTALS')
    lbl.font = Font(name='Calibri', bold=True, size=10, color='1F4E79')
    lbl.fill = PatternFill('solid', fgColor=TOT_COLOR)
    lbl.border = thin_b
    lbl.alignment = Alignment(horizontal='center')
    for ci in range(1, num_cols):
        cell = ws.cell(row=tot_xl, column=ci + 1)
        cell.fill = PatternFill('solid', fgColor=TOT_COLOR)
        cell.border = thin_b
        cell.font   = Font(name='Calibri', bold=True, size=10)
        if ci == 4:   # Qty total
            cell.value = round(totals['qty'], 2)
            cell.alignment = Alignment(horizontal='right')
        elif ci == 6:  # Price total
            cell.value = round(totals['price'], 2)
            cell.number_format = INR_FORMAT
            cell.alignment = Alignment(horizontal='right')
        elif ci == 7:  # MRP total
            cell.value = round(totals['mrp'], 2)
            cell.number_format = INR_FORMAT
            cell.alignment = Alignment(horizontal='right')

    # ── Auto-size columns ──────────────────────────────────────────────
    for ci_idx, col_cells in enumerate(ws.columns):
        max_len = 0
        col_letter = get_column_letter(ci_idx + 1)
        for cell in col_cells:
            try:
                max_len = max(max_len, len(str(cell.value or '')))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 40)

    # ── Page setup ─────────────────────────────────────────────────────
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.print_options.horizontalCentered = True
    ws.oddFooter.left.text   = f'Generated by Sanjana Pro  |  {gen_at}'
    ws.oddFooter.center.text = 'Page &P of &N'
    ws.oddFooter.right.text  = shop_name

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f'Stock_Inventory_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    from flask import send_file as _send_file
    return _send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=fname,
    )


threading.Thread(target=_sync_to_cloud_admin, daemon=True).start()

if __name__ == '__main__':
    import socket, subprocess, os

    # Check if launched by Electron (Electron sets this env var).
    # If yes, skip opening a browser window — Electron handles the UI.
    _launched_by_electron = os.environ.get('ELECTRON_PARENT') == '1'

    def _open_app_window():
        """Poll until Flask is ready, then open as a standalone app window (NO browser tab)."""
        url = "http://127.0.0.1:5000"
        for _ in range(40):
            time.sleep(0.5)
            try:
                s = socket.create_connection(("127.0.0.1", 5000), timeout=1)
                s.close()
                print("[SANJANA] Server ready — opening app window...")

                # Try Microsoft Edge --app mode (frameless, no address bar)
                edge_paths = [
                    r'C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe',
                    r'C:\Program Files\Microsoft\Edge\Application\msedge.exe',
                ]
                for edge in edge_paths:
                    if os.path.exists(edge):
                        subprocess.Popen([edge, f'--app={url}', '--no-first-run'], creationflags=0x08000000)
                        return

                # Try Google Chrome --app mode
                chrome_paths = [
                    r'C:\Program Files\Google\Chrome\Application\chrome.exe',
                    r'C:\Program Files (x86)\Google\Chrome\Application\chrome.exe',
                ]
                for chrome in chrome_paths:
                    if os.path.exists(chrome):
                        subprocess.Popen([chrome, f'--app={url}', '--no-first-run'], creationflags=0x08000000)
                        return

                # No browser found — server is running, user can open manually
                print(f"[SANJANA] Open manually: {url}")
                return
            except OSError:
                pass
        print(f"[SANJANA] Server did not start in time. Try: {url}")
    # Only open a browser window when running standalone (not via Electron)
    if not _launched_by_electron:
        threading.Thread(target=_open_app_window, daemon=True).start()
    else:
        print("[SANJANA] Running as Electron backend — skipping browser window.")

    print("=" * 55)
    print("  SANJANA PRO — Medical Shop Management")
    print(f"  Data : {DATA_FOLDER}")
    print("  URL  : http://127.0.0.1:5000")
    if not _launched_by_electron:
        print("  App window opens automatically...")
    else:
        print("  Electron is handling the UI window.")
    print("  Close this window to stop.")
    print("=" * 55)

    app.run(host="0.0.0.0", port=5000, debug=False, use_reloader=False)